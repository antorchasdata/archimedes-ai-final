"""Tests for ReferenceCatalogResolver integration into write_leanix_excel.

Uses the ARCHIMEDES_USE_CATALOG_RESOLVER feature flag to gate resolver activation.
"""
from __future__ import annotations

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

sys.path.insert(0, str(Path(__file__).parent.parent))

import pytest

from pipeline.reference_catalog import ResolvedMatch
from pipeline import write as write_mod


def _minimal_enriched():
    """Two requirements → two distinct RSA apps."""
    return [
        {
            "id": "R1", "description": "x", "area": "",
            "module": "FIN", "bcs": ["Finance"], "rsa": "SAP S/4HANA",
            "coverage": "Total", "dev": "", "dev_exp": "",
            "ext_apps": "", "licensing": "", "comment": "",
        },
        {
            "id": "R2", "description": "y", "area": "",
            "module": "HR", "bcs": ["HR"], "rsa": "SAP SuccessFactors",
            "coverage": "Total", "dev": "", "dev_exp": "",
            "ext_apps": "", "licensing": "", "comment": "",
        },
    ]


def test_flag_off_resolver_not_called(tmp_path, monkeypatch):
    monkeypatch.delenv("ARCHIMEDES_USE_CATALOG_RESOLVER", raising=False)
    out = tmp_path / "x.xlsx"
    with patch.object(write_mod, "ReferenceCatalogResolver") as Resolver:
        write_mod.write_leanix_excel(
            enriched=_minimal_enriched(),
            bcs_index={},
            output_path=out,
            client_name="ACME",
        )
    Resolver.assert_not_called()
    assert out.exists()


def test_flag_on_resolver_called_and_externalid_written(tmp_path, monkeypatch):
    monkeypatch.setenv("ARCHIMEDES_USE_CATALOG_RESOLVER", "true")

    fake_resolver = MagicMock()
    fake_resolver.resolve.side_effect = lambda fs_type, names: {
        n: ResolvedMatch(
            name=n,
            external_id=f"lx_APP_{i:03d}" if fs_type == "Application" else f"lx_ITC_{i:03d}",
            status="LINKED",
            confidence="VERYHIGH",
            display_name=n,
            fields={},
        )
        for i, n in enumerate(names, start=1)
    }
    out = tmp_path / "x.xlsx"
    with patch.object(write_mod, "ReferenceCatalogResolver", return_value=fake_resolver), \
         patch.object(write_mod, "_resolver_credentials", return_value=("https://x", "tok")):
        write_mod.write_leanix_excel(
            enriched=_minimal_enriched(),
            bcs_index={},
            output_path=out,
            client_name="ACME",
        )

    # resolver.resolve called for both types
    assert fake_resolver.resolve.call_count == 2
    types_called = [c.args[0] for c in fake_resolver.resolve.call_args_list]
    assert "Application" in types_called
    assert "ITComponent" in types_called

    # Verify externalId landed in the Application sheet
    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["Application"]
    # Row 1 = technical keys
    headers = [c.value for c in ws[1]]
    ext_col = headers.index("externalId") + 1
    data_rows = [ws.cell(row=r, column=ext_col).value
                 for r in range(3, ws.max_row + 1)]
    assert any(v and v.startswith("lx_APP_") for v in data_rows)


def test_flag_on_cleanup_invoked(tmp_path, monkeypatch):
    monkeypatch.setenv("ARCHIMEDES_USE_CATALOG_RESOLVER", "1")
    fake_resolver = MagicMock()
    fake_resolver.resolve.return_value = {}
    out = tmp_path / "x.xlsx"
    with patch.object(write_mod, "ReferenceCatalogResolver", return_value=fake_resolver), \
         patch.object(write_mod, "_resolver_credentials", return_value=("https://x", "tok")):
        write_mod.write_leanix_excel(
            enriched=_minimal_enriched(),
            bcs_index={},
            output_path=out,
            client_name="ACME",
        )
    fake_resolver.cleanup.assert_called_once()


def test_flag_on_resolver_failure_is_non_fatal(tmp_path, monkeypatch):
    """If resolver construction explodes, the Excel must still be written."""
    monkeypatch.setenv("ARCHIMEDES_USE_CATALOG_RESOLVER", "true")
    out = tmp_path / "x.xlsx"
    with patch.object(write_mod, "ReferenceCatalogResolver", side_effect=RuntimeError("boom")), \
         patch.object(write_mod, "_resolver_credentials", return_value=("https://x", "tok")):
        write_mod.write_leanix_excel(
            enriched=_minimal_enriched(),
            bcs_index={},
            output_path=out,
            client_name="ACME",
        )
    assert out.exists()


def test_push_payload_includes_external_id_when_present():
    """When a row carries externalId, the createFactSheet payload includes it."""
    from pipeline.push_ldif import build_create_factsheet_payload

    row = {
        "name": "SAP S/4HANA",
        "type": "Application",
        "externalId": "lx_APP_000123",
        "description": "ERP",
    }
    payload = build_create_factsheet_payload(row)
    assert payload.get("externalId") == "lx_APP_000123"


def test_push_payload_omits_external_id_when_absent():
    from pipeline.push_ldif import build_create_factsheet_payload

    row = {"name": "Custom App", "type": "Application", "description": ""}
    payload = build_create_factsheet_payload(row)
    assert "externalId" not in payload or payload["externalId"] in (None, "")


def test_link_apps_skips_rows_with_external_id():
    """Post-push linker must not try to link rows that already carried externalId.

    The pre-creation Reference Catalog resolver populates externalId in the
    staging Excel. Once a row is created in LeanIX with that externalId, the
    fact sheet is already linked to the catalog — the post-push linker must
    skip it to avoid spurious work and conflicting links.
    """
    from pipeline.write import _link_apps_to_catalog

    app_id_cache = {"SAP S/4HANA": "ws-uuid-1", "Custom App": "ws-uuid-2"}
    rows_by_name = {
        "SAP S/4HANA": {"externalId": "lx_APP_000123"},
        "Custom App": {},  # no externalId — eligible for post-push linking
    }

    # Mock the bearer so the function proceeds past auth.
    # `requests` is imported locally inside the function (`import requests as _req`);
    # patch the module-level `requests` so the local import resolves to our mock.
    fake_requests = MagicMock()
    fake_requests.post.return_value = MagicMock(
        status_code=200,
        json=lambda: {"data": {}},
        raise_for_status=MagicMock(),
    )
    fake_requests.put.return_value = MagicMock(
        status_code=200,
        json=lambda: {},
        raise_for_status=MagicMock(),
    )
    with patch.object(write_mod, "get_bearer", return_value="fake-token"), \
         patch.dict("sys.modules", {"requests": fake_requests}):
        _link_apps_to_catalog(
            "https://x", "tok", app_id_cache, {},
            rows_by_name=rows_by_name,
        )

    # The skipped app's workspace UUID must never appear in any POST body.
    called_bodies = [
        str(c.kwargs.get("json", "")) + str(c.args)
        for c in fake_requests.post.call_args_list
    ]
    assert all("ws-uuid-1" not in body for body in called_bodies), \
        "Pre-linked app SAP S/4HANA must not be sent to batch-links"


def test_catalog_resolution_report_written(tmp_path, monkeypatch):
    """When the resolver runs, a catalog_resolution_report.json is written
    next to the Excel with one entry per resolved name per type."""
    monkeypatch.setenv("ARCHIMEDES_USE_CATALOG_RESOLVER", "true")

    fake_resolver = MagicMock()
    fake_resolver.resolve.side_effect = lambda fs_type, names: {
        n: ResolvedMatch(
            name=n,
            external_id=f"lx_APP_{i:03d}" if fs_type == "Application" else f"lx_ITC_{i:03d}",
            display_name=n, confidence="VERYHIGH", status="LINKED", fields={},
        )
        for i, n in enumerate(names, start=1)
    }
    out = tmp_path / "x.xlsx"
    with patch.object(write_mod, "ReferenceCatalogResolver", return_value=fake_resolver), \
         patch.object(write_mod, "_resolver_credentials", return_value=("https://x", "tok")):
        write_mod.write_leanix_excel(
            enriched=_minimal_enriched(),
            bcs_index={},
            output_path=out,
            client_name="ACME",
        )

    import json as _json
    report_path = tmp_path / "catalog_resolution_report.json"
    assert report_path.exists(), "catalog_resolution_report.json must be written next to the Excel"
    body = _json.loads(report_path.read_text())
    assert "Application" in body and "ITComponent" in body
    assert any(e.get("status") == "LINKED" for e in body["Application"])
    assert all(e.get("external_id", "").startswith("lx_APP_") for e in body["Application"])
