# tests/test_catalog_report_e2e.py
"""End-to-end: from raw inputs in a session dir, hit the endpoint and get a usable HTML + XLSX."""
import json
import uuid
from pathlib import Path

import openpyxl
from fastapi.testclient import TestClient

from archimedes_wizard import app, _sessions, OUTPUT_DIR


def test_e2e_session_to_report():
    session_id = str(uuid.uuid4())
    session_dir = OUTPUT_DIR / session_id
    session_dir.mkdir(parents=True, exist_ok=True)

    (session_dir / "catalog_resolution_report.json").write_text(json.dumps({"entries": [
        {"name": "SAP S/4HANA", "type": "Application", "status": "LINKED",
         "confidence": "VERYHIGH", "external_id": "lx_APP_019278",
         "suggested_name": "SAP S/4HANA Cloud", "suggested_score": 1.0},
        {"name": "SAP Ariba",   "type": "Application", "status": "CUSTOM",
         "confidence": "HIGH",
         "suggested_name": "SAP Ariba Buying", "suggested_score": 0.82},
        {"name": "Legacy ERP",  "type": "Application", "status": "CUSTOM",
         "confidence": "NONE"},
    ]}))
    (session_dir / "push_uuid_map.json").write_text(json.dumps({
        "workspace": "demo-eu-3",
        "base_url":  "https://demo-eu-3.leanix.net",
        "entries": {
            "Application::SAP S/4HANA": {"uuid": "uuid-s4", "created": True},
            "Application::SAP Ariba":   {"uuid": "uuid-ariba", "created": True},
            "Application::Legacy ERP":  {"uuid": "uuid-legacy", "created": True},
        },
        "failed": [],
    }))
    _sessions[session_id] = {
        "client_name": "Acme",
        "output_dir":  session_dir,
        "out_baseline": None, "out_target": None,
    }

    client = TestClient(app)

    # HTML
    resp = client.get(f"/api/session/{session_id}/catalog-report")
    assert resp.status_code == 200
    html = resp.text
    assert "1 linked" in html
    assert "1 to review" in html
    assert "1 custom" in html
    assert "factsheet/Application/uuid-s4" in html       # Open FS link for LINKED
    assert "factsheet/Application/uuid-ariba" in html    # Open FS link for REVIEW
    assert "factsheet/Application/uuid-legacy" in html   # Open FS link for CUSTOM
    assert "/inventory/referenceCatalog?q=SAP%20Ariba" in html  # Catalog search only on REVIEW
    assert html.count("/inventory/referenceCatalog?q=") == 1    # not on LINKED or CUSTOM

    # XLSX
    resp = client.get(f"/api/session/{session_id}/catalog-report.xlsx")
    assert resp.status_code == 200
    out = session_dir / "catalog_report.xlsx"
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    ws = wb["Catalog Report"]
    assert ws.max_row == 4   # header + 3 data rows
    # Header order
    assert [c.value for c in ws[1]] == [
        "Status", "Name", "Type", "Suggested match", "Suggested score",
        "externalId", "Open FS URL", "Search catalog URL",
    ]
