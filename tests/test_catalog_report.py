# tests/test_catalog_report.py
from pipeline.catalog_report import build_rows, sort_rows, count_rows

def test_build_rows_empty_inputs():
    rows = build_rows(
        resolution={"entries": []},
        uuid_map={"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net", "entries": {}, "failed": []},
    )
    assert rows == []


_UUID_MAP_EMPTY = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net", "entries": {}, "failed": []}

def test_linked_when_status_linked():
    res = {"entries": [{"name": "X", "type": "Application", "status": "LINKED", "confidence": "VERYHIGH", "external_id": "lx_APP_1"}]}
    rows = build_rows(res, _UUID_MAP_EMPTY)
    assert rows[0].status == "LINKED"

def test_review_when_high_confidence_custom():
    res = {"entries": [{"name": "X", "type": "Application", "status": "CUSTOM", "confidence": "HIGH"}]}
    rows = build_rows(res, _UUID_MAP_EMPTY)
    assert rows[0].status == "REVIEW"

def test_review_when_medium_confidence_custom():
    res = {"entries": [{"name": "X", "type": "Application", "status": "CUSTOM", "confidence": "MEDIUM"}]}
    rows = build_rows(res, _UUID_MAP_EMPTY)
    assert rows[0].status == "REVIEW"

def test_custom_when_low_confidence():
    res = {"entries": [{"name": "X", "type": "Application", "status": "CUSTOM", "confidence": "LOW"}]}
    rows = build_rows(res, _UUID_MAP_EMPTY)
    assert rows[0].status == "CUSTOM"

def test_custom_when_no_confidence():
    res = {"entries": [{"name": "X", "type": "Application", "status": "CUSTOM", "confidence": "NONE"}]}
    rows = build_rows(res, _UUID_MAP_EMPTY)
    assert rows[0].status == "CUSTOM"

def test_uuid_attached_when_present_in_map():
    res = {"entries": [{"name": "SAP S/4HANA", "type": "Application", "status": "LINKED", "confidence": "VERYHIGH"}]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {"Application::SAP S/4HANA": {"uuid": "abc123", "created": True}}, "failed": []}
    rows = build_rows(res, umap)
    assert rows[0].fs_uuid == "abc123"
    assert rows[0].push_failed is False

def test_uuid_missing_when_not_in_map():
    res = {"entries": [{"name": "Custom App", "type": "Application", "status": "CUSTOM", "confidence": "NONE"}]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net", "entries": {}, "failed": []}
    rows = build_rows(res, umap)
    assert rows[0].fs_uuid is None
    assert rows[0].push_failed is False

def test_push_failed_marked():
    res = {"entries": [{"name": "Broken App", "type": "Application", "status": "CUSTOM", "confidence": "NONE"}]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {}, "failed": [{"name": "Broken App", "type": "Application", "error": "boom"}]}
    rows = build_rows(res, umap)
    assert rows[0].fs_uuid is None
    assert rows[0].push_failed is True

def test_rows_sorted_linked_review_custom():
    res = {"entries": [
        {"name": "C", "type": "Application", "status": "CUSTOM",  "confidence": "NONE"},
        {"name": "L", "type": "Application", "status": "LINKED",  "confidence": "VERYHIGH"},
        {"name": "R", "type": "Application", "status": "CUSTOM",  "confidence": "HIGH"},
    ]}
    rows = sort_rows(build_rows(res, _UUID_MAP_EMPTY))
    assert [r.name for r in rows] == ["L", "R", "C"]

def test_counters():
    res = {"entries": [
        {"name": "L1", "type": "Application", "status": "LINKED",  "confidence": "VERYHIGH"},
        {"name": "L2", "type": "Application", "status": "LINKED",  "confidence": "VERYHIGH"},
        {"name": "R1", "type": "Application", "status": "CUSTOM",  "confidence": "HIGH"},
        {"name": "C1", "type": "Application", "status": "CUSTOM",  "confidence": "NONE"},
    ]}
    counts = count_rows(build_rows(res, _UUID_MAP_EMPTY))
    assert counts == {"LINKED": 2, "REVIEW": 1, "CUSTOM": 1}

from pipeline.catalog_report import build_fs_url, build_catalog_search_url

def test_build_fs_url_uses_base_and_uuid():
    url = build_fs_url(
        base_url="https://demo-eu-3.leanix.net",
        workspace="demo-eu-3",
        fs_type="Application",
        uuid="abc123",
    )
    assert url == "https://demo-eu-3.leanix.net/demo-eu-3/factsheet/Application/abc123"

def test_build_fs_url_returns_none_without_uuid():
    assert build_fs_url("https://demo-eu-3.leanix.net", "demo-eu-3", "Application", None) is None

def test_build_catalog_search_url_url_encodes_name():
    url = build_catalog_search_url(
        base_url="https://demo-eu-3.leanix.net",
        workspace="demo-eu-3",
        name="SAP Ariba Buying",
    )
    assert url == "https://demo-eu-3.leanix.net/demo-eu-3/inventory/referenceCatalog?q=SAP%20Ariba%20Buying"

from pipeline.catalog_report import render_html

def test_render_html_contains_title_and_counts():
    res = {"entries": [
        {"name": "L", "type": "Application", "status": "LINKED",  "confidence": "VERYHIGH", "external_id": "lx_APP_1", "suggested_name": "L Cloud", "suggested_score": 1.0},
        {"name": "R", "type": "Application", "status": "CUSTOM",  "confidence": "HIGH",     "suggested_name": "R Buying", "suggested_score": 0.82},
        {"name": "C", "type": "Application", "status": "CUSTOM",  "confidence": "NONE"},
    ]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {"Application::L": {"uuid": "ll-1", "created": True},
                        "Application::R": {"uuid": "rr-2", "created": True},
                        "Application::C": {"uuid": "cc-3", "created": True}}, "failed": []}
    html = render_html(res, umap, client_name="Acme")
    assert "Catalog Linking Review" in html
    assert "Acme" in html
    assert "1 linked" in html
    assert "1 to review" in html
    assert "1 custom" in html

def test_render_html_includes_fs_link_for_linked_row():
    res = {"entries": [{"name": "L", "type": "Application", "status": "LINKED", "confidence": "VERYHIGH", "external_id": "lx_APP_1"}]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {"Application::L": {"uuid": "ll-1", "created": True}}, "failed": []}
    html = render_html(res, umap, client_name="Acme")
    assert "https://demo-eu-3.leanix.net/demo-eu-3/factsheet/Application/ll-1" in html

def test_render_html_search_link_only_on_review_rows():
    res = {"entries": [
        {"name": "L", "type": "Application", "status": "LINKED", "confidence": "VERYHIGH", "external_id": "lx_APP_1"},
        {"name": "R", "type": "Application", "status": "CUSTOM", "confidence": "HIGH",     "suggested_name": "R Buying"},
        {"name": "C", "type": "Application", "status": "CUSTOM", "confidence": "NONE"},
    ]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {"Application::L": {"uuid": "ll-1", "created": True},
                        "Application::R": {"uuid": "rr-2", "created": True},
                        "Application::C": {"uuid": "cc-3", "created": True}}, "failed": []}
    html = render_html(res, umap, client_name="Acme")
    # Only the REVIEW row should produce a referenceCatalog search link
    assert html.count("/inventory/referenceCatalog?q=") == 1
    assert "q=R" in html

def test_render_html_uses_status_css_classes():
    res = {"entries": [
        {"name": "L", "type": "Application", "status": "LINKED", "confidence": "VERYHIGH", "external_id": "lx_APP_1"},
        {"name": "R", "type": "Application", "status": "CUSTOM", "confidence": "HIGH"},
        {"name": "C", "type": "Application", "status": "CUSTOM", "confidence": "NONE"},
    ]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net", "entries": {}, "failed": []}
    html = render_html(res, umap, client_name="Acme")
    assert 'class="row row-linked"' in html
    assert 'class="row row-review"' in html
    assert 'class="row row-custom"' in html

def test_render_html_push_failed_row_marked():
    res = {"entries": [{"name": "Broken", "type": "Application", "status": "CUSTOM", "confidence": "NONE"}]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {}, "failed": [{"name": "Broken", "type": "Application", "error": "boom"}]}
    html = render_html(res, umap, client_name="Acme")
    assert "PUSH FAILED" in html


def test_render_html_escapes_dangerous_chars_in_name():
    res = {"entries": [{"name": "<script>alert(1)</script>", "type": "Application",
                        "status": "LINKED", "confidence": "VERYHIGH", "external_id": "lx_APP_1"}]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {}, "failed": []}
    html = render_html(res, umap, client_name="Acme")
    assert "<script>alert(1)</script>" not in html
    assert "&lt;script&gt;alert(1)&lt;/script&gt;" in html


def test_render_html_escapes_quotes_in_workspace_for_href():
    res = {"entries": [{"name": "L", "type": "Application", "status": "LINKED",
                        "confidence": "VERYHIGH", "external_id": "lx_APP_1"}]}
    umap = {"workspace": 'ws"oops', "base_url": "https://demo-eu-3.leanix.net",
            "entries": {"Application::L": {"uuid": "ll-1", "created": True}}, "failed": []}
    html = render_html(res, umap, client_name="Acme")
    # Raw `ws"oops` inside an href attribute would close the attribute.
    assert 'href="https://demo-eu-3.leanix.net/ws"oops/' not in html
    # The double quote must be escaped to &quot;
    assert "ws&quot;oops" in html


def test_render_html_escapes_confidence_html():
    res = {"entries": [{"name": "L", "type": "Application", "status": "LINKED",
                        "confidence": "<b>X</b>", "external_id": "lx_APP_1"}]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {}, "failed": []}
    html = render_html(res, umap, client_name="Acme")
    assert "<b>X</b>" not in html
    assert "&lt;b&gt;X&lt;/b&gt;" in html


import openpyxl
from pipeline.catalog_report import render_xlsx

def test_render_xlsx_writes_workbook_with_one_sheet(tmp_path):
    res = {"entries": [
        {"name": "L", "type": "Application", "status": "LINKED",  "confidence": "VERYHIGH", "external_id": "lx_APP_1", "suggested_name": "L Cloud", "suggested_score": 1.0},
        {"name": "R", "type": "Application", "status": "CUSTOM",  "confidence": "HIGH",     "suggested_name": "R Buying", "suggested_score": 0.82},
    ]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {"Application::L": {"uuid": "ll-1", "created": True},
                        "Application::R": {"uuid": "rr-2", "created": True}}, "failed": []}
    out = tmp_path / "report.xlsx"
    render_xlsx(res, umap, client_name="Acme", out_path=out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Catalog Report"]
    ws = wb["Catalog Report"]
    headers = [c.value for c in ws[1]]
    assert headers == ["Status", "Name", "Type", "Suggested match", "Suggested score", "externalId", "Open FS URL", "Search catalog URL"]
    assert ws.max_row == 3

def test_render_xlsx_open_fs_is_hyperlink(tmp_path):
    res = {"entries": [{"name": "L", "type": "Application", "status": "LINKED", "confidence": "VERYHIGH", "external_id": "lx_APP_1"}]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {"Application::L": {"uuid": "ll-1", "created": True}}, "failed": []}
    out = tmp_path / "report.xlsx"
    render_xlsx(res, umap, client_name="Acme", out_path=out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Catalog Report"]
    cell = ws.cell(row=2, column=7)
    assert cell.hyperlink is not None
    assert "factsheet/Application/ll-1" in cell.hyperlink.target


import json
import pytest
from pipeline.catalog_report import generate_report

def test_generate_report_writes_both_files(tmp_path):
    res = {"entries": [{"name": "L", "type": "Application", "status": "LINKED", "confidence": "VERYHIGH", "external_id": "lx_APP_1"}]}
    umap = {"workspace": "demo-eu-3", "base_url": "https://demo-eu-3.leanix.net",
            "entries": {"Application::L": {"uuid": "ll-1", "created": True}}, "failed": []}
    (tmp_path / "catalog_resolution_report.json").write_text(json.dumps(res))
    (tmp_path / "push_uuid_map.json").write_text(json.dumps(umap))

    generate_report(session_dir=tmp_path, client_name="Acme")

    assert (tmp_path / "catalog_report.html").exists()
    assert (tmp_path / "catalog_report.xlsx").exists()

def test_generate_report_raises_when_resolution_missing(tmp_path):
    (tmp_path / "push_uuid_map.json").write_text("{}")
    with pytest.raises(FileNotFoundError, match="Resolution report missing"):
        generate_report(tmp_path, "Acme")

def test_generate_report_raises_when_uuid_map_missing(tmp_path):
    (tmp_path / "catalog_resolution_report.json").write_text('{"entries": []}')
    with pytest.raises(FileNotFoundError, match="UUID map missing"):
        generate_report(tmp_path, "Acme")
