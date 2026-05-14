"""
run.py — Orchestrator for the Archimedes AI pipeline.

Commands:

  1. pipeline — Interactive 5-step pipeline (recommended entry point)
     python run.py pipeline --client <name> [--output-dir OUTPUT]

     Step 0: Catalog version check (RBA/RSA) with optional update
     Step 1: Baseline AS-IS from OnPrem/Cloud Systems Excel files
     Step 2: Target TO-BE from requirements Excel and/or PDF
              + optional SAP Help Portal contrast (help.sap.com)
     Step 3: Architecture diagrams/images (optional)
     Step 4: Generate LeanIX import Excel files

     Outputs:
       output/<client>_baseline.xlsx      ← AS-IS Applications (Baseline)
       output/<client>_target.xlsx        ← TO-BE fact sheets (BC, App, Org, Initiative, ITC)

  2. enrich — Extract → enrich → validate → write (requirements only)
     python run.py enrich <input_file> --client <name> [options]

     Outputs:
       output/<stem>_enriched.xlsx        ← client deliverable (cols H–P)
       output/<client>_leanix_import.xlsx ← staging file for LeanIX (review before push)

  3. push — Read staging Excel → push to LeanIX
     python run.py push <leanix_import.xlsx> --client <name>

     Run after reviewing / editing the staging file.

Options for enrich:
  --no-validate         Skip validation step
  --client NAME         Client name used as tag in LeanIX (default: unknown)
  --output-dir OUTPUT   Directory for output files (default: output)

Environment variables (see .env.example):
  ANTHROPIC_API_KEY, LEANIX_API_TOKEN, LEANIX_BASE_URL
  ENRICH_MODEL, ENRICH_BATCH_SIZE, LOG_LEVEL
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()

sys.path.insert(0, str(Path(__file__).parent / "pipeline"))

from pipeline.extract  import extract     # noqa: E402
from pipeline.enrich   import enrich      # noqa: E402
from pipeline.validate import validate    # noqa: E402
from pipeline.write    import write, push_leanix, write_leanix_excel_from_xlsx  # noqa: E402
from pipeline.catalog  import check_catalogs      # noqa: E402
from pipeline.footprint import generate_baseline  # noqa: E402
from pipeline.pdf_extract import extract_pdf_factsheets   # noqa: E402
from pipeline.image_extract import extract_image_factsheets  # noqa: E402
from pipeline.help_contrast import run_contrast, print_contrast_summary  # noqa: E402


def _setup_logging(level: str) -> None:
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)-8s %(name)s — %(message)s",
        datefmt="%H:%M:%S",
    )


def cmd_pipeline(args: argparse.Namespace) -> None:
    """Interactive 5-step LeanIX pipeline."""
    import json
    import re
    import anthropic as _anthropic

    logger = logging.getLogger("archimedes")
    client_name = re.sub(r"[^\w\-]", "_", args.client)  # strip path separators
    if client_name != args.client:
        logger.warning("Client name sanitized: '%s' → '%s'", args.client, client_name)
    output_dir  = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print("\n" + "═" * 60)
    print("  ARCHIMEDES AI — LeanIX Pipeline")
    print(f"  Cliente: {client_name}")
    print(f"  Output:  {output_dir}")
    print("═" * 60)

    # ── PASO 0 — Catálogos ────────────────────────────────────────────────────
    ok = check_catalogs(interactive=True)
    if not ok:
        logger.error("Pipeline aborted by user.")
        return

    anthropic_client = _anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    model = os.getenv("ENRICH_MODEL", "claude-sonnet-4-6")

    # ── PASO 1 — Footprint / Baseline (AS-IS) ─────────────────────────────────
    print("\n" + "═" * 60)
    print("  PASO 1 — Footprint / Baseline (AS-IS)")
    print("═" * 60)

    onprem_path = None
    cloud_path  = None

    op_input = input("  ¿Tienes un fichero OnPrem Systems.xlsx? Ruta (Enter para omitir): ").strip()
    if op_input:
        p = Path(op_input).expanduser()
        if p.exists():
            onprem_path = p
        else:
            print(f"  ⚠  No encontrado: {p} — omitiendo.")

    cl_input = input("  ¿Tienes un fichero Cloud Systems.xlsx? Ruta (Enter para omitir): ").strip()
    if cl_input:
        p = Path(cl_input).expanduser()
        if p.exists():
            cloud_path = p
        else:
            print(f"  ⚠  No encontrado: {p} — omitiendo.")

    baseline_result = None
    if onprem_path or cloud_path:
        baseline_out = output_dir / f"{client_name}_baseline.xlsx"
        baseline_result = generate_baseline(
            output_path=baseline_out,
            client_name=client_name,
            onprem_path=onprem_path,
            cloud_path=cloud_path,
        )
        print(f"\n  ✓ Baseline generado: {baseline_out}")
        print(f"    OnPremise: {baseline_result['n_onprem']}  |  Cloud: {baseline_result['n_cloud']}  |  Total: {baseline_result['n_total']}")
    else:
        print("  → Sin ficheros de footprint. Paso omitido.")

    # ── PASO 2 — Requerimientos / PDF (Target TO-BE) ──────────────────────────
    print("\n" + "═" * 60)
    print("  PASO 2 — Requerimientos / Target (TO-BE)")
    print("═" * 60)

    req_excel_path   = None
    req_enriched_xlsx = None   # Excel already enriched by map_requirements.py
    target_json_path = None

    req_input = input("  ¿Tienes un Excel de requerimientos? Ruta (Enter para omitir): ").strip()
    if req_input:
        p = Path(req_input).expanduser()
        if p.exists():
            req_excel_path = p
        else:
            print(f"  ⚠  No encontrado: {p} — omitiendo.")

    if req_excel_path:
        # Detect if already enriched by map_requirements.py (has data in col O/P from row 9)
        import openpyxl as _xl
        wb_check = _xl.load_workbook(str(req_excel_path))
        ws_check = wb_check.active
        already_enriched = any(
            ws_check.cell(r, 15).value or ws_check.cell(r, 16).value
            for r in range(9, min(15, ws_check.max_row + 1))
        )

        if already_enriched:
            print(f"  → Excel ya enriquecido detectado (columnas O/P presentes). Usando directamente.")
            req_enriched_xlsx = req_excel_path
        else:
            print(f"  → Extrayendo y enriqueciendo con Claude API + catálogos RBA/RSA …")
            raw_path = extract(req_excel_path, output_dir)
            enriched_path = enrich(raw_path, output_dir)
            if not args.no_validate:
                print("  → Validando …")
                ok = validate(enriched_path, output_dir)
                if not ok:
                    print(f"  ⚠  Validación con errores. Revisa {output_dir}/validation_report.json")
            target_json_path = enriched_path

    # ── Contraste SAP Help Portal (opcional) ──────────────────────────────────
    help_contrast_path = None
    if target_json_path:
        contrast_answer = input("  ¿Contrastar el mapeo RSA con SAP Help Portal (help.sap.com)? [s/N]: ").strip().lower()
        if contrast_answer in ("s", "si", "sí", "y", "yes"):
            print("  → Consultando SAP Help Portal para validar productos RSA …")
            help_contrast_path = run_contrast(target_json_path, output_dir)
            print_contrast_summary(help_contrast_path)
        else:
            print("  → Contraste SAP Help Portal omitido.")

    # PDF additional info
    pdf_factsheets: dict = {}
    pdf_input = input("  ¿Tienes un PDF con información adicional del cliente? Ruta (Enter para omitir): ").strip()
    if pdf_input:
        p = Path(pdf_input).expanduser()
        if p.exists():
            print(f"  → Extrayendo fact sheets del PDF {p.name} con Claude API …")
            pdf_factsheets = extract_pdf_factsheets(p, client_name, anthropic_client, model)
            print(f"  ✓ PDF: {pdf_factsheets.get('summary', '')}")
            n_apps = len(pdf_factsheets.get("applications", []))
            n_bcs  = len(pdf_factsheets.get("business_capabilities", []))
            print(f"    Apps: {n_apps}  |  BCs: {n_bcs}  |  Orgs: {len(pdf_factsheets.get('organizations', []))}  |  Initiatives: {len(pdf_factsheets.get('initiatives', []))}  |  ITCs: {len(pdf_factsheets.get('it_components', []))}")
        else:
            print(f"  ⚠  No encontrado: {p} — omitiendo.")

    # ── PASO 3 — Imágenes / Diagramas ─────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  PASO 3 — Imágenes / Diagramas de arquitectura")
    print("═" * 60)

    image_factsheets: dict = {}
    img_answer = input("  ¿Tienes imágenes o diagramas de arquitectura? [s/N]: ").strip().lower()
    if img_answer in ("s", "si", "sí", "y", "yes"):
        print("  Introduce las rutas de las imágenes una por línea.")
        print("  Formatos soportados: .png, .jpg, .jpeg, .gif, .webp")
        print("  (Enter en línea vacía para terminar)")
        image_paths = []
        while True:
            line = input("    Imagen: ").strip()
            if not line:
                break
            p = Path(line).expanduser()
            if p.exists():
                image_paths.append(p)
            else:
                print(f"    ⚠  No encontrado: {p}")

        if image_paths:
            print(f"  → Analizando {len(image_paths)} imagen(es) con Claude API …")
            image_factsheets = extract_image_factsheets(image_paths, client_name, anthropic_client, model)
            print(f"  ✓ Diagramas: {image_factsheets.get('summary', '')}")
            print(f"    Apps: {len(image_factsheets.get('applications', []))}  |  BCs: {len(image_factsheets.get('business_capabilities', []))}  |  ITCs: {len(image_factsheets.get('it_components', []))}  |  Interfaces: {len(image_factsheets.get('interfaces', []))}")
        else:
            print("  → Sin imágenes válidas. Paso omitido.")
    else:
        print("  → Paso omitido.")

    # ── PASO 4 — Generación outputs LeanIX ────────────────────────────────────
    print("\n" + "═" * 60)
    print("  PASO 4 — Generación outputs LeanIX")
    print("═" * 60)

    out_enriched  = None
    out_target    = None
    out_supp      = None

    # Assemble supplementary dict from PDF + images (populated in steps 2-3)
    supplementary = {}
    if pdf_factsheets:
        supplementary["from_pdf"] = pdf_factsheets
    if image_factsheets:
        supplementary["from_images"] = image_factsheets

    if req_enriched_xlsx:
        # Already-enriched Excel from map_requirements.py → generate LeanIX target directly
        out_enriched = req_enriched_xlsx
        out_target   = output_dir / f"{client_name}_target_leanix.xlsx"
        write_leanix_excel_from_xlsx(req_enriched_xlsx, out_target, client_name, supplementary=supplementary or None)
        print(f"\n  ✓ [1] Requerimientos enriquecidos (cliente):  {out_enriched}")
        print(f"  ✓ [2] Target LeanIX importable (TO-BE):       {out_target}")
    elif target_json_path and req_excel_path:
        out_enriched, out_target = write(
            target_json_path, req_excel_path, output_dir, client_name=client_name,
            supplementary=supplementary or None,
        )
        print(f"\n  ✓ [1] Requerimientos enriquecidos (cliente):  {out_enriched}")
        print(f"  ✓ [2] Target LeanIX importable (TO-BE):       {out_target}")

    # Also save supplementary JSON as backup reference
    if supplementary:
        out_supp = output_dir / f"{client_name}_supplementary_factsheets.json"
        out_supp.write_text(json.dumps(supplementary, ensure_ascii=False, indent=2))
        print(f"  ✓ [3] Fact sheets adicionales (PDF+imágenes): {out_supp}")
        if out_target:
            print(f"      → Ya incluidos en {out_target.name}")
        else:
            print(f"      → Revisa y completa manualmente antes de importar en LeanIX.")

    if not (target_json_path or pdf_factsheets or image_factsheets):
        print("  → Sin datos Target. Paso omitido.")

    # ── PASO 5 — Import a LeanIX ───────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  PASO 5 — Import a LeanIX (opcional)")
    print("═" * 60)
    print()
    print("  Ficheros listos para importar:")
    if baseline_result and baseline_result.get("output_path"):
        print(f"    [A] Baseline AS-IS:  {baseline_result['output_path']}")
    if out_target:
        print(f"    [B] Target TO-BE:    {out_target}")
    print()

    import_answer = input("  ¿Importar ahora en LeanIX vía API? [s/N]: ").strip().lower()
    if import_answer in ("s", "si", "sí", "y", "yes"):
        if not os.environ.get("LEANIX_API_TOKEN") or not os.environ.get("LEANIX_BASE_URL"):
            print("  ⚠  LEANIX_API_TOKEN y/o LEANIX_BASE_URL no configurados en .env")
            print("     Configúralos y ejecuta manualmente:")
            if baseline_result and baseline_result.get("output_path"):
                print(f"     python3 run.py push {baseline_result['output_path']} --client {client_name}")
            if out_target:
                print(f"     python3 run.py push {out_target} --client {client_name}")
        else:
            which = input("  ¿Qué importar? [A=Baseline / B=Target / AB=ambos]: ").strip().upper()
            if which in ("A", "AB") and baseline_result and baseline_result.get("output_path"):
                print(f"  → Importando Baseline …")
                push_leanix(baseline_result["output_path"], client_name)
                print(f"  ✓ Baseline importado.")
            if which in ("B", "AB") and out_target:
                print(f"  → Importando Target …")
                push_leanix(out_target, client_name)
                print(f"  ✓ Target importado.")
    else:
        print("  → Import manual. Usa LeanIX: Settings → Import/Export → Import from Excel")
        print("    Orden: BusinessCapability → Application → Initiative → ITComponent")

    # ── Resumen final ──────────────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  RESUMEN FINAL")
    print("═" * 60)
    if baseline_result and baseline_result.get("output_path"):
        print(f"  [1] Baseline AS-IS:              {baseline_result['output_path']}")
        print(f"      {baseline_result['n_total']} apps  ({baseline_result['n_onprem']} on-prem + {baseline_result['n_cloud']} cloud)")
    if out_enriched:
        print(f"  [2] Requerimientos enriquecidos: {out_enriched}")
    if out_target:
        print(f"  [3] Target LeanIX importable:    {out_target}")
    if out_supp:
        print(f"  [4] Fact sheets adicionales:     {out_supp}")
    if help_contrast_path:
        print(f"  [5] Contraste SAP Help Portal:   {help_contrast_path}")
    print("═" * 60 + "\n")


def cmd_enrich(args: argparse.Namespace) -> None:
    logger = logging.getLogger("archimedes")

    input_path = Path(args.input_file)
    output_dir = Path(args.output_dir)

    if not input_path.exists():
        logger.error("Input file not found: %s", input_path)
        sys.exit(1)

    logger.info("═" * 60)
    logger.info("Archimedes AI — enrich")
    logger.info("Input:  %s", input_path)
    logger.info("Client: %s", args.client)
    logger.info("Output: %s", output_dir)
    logger.info("═" * 60)

    # Step 1 — Extract
    logger.info("Step 1/4 — Extract")
    raw_path = extract(input_path, output_dir)

    # Step 2 — Enrich
    logger.info("Step 2/4 — Enrich")
    enriched_path = enrich(raw_path, output_dir)

    # Step 3 — Validate
    if not args.no_validate:
        logger.info("Step 3/4 — Validate")
        ok = validate(enriched_path, output_dir)
        if not ok:
            logger.error(
                "Validation failed. Fix errors in %s/validation_report.json "
                "or re-run with --no-validate to skip.",
                output_dir,
            )
            sys.exit(1)
    else:
        logger.info("Step 3/4 — Validate (SKIPPED)")

    # Step 4 — Write
    logger.info("Step 4/4 — Write")
    out_excel, out_staging = write(
        enriched_path,
        input_path,
        output_dir,
        client_name=args.client,
    )

    logger.info("═" * 60)
    logger.info("Done.")
    logger.info("  Client Excel  : %s", out_excel)
    logger.info("  LeanIX import : %s", out_staging)
    logger.info("  Enriched JSON : %s", enriched_path)
    if not args.no_validate:
        logger.info("  Validation    : %s/validation_report.json", output_dir)
    logger.info("─" * 60)
    logger.info("Review %s, then run:", out_staging)
    logger.info("  python run.py push %s --client %s", out_staging, args.client)
    logger.info("═" * 60)


def cmd_push(args: argparse.Namespace) -> None:
    logger = logging.getLogger("archimedes")
    staging = Path(args.staging_file)

    if not staging.exists():
        logger.error("Staging file not found: %s", staging)
        sys.exit(1)

    logger.info("═" * 60)
    logger.info("Archimedes AI — push to LeanIX")
    logger.info("Staging: %s", staging)
    logger.info("Client:  %s", args.client)
    logger.info("═" * 60)

    push_leanix(staging, client_name=args.client)

    logger.info("═" * 60)
    logger.info("LeanIX push complete.")
    logger.info("═" * 60)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Archimedes AI — SAP requirements enrichment pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # pipeline sub-command
    p_pipeline = sub.add_parser("pipeline", help="Interactive 5-step LeanIX pipeline")
    p_pipeline.add_argument("--client", default="unknown", help="Client name (used as LeanIX tag)")
    p_pipeline.add_argument("--output-dir", default="output", help="Output directory")
    p_pipeline.add_argument("--no-validate", action="store_true", help="Skip validation step")

    # enrich sub-command
    p_enrich = sub.add_parser("enrich", help="Extract → enrich → validate → write")
    p_enrich.add_argument("input_file", help="Path to .xlsx, .xls, or .pdf")
    p_enrich.add_argument("--no-validate", action="store_true", help="Skip validation")
    p_enrich.add_argument("--client", default="unknown", help="Client name (used as LeanIX tag)")
    p_enrich.add_argument("--output-dir", default="output", help="Output directory")

    # push sub-command
    p_push = sub.add_parser("push", help="Push staging Excel to LeanIX")
    p_push.add_argument("staging_file", help="Path to <client>_leanix_import.xlsx")
    p_push.add_argument("--client", default="unknown", help="Client name (used as LeanIX tag)")

    args = parser.parse_args()
    _setup_logging(os.getenv("LOG_LEVEL", "INFO"))

    if args.command == "pipeline":
        cmd_pipeline(args)
    elif args.command == "enrich":
        cmd_enrich(args)
    else:
        cmd_push(args)


if __name__ == "__main__":
    main()
