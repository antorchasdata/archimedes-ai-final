"""
pipeline/industry_reference.py — SAP Industry Reference Architecture whitespace analysis.

Fetches SAP-recommended products for a given industry from the SAP Business Accelerator Hub
OData API (public, no auth required), crosses with the RSA catalog, and compares against
the client's Baseline to identify the whitespace (apps in the reference but not in the client).
"""

from __future__ import annotations

import json
import logging
import urllib.request
import urllib.parse
from pathlib import Path
from typing import Optional

logger = logging.getLogger("archimedes.industry_reference")

# ── SAP Business Accelerator Hub OData ────────────────────────────────────────
_API_BASE = "https://api.sap.com/odata/1.0/catalog.svc/ContentPackages"
_FIELDS   = "DisplayName,Industries,Products"

# ── Industry name mapping (API Hub label → display label) ─────────────────────
INDUSTRIES = {
    "eco":    "Engineering Construction and Operations",
    "retail": "Retail",
    "aad":    "Aerospace and Defense",
    "auto":   "Automotive",
    "chem":   "Chemicals",
    "cp":     "Consumer Products",
    "fs":     "Financial Services",
    "hc":     "Healthcare",
    "ht":     "High Tech",
    "imc":    "Industrial Machinery and Components",
    "ls":     "Life Sciences",
    "media":  "Media",
    "mill":   "Mill Products",
    "mining": "Mining",
    "oilgas": "Oil and Gas",
    "ps":     "Professional Services",
    "pubsec": "Public Sector",
    "telco":  "Telecommunications",
    "travel": "Travel and Transportation",
    "utils":  "Utilities",
    "wd":     "Wholesale Distribution",
}

# ── Non-SAP product codes to exclude ──────────────────────────────────────────
_NON_SAP_PREFIXES = (
    "Amazon", "Atlassian", "CELUM", "Cobrainer", "Degreed", "ELK",
    "Government", "JIRA", "KTernAI", "MQTT", "Microsoft", "PiLog",
    "Planon", "Pricefx", "Replicon", "ServiceNow", "Sirion", "Splunk",
    "ThirdParty", "Twenty5", "DOST",
)

# ── Product code → display name normalisation ─────────────────────────────────
_PRODUCT_NORMALISE = {
    "SAPS4HANA":                          "SAP S/4HANA",
    "SAPS4HANACloud":                     "SAP S/4HANA Cloud",
    "SAPS4HANACloudPrivateEdition":       "SAP S/4HANA Cloud Private Edition",
    "SAPS4HANACloudPublicEdition":        "SAP S/4HANA Cloud Public Edition",
    "SAPAriba":                           "SAP Ariba",
    "SAPAribaCloudIntegrationGateway":    "SAP Ariba Cloud Integration Gateway",
    "SAPAriabProcurement":                "SAP Ariba Procurement",
    "SAPSuccessFactors":                  "SAP SuccessFactors",
    "SAPSuccessFactorsEmployeeCentral":   "SAP SuccessFactors Employee Central",
    "SAPSuccessFactorsLearning":          "SAP SuccessFactors Learning",
    "SAPSuccessFactorsRecruitingManagement": "SAP SuccessFactors Recruiting Management",
    "SAPIntegratedBusinessPlanningforSupplyChain": "SAP Integrated Business Planning",
    "SAPProjectandResourceManagement":    "SAP Project and Resource Management",
    "SAPFieldServiceManagement":          "SAP Field Service Management",
    "SAPFieldglass":                      "SAP Fieldglass Vendor Management System",
    "SAPAssetPerformanceManagement":      "SAP Asset Performance Management",
    "SAPAssetIntelligenceNetwork":        "SAP Asset Intelligence Network",
    "SAPCloudALM":                        "SAP Cloud ALM",
    "SAPCloudPlatform":                   "SAP Business Technology Platform",
    "SAPCloudPlatformIntegration":        "SAP Integration Suite",
    "SAPCloudPlatformIntegrationSuite":   "SAP Integration Suite",
    "SAPCloudPlatformBusinessRules":      "SAP Build Process Automation",
    "SAPCloudPlatformWorkflow":           "SAP Build Process Automation",
    "SAPCloudPlatformWorkflowManagement": "SAP Build Process Automation",
    "SAPProcessAutomation":               "SAP Process Automation",
    "SAPBuildProcessAutomation":          "SAP Build Process Automation",
    "SAPProcessOrchestration":            "SAP Process Orchestration",
    "SAPAnalyticsCloud":                  "SAP Analytics Cloud",
    "SAPHANA":                            "SAP HANA",
    "SAPHCM":                             "SAP HCM",
    "SAPBusinessByDesign":                "SAP Business ByDesign",
    "SAPBusinessSuite":                   "SAP Business Suite",
    "SAPERP":                             "SAP ERP",
    "SAPERPCentralComponent":             "SAP ERP Central Component",
    "SAPEWM":                             "SAP Extended Warehouse Management",
    "SAPS4HANAAssetManagement":           "SAP S/4HANA Asset Management",
    "SAPS4HANAUtilities":                 "SAP S/4HANA for Utilities",
    "SAPCPQ":                             "SAP Configure Price Quote",
    "SAPCRM":                             "SAP CRM",
    "SAPSolutionManager":                 "SAP Solution Manager",
    "SAPServiceandAssetManager":          "SAP Service and Asset Manager",
    "SAPInternetofThings":                "SAP IoT",
    "SAPEventMesh":                       "SAP Event Mesh",
    "SAPDocumentCompliance":              "SAP Document Compliance",
    "SAPDocumentandReportingCompliance":  "SAP Document and Reporting Compliance",
    "SAPDocumentManagementservice":       "SAP Document Management Service",
    "SAPEntitlementManagement":           "SAP Entitlement Management",
    "SAPSubscriptionBilling":             "SAP Subscription Billing",
    "SAPVariantConfigurationandPricing":  "SAP Variant Configuration and Pricing",
    "SAPLogisticsBusinessNetworkglobaltrackandtraceoption": "SAP Logistics Business Network",
    "SAPReturnablePackagingManagement":   "SAP Returnable Packaging Management",
    "SAPMobileServices":                  "SAP Mobile Services",
    "SAPSignavioProcessManager":          "SAP Signavio Process Manager",
    "SAPSRM":                             "SAP Supplier Relationship Management",
    "SAPCloudforCustomer":                "SAP Sales Cloud",
    "SAPMarketingCloud":                  "SAP Marketing Cloud",
    "SAPEMobility":                       "SAP E-Mobility",
    "SAPCloudPlatformInternetofThings":   "SAP IoT",
    "SAPCloudPlatformMasterDataIntegration": "SAP Master Data Integration",
    "SAPCloudPlatformProcessVisibility":  "SAP Process Insights",
    "SAPIndustryProcessFramework":        "SAP Industry Process Framework",
    "SAPBusinessTechnologyPlatform":      "SAP Business Technology Platform",
    "SAPERPoptionforedocumentprocessing": "SAP ERP",
    "SAPDocumentComplianceonpremiseedition": "SAP Document Compliance",
    "SAPS4HANACloudforIntelligentProductDesign": "SAP S/4HANA Cloud for Intelligent Product Design",
    "SAPS4HANAFinance":                   "SAP S/4HANA Finance",
    "SAPLogisticsBusinessNetwork":        "SAP Logistics Business Network",
    "SAPS4HANAAssetManagementforresourcescheduling": "SAP S/4HANA Asset Management",
    "SAPS4HANAforprocurementplanning":    "SAP S/4HANA for Procurement Planning",
    "SAPReturnablePackagingManagement":   "SAP Returnable Packaging Management",
}


def _fetch_industry_products(industry_label: str) -> list[str]:
    """Fetch all SAP product names for a given industry label from SAP API Hub."""
    all_packages = []
    skip = 0

    while True:
        params = urllib.parse.urlencode({
            "$select": _FIELDS,
            "$top":    200,
            "$skip":   skip,
            "$format": "json",
        })
        url = f"{_API_BASE}?{params}"
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        try:
            with urllib.request.urlopen(req, timeout=20) as r:
                data = json.loads(r.read())
        except Exception as exc:
            logger.warning("API Hub fetch error at skip=%d: %s", skip, exc)
            break

        results = data.get("d", {}).get("results", [])
        if not results:
            break
        all_packages.extend(results)
        skip += len(results)
        if len(results) < 200:
            break

    logger.info("Fetched %d total packages from SAP API Hub", len(all_packages))

    # Filter by industry + has products
    industry_packages = [
        r for r in all_packages
        if industry_label in (r.get("Industries") or "")
        and r.get("Products")
    ]
    logger.info("%d packages for industry '%s'", len(industry_packages), industry_label)

    # Extract + normalize unique SAP product names
    product_names: set[str] = set()
    for pkg in industry_packages:
        for raw in pkg["Products"].split(","):
            raw = raw.strip()
            if not raw:
                continue
            # Skip non-SAP
            if any(raw.startswith(p) for p in _NON_SAP_PREFIXES):
                continue
            # Normalize code → display name
            name = _PRODUCT_NORMALISE.get(raw, raw)
            product_names.add(name)

    return sorted(product_names)


def get_industry_reference(industry_key: str) -> dict:
    """
    Returns the SAP reference architecture for an industry.

    Returns:
        {
            "industry_key": "eco",
            "industry_label": "Engineering Construction and Operations",
            "products": ["SAP S/4HANA", "SAP Ariba", ...],
            "n_products": 42,
        }
    """
    industry_label = INDUSTRIES.get(industry_key)
    if not industry_label:
        raise ValueError(f"Unknown industry key: {industry_key!r}. Valid keys: {list(INDUSTRIES)}")

    products = _fetch_industry_products(industry_label)

    return {
        "industry_key":   industry_key,
        "industry_label": industry_label,
        "products":       products,
        "n_products":     len(products),
    }


def _load_app_names_from_excel(path: Path) -> set[str]:
    """Read Application sheet col C (Name, row 3+) from a LeanIX-format Excel."""
    names: set[str] = set()
    if not path or not path.exists():
        return names
    try:
        import openpyxl as xl
        wb = xl.load_workbook(str(path), read_only=True, data_only=True)
        if "Application" not in wb.sheetnames:
            return names
        ws = wb["Application"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            # col C = index 2 = Name
            name = row[2] if len(row) > 2 else None
            if name:
                names.add(str(name).strip().lower())
    except Exception as exc:
        logger.warning("Could not read %s: %s", path, exc)
    return names


def compute_whitespace(
    reference_products: list[str],
    baseline_path: Optional[Path],
    target_path: Optional[Path] = None,
) -> list[dict]:
    """
    Compare reference products against the client's Baseline AND Target Excels.

    Returns list of dicts:
        {
            "name": "SAP Signavio Process Manager",
            "in_baseline": False,
            "in_target": False,
            "include": False,
        }
    """
    baseline_names = _load_app_names_from_excel(baseline_path)
    target_names   = _load_app_names_from_excel(target_path)
    known_names    = baseline_names | target_names

    result = []
    for prod in reference_products:
        key = prod.strip().lower()
        result.append({
            "name":        prod,
            "in_baseline": key in baseline_names,
            "in_target":   key in target_names,
            "in_client":   key in known_names,   # already covered (baseline OR target)
            "include":     False,
        })

    return result


def add_reference_to_target_excel(
    target_path: Path,
    selected_products: list[str],
    industry_label: str,
    client_name: str = "",
) -> Path:
    """
    Adds selected reference products to the Target LeanIX Excel.
    Appends rows to the Application sheet with the full LeanIX column format
    and tag: "Target Reference".
    Returns the modified path (same file, modified in-place).
    """
    import openpyxl as xl
    from pipeline.write import _COLS_APPLICATION, _sheet_header, _sheet_row

    tag = "Target Reference"
    description = f"SAP recommended product for {industry_label} — from SAP Reference Architecture."
    if client_name:
        description += f" Client: {client_name}."

    wb = xl.load_workbook(str(target_path))

    if "Application" not in wb.sheetnames:
        ws = wb.create_sheet("Application")
        _sheet_header(ws, _COLS_APPLICATION)
        next_row = 3
    else:
        ws = wb["Application"]
        next_row = ws.max_row + 1

    keys_app = [c[0] for c in _COLS_APPLICATION]

    for prod in selected_products:
        vals = {
            "id": "", "type": "Application", "name": prod,
            "description": description,
            "alias": "", "externalId": "",
            "lifecycle_phase": "plan", "lifecycle_startDate": "", "lifecycle_endDate": "",
            "businessCriticality": "", "functionalSuitability": "",
            "technicalSuitability": "", "lxHostingType": "saas", "lxState": "DRAFT",
            "tags": tag,
            "relApplicationToBusinessCapability": "",
            "relApplicationToITComponent": "", "relToParent": "",
        }
        _sheet_row(ws, next_row, [vals.get(k, "") for k in keys_app])
        next_row += 1

    wb.save(str(target_path))
    logger.info("Added %d reference products to %s", len(selected_products), target_path.name)

    return target_path
