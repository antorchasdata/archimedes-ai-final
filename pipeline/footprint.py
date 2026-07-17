"""
footprint.py — Baseline AS-IS generator.

Reads OnPrem Systems.xlsx and/or Cloud Systems.xlsx and generates a
LeanIX-ready Excel with Applications tagged Baseline;OnPremise / Baseline;Cloud.

Public API:
    generate_baseline(onprem_path, cloud_path, output_path, client_name) -> dict
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

logger = logging.getLogger(__name__)

# ── Colors ─────────────────────────────────────────────────────────────────────
ROW_ONPREM = "E8F5E9"   # light green
ROW_CLOUD  = "E3F2FD"   # light blue
COLOR_MANDATORY = "002A86"
COLOR_OPTIONAL  = "0070F2"
COLOR_RELATION  = "107E3E"
COLOR_READONLY  = "A9B4BE"
COLOR_TRANS_ROW = "EAF4FF"

# ── Column schema ──────────────────────────────────────────────────────────────
COLUMNS = [
    "id", "type", "name", "description", "alias", "externalId",
    "lifecycle_phase", "lifecycle_startDate", "lifecycle_endDate",
    "businessCriticality", "functionalSuitability", "technicalSuitability",
    "lxHostingType", "lxState", "tags",
    "relApplicationToITComponent", "relApplicationToBusinessCapability", "relToParent",
]

TRANSLATIONS = {
    "id": "ID", "type": "Type", "name": "Name", "description": "Description",
    "alias": "Alias", "externalId": "External ID",
    "lifecycle_phase": "Lifecycle Phase", "lifecycle_startDate": "Lifecycle Start Date",
    "lifecycle_endDate": "Lifecycle End Date", "businessCriticality": "Business Criticality",
    "functionalSuitability": "Functional Fit", "technicalSuitability": "Technical Fit",
    "lxHostingType": "Hosting Type", "lxState": "Quality Seal", "tags": "Tags",
    "relApplicationToITComponent": "IT Components",
    "relApplicationToBusinessCapability": "Business Capabilities",
    "relToParent": "Parent Application",
}

CATEGORIES = {
    "id": "readonly", "type": "mandatory", "name": "mandatory",
    "description": "optional", "alias": "optional", "externalId": "optional",
    "lifecycle_phase": "optional", "lifecycle_startDate": "optional",
    "lifecycle_endDate": "optional", "businessCriticality": "optional",
    "functionalSuitability": "optional", "technicalSuitability": "optional",
    "lxHostingType": "optional", "lxState": "optional", "tags": "optional",
    "relApplicationToITComponent": "relation",
    "relApplicationToBusinessCapability": "relation",
    "relToParent": "relation",
}

HEADER_COLORS = {
    "mandatory": COLOR_MANDATORY, "optional": COLOR_OPTIONAL,
    "relation": COLOR_RELATION,   "readonly": COLOR_READONLY,
}

COL_WIDTHS = {
    "id": 36, "type": 14, "name": 38, "description": 50, "alias": 14,
    "externalId": 18, "lifecycle_phase": 16, "lifecycle_startDate": 20,
    "lifecycle_endDate": 18, "businessCriticality": 22, "functionalSuitability": 20,
    "technicalSuitability": 20, "lxHostingType": 18, "lxState": 22, "tags": 28,
    "relApplicationToITComponent": 45,
    "relApplicationToBusinessCapability": 55, "relToParent": 30,
}

# ── Product mappings ───────────────────────────────────────────────────────────
PRODUCT_PRIORITY = {
    "SAP S/4HANA": 0, "SAP S/4HANA FOUNDATION": 1, "ABAP PLATFORM": 2,
    "SAP ERP": 3, "SAP HANA PLATFORM EDITION": 4, "SAP SOLUTION MANAGER": 5,
    "SAP NETWEAVER": 6, "NW AS ABAP INNOVATION PACKAGE": 7,
}

PRODUCT_TO_ITC = {
    "SAP ERP":                       "SAP ERP 6.0",
    "SAP NETWEAVER":                 "SAP NetWeaver",
    "SAP HANA PLATFORM EDITION":     "SAP HANA Platform",
    "SAP S/4HANA":                   "SAP S/4HANA",
    "SAP S/4HANA FOUNDATION":        "SAP S/4HANA Foundation",
    "ABAP PLATFORM":                 "ABAP Platform",
    "NW AS ABAP INNOVATION PACKAGE": "SAP NetWeaver",
    "SAP SOLUTION MANAGER":          "SAP Solution Manager",
}

PRODUCT_TO_BC = {
    "SAP ERP":     "Finance / Accounting and Financial Close",
    "SAP S/4HANA": "Finance / Accounting and Financial Close;Sourcing and Procurement / Operational Procurement;Supply Chain Execution / Inventory Management",
    "SAP S/4HANA FOUNDATION": "",
    "SAP HANA PLATFORM EDITION": "",
    "SAP NETWEAVER": "",
    "ABAP PLATFORM": "",
    "NW AS ABAP INNOVATION PACKAGE": "",
    "SAP SOLUTION MANAGER": "",
}

CLOUD_ROLE_TO_APP: dict[str, tuple[str, str, str]] = {
    "SAP Unified Account":                              ("SAP Unified Account",                 "BTP",  "saas"),
    "Identity Authentication":                          ("SAP Identity Authentication",          "BTP",  "saas"),
    "Identity Provisioning":                            ("SAP Identity Provisioning",            "BTP",  "saas"),
    "SAP Datasphere, SAP BW Bridge":                    ("SAP Datasphere with BW Bridge",        "Data", "saas"),
    "SAP Datasphere":                                   ("SAP Datasphere",                       "Data", "saas"),
    "SAP Analytics Cloud":                              ("SAP Analytics Cloud",                  "Data", "saas"),
    "SAP Signature Management by DocuSign":             ("SAP Signature Management by DocuSign", "Other","saas"),
    "SAP Business Network":                             ("SAP Business Network",                 "SCM",  "saas"),
    "SAP Ariba Procurement":                            ("SAP Ariba Procurement",                "Proc", "saas"),
    "SAP Ariba Sourcing":                               ("SAP Ariba Sourcing",                   "Proc", "saas"),
    "SAP Ariba Shopping":                               ("SAP Ariba Shopping",                   "Proc", "saas"),
    "SAP Build Work Zone, standard edition":            ("SAP Build Work Zone",                  "BTP",  "saas"),
    "Cloud Management Tools":                           ("SAP BTP Cloud Management Tools",       "BTP",  "saas"),
    "foundational services for SAP BTP":                ("SAP BTP Foundational Services",        "BTP",  "paas"),
    "lifecycle management for SAP BTP, Cloud Foundry runtime": ("SAP BTP CF Lifecycle Mgmt",    "BTP",  "paas"),
    "Joule":                                            ("SAP Joule",                            "AI",   "saas"),
    "SAP ID Service":                                   ("SAP ID Service",                       "BTP",  "saas"),
    "SAP BTP, Cloud Foundry runtime":                   ("SAP BTP Cloud Foundry Runtime",        "BTP",  "paas"),
    "SAP Enable Now":                                   ("SAP Enable Now",                       "Other","saas"),
    "SAP Business Technology Platform":                 ("SAP Business Technology Platform",     "BTP",  "paas"),
    "SAP Business Technology Platform Subaccount":      ("SAP BTP Subaccount",                  "BTP",  "paas"),
    "Cloud Foundry Organization":                       ("SAP BTP CF Organization",              "BTP",  "paas"),
    "SAP Audit Log service":                            ("SAP Audit Log Service",                "BTP",  "saas"),
    "SAP HANA Cloud":                                   ("SAP HANA Cloud",                      "Data", "paas"),
    "SAP SuccessFactors HCM":                           ("SAP SuccessFactors HCM",              "HCM",  "saas"),
    "SAP Cloud Integration":                            ("SAP Integration Suite",               "BTP",  "saas"),
    "SAP Cloud Portal service":                         ("SAP Cloud Portal Service",             "BTP",  "saas"),
    "SAP Jam Collaboration":                            ("SAP Jam Collaboration",               "Other","saas"),
    "SAP Information Capture, Core by OpenText":        ("SAP Information Capture by OpenText", "Other","saas"),
}

CLOUD_ROLE_TO_ITC: dict[str, str] = {
    "SAP Analytics Cloud":             "SAP Analytics Cloud",
    "SAP Datasphere":                  "SAP Datasphere",
    "SAP Datasphere, SAP BW Bridge":   "SAP Datasphere;SAP BW Bridge",
    "SAP Ariba Procurement":           "SAP Ariba;SAP Integration Suite",
    "SAP Ariba Sourcing":              "SAP Ariba;SAP Integration Suite",
    "SAP Ariba Shopping":              "SAP Ariba",
    "SAP Business Network":            "SAP Business Network",
    "SAP Build Work Zone, standard edition": "SAP BTP",
    "SAP BTP, Cloud Foundry runtime":  "SAP BTP",
    "SAP HANA Cloud":                  "SAP HANA Cloud",
    "SAP SuccessFactors HCM":          "SAP SuccessFactors;SAP Integration Suite",
    "Identity Authentication":         "SAP Identity Authentication Service",
    "Identity Provisioning":           "SAP Identity Authentication Service",
    "Joule":                           "SAP Joule",
    "SAP Enable Now":                  "SAP Enable Now",
    "SAP Jam Collaboration":           "SAP Jam Collaboration",
    "SAP Information Capture, Core by OpenText": "SAP Information Capture by OpenText",
}

CLOUD_ROLE_TO_BC: dict[str, str] = {
    "SAP Analytics Cloud":           "Finance / Financial Planning and Analysis",
    "SAP Datasphere":                "Finance / Financial Planning and Analysis",
    "SAP Datasphere, SAP BW Bridge": "Finance / Financial Planning and Analysis",
    "SAP Ariba Procurement":         "Sourcing and Procurement / Operational Procurement;Sourcing and Procurement / Procurement Contract Management",
    "SAP Ariba Sourcing":            "Sourcing and Procurement / Procurement Contract Management;Sourcing and Procurement / Supplier Management",
    "SAP Ariba Shopping":            "Sourcing and Procurement / Operational Procurement",
    "SAP Business Network":          "Sourcing and Procurement / Supplier Management",
    "SAP SuccessFactors HCM":        "",
    "Joule":                         "",
    "SAP Enable Now":                "",
}


def _criticality(install: str, _sol_area: str = "") -> str:
    if "INFRA" in install or "S/4HANA" in install:
        return "missionCritical"
    if "ERP" in install or "SOLMAN" in install:
        return "businessCritical"
    return "businessOperational"


# ── Header lookup helpers ──────────────────────────────────────────────────────

def _header_index(ws, aliases: list[str]) -> int | None:
    """Return the 1-based column index of the first header matching any alias
    (case-insensitive, exact match). Returns None if no header matches."""
    header_row = {
        str(ws.cell(1, c).value).strip().lower(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }
    for a in aliases:
        idx = header_row.get(a.strip().lower())
        if idx is not None:
            return idx
    return None


def _row_dict(ws, r: int, cols: dict[str, int | None]) -> dict[str, Any]:
    """Build a dict {logical_name: cell_value} using pre-resolved column indices."""
    return {
        name: (ws.cell(r, idx).value if idx is not None else None)
        for name, idx in cols.items()
    }


# ── On-premise extraction ──────────────────────────────────────────────────────

# Aliases: first match wins. Ordered so newer/wider export headers are preferred.
_ONPREM_HEADERS = {
    "sid":         ["System ID", "SID"],
    "install":     ["Installation Name", "Install"],
    "system_role": ["System Role", "Business Type"],
    "product":     ["Product Line Description", "Product"],
    "version":     ["Product Version", "Version"],
    "database":    ["Data Base", "Database"],
}


def _read_onprem(path: Path, client_name: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path)
    ws = wb["sheet1"]

    cols = {k: _header_index(ws, aliases) for k, aliases in _ONPREM_HEADERS.items()}
    missing = [k for k, v in cols.items() if v is None and k in ("sid", "install", "product")]
    if missing:
        logger.warning("OnPrem: missing required columns %s in %s", missing, path.name)
        return []

    sid_groups: dict[tuple, list[dict[str, Any]]] = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        row = _row_dict(ws, r, cols)
        if not any(row.values()):
            continue
        # Filter productive systems. Older exports had "Productive"/"Non-Productive"
        # in a dedicated column; if missing, accept all rows.
        role = row.get("system_role")
        if role is not None and str(role).strip() != "Productive":
            continue
        if not row["sid"] or not row["install"]:
            continue
        sid_groups[(row["sid"], row["install"])].append(row)

    seen: set[str] = set()
    apps: list[dict] = []

    for (sid, install), rows in sid_groups.items():
        rows_sorted = sorted(
            rows, key=lambda r: PRODUCT_PRIORITY.get(str(r.get("product") or "").upper(), 99)
        )
        primary = rows_sorted[0]
        install_clean = str(install).strip()
        primary_product = str(primary.get("product") or "").upper()

        name_map = {
            "SAP ERP (HR)":    f"SAP ERP HR ({sid})",
            "SAP ERP (INFRA)": f"SAP S/4HANA ({sid})" if "S/4HANA" in primary_product else f"SAP ERP ({sid})",
            "SAP ERP":         f"SAP ERP ({sid})",
            "SAP NW":          f"SAP NetWeaver ({sid})",
            "SOLMAN":          f"SAP Solution Manager ({sid})",
        }
        name = name_map.get(install_clean, f"{install_clean} ({sid})")
        if name in seen:
            continue
        seen.add(name)

        all_products = {
            PRODUCT_TO_ITC[p]
            for r in rows
            if (p := str(r.get("product") or "").upper()) and p in PRODUCT_TO_ITC
        }
        itc = ";".join(sorted(all_products))
        bc  = PRODUCT_TO_BC.get(primary_product, "")

        version = primary.get("version") or ""
        lifecycle = "phaseOut" if primary_product == "SAP ERP" or "ERP 6" in str(version) else "active"
        desc = (
            f"On-premise SAP system. SID: {sid}. Install: {install_clean}. "
            f"Product: {primary.get('product') or ''}. Version: {version}."
        )
        if primary.get("database"):
            desc += f" Database: {primary['database']}."

        apps.append({
            "id": "", "type": "Application", "name": name, "description": desc,
            "alias": sid, "externalId": sid,
            "lifecycle_phase": lifecycle, "lifecycle_startDate": "", "lifecycle_endDate": "",
            "businessCriticality": _criticality(install_clean),
            "functionalSuitability": "", "technicalSuitability": "",
            "lxHostingType": "onPremise", "lxState": "DRAFT",
            "tags": f"Baseline;OnPremise;{client_name}",
            "relApplicationToITComponent": itc,
            "relApplicationToBusinessCapability": bc,
            "relToParent": "",
            "_source": "onprem",
        })

    logger.info("OnPrem: %d applications from %s", len(apps), path.name)
    return apps


# ── Cloud extraction ───────────────────────────────────────────────────────────

_CLOUD_HEADERS = {
    "business_type":    ["Business Type"],
    "lifecycle_status": ["Lifecycle Status"],
    "role":             ["System Role", "Tenant Role"],
    "sol_area":         ["Solution Area"],
    "sub_sol_area":     ["Sub-Solution Area"],
    "external_id":      ["External ID", "External Name"],
    "data_center":      ["Data Center Description", "Data Center External Description"],
}


def _is_productive_business_type(v: Any) -> bool:
    if not v:
        return False
    s = str(v).strip().lower()
    # Old exports used "Productive"; new ones use "Production Tenant" / "Parent Production Tenant".
    return "product" in s and "test" not in s


def _read_cloud(path: Path, client_name: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path)
    ws = wb["sheet1"]

    cols = {k: _header_index(ws, aliases) for k, aliases in _CLOUD_HEADERS.items()}
    if cols["role"] is None:
        logger.warning("Cloud: missing 'System Role' column in %s", path.name)
        return []

    seen: set[str] = set()
    apps: list[dict] = []

    for r in range(2, ws.max_row + 1):
        row = _row_dict(ws, r, cols)
        if not any(row.values()):
            continue
        # Filters: only live production tenants. If a column is missing, skip that filter.
        if cols["lifecycle_status"] is not None and str(row.get("lifecycle_status") or "").strip() != "Live":
            continue
        if cols["business_type"] is not None and not _is_productive_business_type(row.get("business_type")):
            continue

        role = row.get("role")
        if not role or role in seen:
            continue
        seen.add(role)

        if role in CLOUD_ROLE_TO_APP:
            app_name, _area, hosting = CLOUD_ROLE_TO_APP[role]
        else:
            app_name, hosting = role, "saas"

        sol_area = row.get("sol_area") or ""
        data_center = row.get("data_center") or ""
        sub_area = row.get("sub_sol_area") or ""
        desc = f"Cloud SAP application. Role: {role}. Solution area: {sol_area}. Data center: {data_center}."
        if sub_area:
            desc += f" Sub-area: {str(sub_area)[:80]}."

        apps.append({
            "id": "", "type": "Application", "name": app_name, "description": desc,
            "alias": "", "externalId": row.get("external_id") or "",
            "lifecycle_phase": "active", "lifecycle_startDate": "", "lifecycle_endDate": "",
            "businessCriticality": "businessCritical",
            "functionalSuitability": "", "technicalSuitability": "",
            "lxHostingType": hosting, "lxState": "DRAFT",
            "tags": f"Baseline;Cloud;{str(sol_area).replace(' ', '_')};{client_name}",
            "relApplicationToITComponent": CLOUD_ROLE_TO_ITC.get(role, ""),
            "relApplicationToBusinessCapability": CLOUD_ROLE_TO_BC.get(role, ""),
            "relToParent": "",
            "_source": "cloud",
        })

    logger.info("Cloud: %d applications from %s", len(apps), path.name)
    return apps


# ── Excel writer ───────────────────────────────────────────────────────────────

def _write_excel(apps: list[dict], output_path: Path, client_name: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    for col_idx, key in enumerate(COLUMNS, start=1):
        cat = CATEGORIES[key]
        c1 = ws.cell(row=1, column=col_idx, value=key)
        c1.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c1.fill = PatternFill("solid", fgColor=HEADER_COLORS[cat])
        c1.alignment = Alignment(horizontal="left", vertical="center")

        c2 = ws.cell(row=2, column=col_idx, value=TRANSLATIONS[key])
        c2.font = Font(name="Calibri", bold=True, color="223548", size=9)
        c2.fill = PatternFill("solid", fgColor=COLOR_TRANS_ROW)
        c2.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, app in enumerate(apps, start=3):
        bg = ROW_ONPREM if app["_source"] == "onprem" else ROW_CLOUD
        for col_idx, key in enumerate(COLUMNS, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=app.get(key, ""))
            c.font = Font(name="Calibri", size=9, color="223548")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, key in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(key, 20)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for r in range(3, len(apps) + 3):
        ws.row_dimensions[r].height = 16
    ws.freeze_panes = "D3"

    n_onprem = sum(1 for a in apps if a["_source"] == "onprem")
    n_cloud  = sum(1 for a in apps if a["_source"] == "cloud")

    readme = wb.create_sheet("ReadMe")
    readme_rows = [
        (f"LeanIX Import — Applications Baseline ({client_name})",),
        (f"Source: OnPrem Systems.xlsx + Cloud Systems.xlsx",),
        ("",),
        ("COLOR LEGEND",),
        ("Green rows = On-premise systems",),
        ("Blue rows  = Cloud systems",),
        ("",),
        ("IMPORT RULES",),
        ("- Leave 'id' column EMPTY — new fact sheets will be created.",),
        ("- Relations use EXACT display names of existing LeanIX fact sheets.",),
        ("- Multiple values separated by semicolon (;) without spaces.",),
        ("- 'lxState' = DRAFT — approve manually after review.",),
        ("- Lifecycle 'phaseOut' applied to SAP ERP 6.0 systems.",),
        ("- Import via: Inventory > Inventory Tools > Import from Excel",),
        ("- Order: Organization → BusinessCapability → Application → Initiative → ITComponent",),
        ("",),
        ("STATS",),
        (f"Total applications: {len(apps)}",),
        (f"On-premise (green): {n_onprem}",),
        (f"Cloud (blue): {n_cloud}",),
    ]
    for r_idx, row in enumerate(readme_rows, start=1):
        c = readme.cell(row=r_idx, column=1, value=row[0])
        if r_idx == 1:
            c.font = Font(name="Calibri", size=13, bold=True, color="002A86")
        elif row[0] in ("COLOR LEGEND", "IMPORT RULES", "STATS"):
            c.font = Font(name="Calibri", size=10, bold=True, color="002A86")
        else:
            c.font = Font(name="Calibri", size=9)
        c.alignment = Alignment(horizontal="left", wrap_text=True)
    readme.column_dimensions["A"].width = 80

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Baseline saved: %s (%d apps)", output_path, len(apps))


# ── Public API ─────────────────────────────────────────────────────────────────

def generate_baseline(
    output_path: str | Path,
    client_name: str,
    onprem_path: str | Path | None = None,
    cloud_path:  str | Path | None = None,
) -> dict[str, Any]:
    """
    Generate the Baseline (AS-IS) LeanIX import Excel.

    Args:
        output_path:  Where to write the output Excel.
        client_name:  Client name used as tag in LeanIX.
        onprem_path:  Path to OnPrem Systems.xlsx (optional).
        cloud_path:   Path to Cloud Systems.xlsx (optional).

    Returns:
        dict with keys: output_path, n_onprem, n_cloud, n_total
    """
    apps: list[dict] = []

    if onprem_path:
        apps.extend(_read_onprem(Path(onprem_path), client_name))
    if cloud_path:
        apps.extend(_read_cloud(Path(cloud_path), client_name))

    if not apps:
        logger.warning("No applications found — check input files and filters.")
        return {"output_path": None, "n_onprem": 0, "n_cloud": 0, "n_total": 0}

    _write_excel(apps, Path(output_path), client_name)

    return {
        "output_path": str(output_path),
        "n_onprem":    sum(1 for a in apps if a["_source"] == "onprem"),
        "n_cloud":     sum(1 for a in apps if a["_source"] == "cloud"),
        "n_total":     len(apps),
    }
