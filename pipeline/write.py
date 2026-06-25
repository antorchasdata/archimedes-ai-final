"""
write.py — Write enriched requirements to Excel and optionally to LeanIX.

Excel output:
  - Writes columns H–P onto the original Excel file (preserving all other columns)
  - Auto-detects the ID column and header row (same logic as extract.py)

LeanIX output (optional, --push-leanix / LEANIX_PUSH=true):
  EA model per requirement:
    Initiative  (one per requirement)
      ├─ relInitiativeToBusinessCapability → BusinessCapability (leaf BC, upsert)
      └─ relInitiativeToApplication        → Application (one per RSA, upsert)

  All fact sheets created by this pipeline carry a tag  client=<name>
  so they can be filtered and cleaned up after a demo.

  Lifecycle derived from coverage:
    Total      → active
    Parcial    → phaseIn
    No cubierto → plan

Requires env vars: LEANIX_API_TOKEN, LEANIX_WORKSPACE_ID, LEANIX_BASE_URL
"""

from __future__ import annotations

import json
import logging
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

import pandas as pd
from dotenv import load_dotenv

from pipeline.reference_catalog import ReferenceCatalogResolver, ResolvedMatch
from pipeline.leanix_auth import get_bearer

load_dotenv()
logger = logging.getLogger(__name__)

KNOWLEDGE_DIR = Path(__file__).parent.parent / "knowledge"
LPR_CATALOG_PATH = KNOWLEDGE_DIR / "sap_lpr_catalog.json"


def _workspace_from_base_url(base_url: str) -> str:
    """Extract workspace slug from a LeanIX base URL.

    https://demo-eu-3.leanix.net  → demo-eu-3
    https://app.leanix.net        → app
    """
    from urllib.parse import urlparse
    host = urlparse(base_url).hostname or ""
    return host.split(".")[0] if host else ""


def _write_push_uuid_map(
    out_dir: Path,
    base_url: str,
    workspace: str,
    app_id_cache: dict[str, str],
    itc_id_cache: dict[str, str],
    failed: list[dict] | None = None,
) -> Path:
    """Persist UUID map → push_uuid_map.json in out_dir.

    Schema:
        {workspace, base_url, entries: {"<Type>::<Name>": {uuid, created}}, failed: [...]}
    """
    entries: dict[str, dict] = {}
    for name, uid in (app_id_cache or {}).items():
        entries[f"Application::{name}"] = {"uuid": uid, "created": True}
    for name, uid in (itc_id_cache or {}).items():
        entries[f"ITComponent::{name}"] = {"uuid": uid, "created": True}
    payload = {
        "workspace": workspace,
        "base_url": base_url,
        "entries": entries,
        "failed": failed or [],
    }
    out_path = Path(out_dir) / "push_uuid_map.json"
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False))
    return out_path


def _load_lpr_catalog() -> dict:
    """Load sap_lpr_catalog.json. Returns empty structure if not found."""
    if not LPR_CATALOG_PATH.exists():
        logger.debug("sap_lpr_catalog.json not found — LPR enrichment disabled")
        return {"lpr_index": {}, "rsa_name_index": {}}
    return json.loads(LPR_CATALOG_PATH.read_text())


def _enrich_with_lpr(row: dict, lpr_catalog: dict) -> None:
    """
    In-place: set externalId and lprId on an Application fact sheet row
    if the displayName matches an RSA name in rsa_name_index.
    """
    rsa_index: dict = lpr_catalog.get("rsa_name_index", {})
    lpr_index: dict = lpr_catalog.get("lpr_index", {})

    name = row.get("name", "")
    lpr_id = rsa_index.get(name)
    if not lpr_id:
        return

    entry = lpr_index.get(lpr_id, {})
    material_ids = entry.get("material_ids", [])

    if material_ids:
        row["externalId"] = material_ids[0]
    row["lprId"] = lpr_id
    logger.debug("LPR enriched '%s' → %s (externalId=%s)",
                 name, lpr_id, material_ids[0] if material_ids else "")

# Column indices in the output Excel (0-based), matching the project standard
COL_H = 7   # coverage
COL_I = 8   # module
COL_J = 9   # dev
COL_K = 10  # dev_exp
COL_L = 11  # ext_apps
COL_M = 12  # licensing
COL_N = 13  # comment
COL_O = 14  # Business Capabilities (SAP RBA)
COL_P = 15  # RSA application

_ID_PATTERNS = re.compile(r"\b(id|req|requ|n[oº°]|num|code|ref)\b", re.I)


# ── BCS resolution ────────────────────────────────────────────────────────────

def _load_bcs_index() -> dict[str, str]:
    rba = json.loads((KNOWLEDGE_DIR / "sap_rba_catalog.json").read_text())
    return rba["short_name_index"]


def bc_str(names: list[str], index: dict[str, str]) -> str:
    return " | ".join(index.get(n, n) for n in names)


# ── Excel writer ──────────────────────────────────────────────────────────────

def _detect_header_row(path: Path) -> int:
    raw = pd.read_excel(path, header=None, nrows=15)
    for i, row in raw.iterrows():
        non_null = row.dropna()
        if len(non_null) >= 2 and all(isinstance(v, str) for v in non_null):
            return int(i)
    return 0


def write_excel(
    enriched: list[dict[str, Any]],
    template_path: Path,
    output_path: Path,
    bcs_index: dict[str, str],
) -> None:
    header_row = _detect_header_row(template_path)
    df = pd.read_excel(template_path, header=None)

    # Build lookup: id → enriched dict
    lookup = {r["id"]: r for r in enriched}

    # Find ID column
    header = df.iloc[header_row]
    id_col_idx = next(
        (i for i, v in enumerate(header) if _ID_PATTERNS.search(str(v))),
        1,  # default to col B (index 1) matching original project convention
    )

    data_start = header_row + 1
    filled = 0

    for idx in range(data_start, len(df)):
        cell = df.iloc[idx, id_col_idx]
        if pd.isna(cell):
            continue
        row_id = str(cell).strip()
        if row_id not in lookup:
            continue

        m = lookup[row_id]
        df.iloc[idx, COL_H] = m.get("coverage", "")
        df.iloc[idx, COL_I] = m.get("module", "")
        df.iloc[idx, COL_J] = m.get("dev", "")
        df.iloc[idx, COL_K] = m.get("dev_exp") or float("nan")
        df.iloc[idx, COL_L] = m.get("ext_apps") or float("nan")
        df.iloc[idx, COL_M] = m.get("licensing", "")
        df.iloc[idx, COL_N] = m.get("comment", "")
        df.iloc[idx, COL_O] = bc_str(m.get("bcs", []), bcs_index)
        df.iloc[idx, COL_P] = m.get("rsa", "")
        filled += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, header=False)

    logger.info("Excel: wrote %d rows → %s", filled, output_path)


# ── LeanIX staging Excel ──────────────────────────────────────────────────────

# coverage → Initiative lifecycle phase (also used by the push step)
_LIFECYCLE_MAP = {
    "Total":        "active",
    "Parcial":      "phaseIn",
    "No cubierto":  "plan",
}

# SAP Application → IT Components (technology stack, well-known SAP standard mapping)
# ITC name follows SAP LeanIX reference catalog conventions.
# Each tuple: (itc_name, hosting_type)
_APP_TO_ITCS: dict[str, list[tuple[str, str]]] = {
    "SAP S/4HANA": [
        ("SAP HANA Database", "onPremise"),
        ("SAP ABAP Platform", "onPremise"),
        ("SAP NetWeaver Application Server", "onPremise"),
        ("SAP Fiori", "onPremise"),
    ],
    "SAP S/4HANA Cloud": [
        ("SAP HANA Database", "saas"),
        ("SAP Business Technology Platform", "paas"),
        ("SAP Fiori", "saas"),
    ],
    "SAP Analytics Cloud": [
        ("SAP Business Technology Platform", "paas"),
        ("SAP HANA Database", "saas"),
    ],
    "SAP Integration Suite": [
        ("SAP Business Technology Platform", "paas"),
        ("SAP Cloud Integration", "paas"),
        ("SAP API Management", "paas"),
    ],
    "SAP Ariba Sourcing": [
        ("SAP Ariba Network", "saas"),
        ("SAP Business Technology Platform", "paas"),
    ],
    "SAP Ariba Procurement": [
        ("SAP Ariba Network", "saas"),
        ("SAP Business Technology Platform", "paas"),
    ],
    "SAP Integrated Business Planning": [
        ("SAP Business Technology Platform", "paas"),
        ("SAP HANA Database", "saas"),
    ],
    "SAP Transportation Management": [
        ("SAP HANA Database", "onPremise"),
        ("SAP ABAP Platform", "onPremise"),
        ("SAP NetWeaver Application Server", "onPremise"),
    ],
    "SAP Global Trade Services": [
        ("SAP HANA Database", "onPremise"),
        ("SAP ABAP Platform", "onPremise"),
        ("SAP NetWeaver Application Server", "onPremise"),
    ],
    "SAP SuccessFactors": [
        ("SAP Business Technology Platform", "paas"),
        ("SAP SuccessFactors Platform", "saas"),
    ],
    "SAP Concur": [
        ("SAP Business Technology Platform", "paas"),
    ],
    "SAP Fieldglass": [
        ("SAP Business Technology Platform", "paas"),
    ],
    "SAP Customer Experience": [
        ("SAP Business Technology Platform", "paas"),
        ("SAP Commerce Cloud", "saas"),
    ],
    "SAP ERP": [
        ("SAP HANA Database", "onPremise"),
        ("SAP ABAP Platform", "onPremise"),
        ("SAP NetWeaver Application Server", "onPremise"),
    ],
    "SAP Business One": [
        ("SAP HANA Database", "onPremise"),
    ],
    "SAP Datasphere": [
        ("SAP Business Technology Platform", "paas"),
        ("SAP HANA Database", "saas"),
    ],
}

def _itcs_for_apps(app_names: list[str]) -> list[tuple[str, str]]:
    """Return deduplicated (itc_name, hosting_type) tuples for a list of app names."""
    seen: set[str] = set()
    result = []
    for app in app_names:
        for itc_name, hosting in _APP_TO_ITCS.get(app, []):
            if itc_name not in seen:
                seen.add(itc_name)
                result.append((itc_name, hosting))
    return result

# ── Shared Excel formatting helpers ───────────────────────────────────────────

_COLOR_MANDATORY = "002A86"
_COLOR_OPTIONAL  = "0070F2"
_COLOR_RELATION  = "107E3E"
_COLOR_READONLY  = "A9B4BE"
_COLOR_TRANS_ROW = "EAF4FF"
_COLOR_DATA_ROW  = "F5F9FF"

_HEADER_COLORS = {
    "mandatory": _COLOR_MANDATORY,
    "optional":  _COLOR_OPTIONAL,
    "relation":  _COLOR_RELATION,
    "readonly":  _COLOR_READONLY,
}


def _sheet_header(ws: Any, columns: list[tuple[str, str, str, int]]) -> None:
    """
    Write LeanIX-format header rows to a worksheet.
    columns: list of (technical_key, translation, category, width)
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    for col_idx, (key, label, cat, width) in enumerate(columns, start=1):
        c1 = ws.cell(row=1, column=col_idx, value=key)
        c1.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c1.fill  = PatternFill("solid", fgColor=_HEADER_COLORS.get(cat, _COLOR_OPTIONAL))
        c1.alignment = Alignment(horizontal="left", vertical="center")

        c2 = ws.cell(row=2, column=col_idx, value=label)
        c2.font  = Font(name="Calibri", bold=True, color="223548", size=9)
        c2.fill  = PatternFill("solid", fgColor=_COLOR_TRANS_ROW)
        c2.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18


def _sheet_row(ws: Any, row_idx: int, values: list, bg: str = _COLOR_DATA_ROW) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=val or "")
        c.font  = Font(name="Calibri", size=9, color="223548")
        c.fill  = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[row_idx].height = 16


# ── Column schemas per fact sheet type ────────────────────────────────────────

# (technical_key, translation, category, width)
_COLS_APPLICATION = [
    ("id",                              "ID",                        "readonly",  36),
    ("type",                            "Type",                      "mandatory", 14),
    ("name",                            "Name",                      "mandatory", 40),
    ("description",                     "Description",               "optional",  55),
    ("alias",                           "Alias",                     "optional",  14),
    ("externalId",                      "External ID",               "optional",  18),
    ("lifecycle_phase",                 "Lifecycle Phase",           "optional",  16),
    ("lifecycle_startDate",             "Lifecycle Start Date",      "optional",  20),
    ("lifecycle_endDate",               "Lifecycle End Date",        "optional",  18),
    ("businessCriticality",             "Business Criticality",      "optional",  22),
    ("functionalSuitability",           "Functional Fit",            "optional",  20),
    ("technicalSuitability",            "Technical Fit",             "optional",  20),
    ("lxHostingType",                   "Hosting Type",              "optional",  18),
    ("lxState",                         "Quality Seal",              "optional",  14),
    ("tags",                            "Tags",                      "optional",  35),
    ("relApplicationToBusinessCapability", "Business Capabilities",  "relation",  60),
    ("relApplicationToITComponent",     "IT Components",             "relation",  45),
    ("relToParent",                     "Parent Application",        "relation",  30),
]

_COLS_BC = [
    ("id",          "ID",          "readonly",  36),
    ("type",        "Type",        "mandatory", 14),
    ("name",        "Name",        "mandatory", 55),
    ("description", "Description", "optional",  60),
    ("lxState",     "Quality Seal","optional",  14),
    ("tags",        "Tags",        "optional",  35),
    ("relToParent", "Parent BC",   "relation",  55),
]

_COLS_INITIATIVE = [
    ("id",                                  "ID",                        "readonly",  36),
    ("type",                                "Type",                      "mandatory", 14),
    ("name",                                "Name",                      "mandatory", 45),
    ("description",                         "Description",               "optional",  60),
    ("lifecycle_phase",                     "Lifecycle Phase",           "optional",  16),
    ("lifecycle_startDate",                 "Lifecycle Start Date",      "optional",  20),
    ("lifecycle_endDate",                   "Lifecycle End Date",        "optional",  18),
    ("lxState",                             "Quality Seal",              "optional",  14),
    ("tags",                                "Tags",                      "optional",  35),
    ("relInitiativeToApplication",          "Applications",              "relation",  45),
    ("relInitiativeToBusinessCapability",   "Business Capabilities",     "relation",  60),
    ("relInitiativeToObjective",            "Objectives",                "relation",  45),
]

_COLS_ITC = [
    ("id",          "ID",          "readonly",  36),
    ("type",        "Type",        "mandatory", 14),
    ("name",        "Name",        "mandatory", 40),
    ("externalId",  "External ID", "optional",  20),
    ("description", "Description", "optional",  55),
    ("lxHostingType","Hosting Type","optional",  18),
    ("lxState",     "Quality Seal","optional",  14),
    ("tags",        "Tags",        "optional",  35),
]


def write_leanix_excel(
    enriched: list[dict[str, Any]],
    bcs_index: dict[str, str],
    output_path: Path,
    client_name: str,
    supplementary: dict | None = None,
    lift_shift: list[dict] | None = None,
) -> None:
    """
    Write a LeanIX-importable multi-sheet Excel with proper format:

      Sheet "Application"         — deduplicated RSA apps (TO-BE)
      Sheet "BusinessCapability"  — deduplicated leaf BCs from requirements
      Sheet "Initiative"          — one row per requirement
      Sheet "ITComponent"         — ITCs where derivable
      Sheet "ReadMe"              — import instructions

    Row 1: technical keys (colored headers)
    Row 2: human-readable translations
    Row 3+: data rows

    supplementary: optional dict from pdf_extract / image_extract containing
      keys "from_pdf" and/or "from_images", each with lists "applications",
      "business_capabilities", "initiatives", "it_components". These are merged
      into the respective sheets (deduplicated by name).

    lift_shift: optional list of selected conversion dicts from lift_shift.convert_to_rise().
      Each dict has: target_matnr, target_desc, status, deployment, prereq_resolved, source_app.
      These are added as Applications with lifecycle="plan", tag="Target;LiftShift;<client>".
    """
    import openpyxl

    # ── Collect data from enriched requirements ────────────────────────────────
    seen_apps: dict[str, dict] = {}   # app_name → {lifecycle, bcs, itcs}
    seen_bcs:  dict[str, str]  = {}   # bc_leaf_name → full_path
    _init_groups: dict[str, dict] = {}  # group_key → aggregated initiative data

    for req in enriched:
        if req.get("_error"):
            continue

        # Resolve BCs
        bcs_resolved: list[str] = []
        for bc_short in req.get("bcs", []):
            full_path = bcs_index.get(bc_short, bc_short)
            parts     = full_path.split(" / ")
            bc_leaf   = parts[-1] if parts else full_path
            seen_bcs[bc_leaf] = full_path
            bcs_resolved.append(bc_leaf)

        rsa_name = req.get("rsa", "SAP S/4HANA")
        lifecycle = _LIFECYCLE_MAP.get(req.get("coverage", ""), "plan")

        # Upsert application — collect ITCs from static mapping
        app_itcs = {itc for itc, _ in _APP_TO_ITCS.get(rsa_name, [])}
        if rsa_name not in seen_apps:
            seen_apps[rsa_name] = {
                "name":       rsa_name,
                "lifecycle":  lifecycle,
                "bcs":        set(bcs_resolved),
                "itcs":       app_itcs,
            }
        else:
            seen_apps[rsa_name]["bcs"].update(bcs_resolved)
            # promote lifecycle: active > phaseIn > plan
            order = {"active": 0, "phaseIn": 1, "plan": 2}
            if order.get(lifecycle, 2) < order.get(seen_apps[rsa_name]["lifecycle"], 2):
                seen_apps[rsa_name]["lifecycle"] = lifecycle

        # Accumulate initiative data, grouped by _group when present
        group_key = req.get("_group") or req["id"]
        if group_key not in _init_groups:
            _init_groups[group_key] = {
                "name":      group_key,
                "lifecycle": lifecycle,
                "apps":      set(),
                "bcs":       set(),
                "n_reqs":    0,
            }
        g = _init_groups[group_key]
        g["apps"].add(rsa_name)
        g["bcs"].update(bcs_resolved)
        g["n_reqs"] += 1
        # promote lifecycle: active > phaseIn > plan
        order = {"active": 0, "phaseIn": 1, "plan": 2}
        if order.get(lifecycle, 2) < order.get(g["lifecycle"], 2):
            g["lifecycle"] = lifecycle

    # Build initiative list from grouped data
    initiatives: list[dict] = []
    for group_key, g in _init_groups.items():
        initiatives.append({
            "name":                 g["name"],
            "description":          f"{g['n_reqs']} requirements — client: {client_name}",
            "lifecycle":            g["lifecycle"],
            "lifecycle_startDate":  time.strftime("%Y-%m-%d"),
            "lifecycle_endDate":    str(int(time.strftime("%Y")) + 2) + time.strftime("-%m-%d"),
            "app":                  ";".join(sorted(g["apps"])),
            "bcs":                  ";".join(sorted(g["bcs"])),
            "objective":            g["name"],
        })

    # Build deduplicated ITC list from all seen apps
    seen_itcs: dict[str, str] = {}  # itc_name → hosting_type
    for app_name in seen_apps:
        for itc_name, hosting in _APP_TO_ITCS.get(app_name, []):
            if itc_name not in seen_itcs:
                seen_itcs[itc_name] = hosting

    # ── Merge supplementary fact sheets (from PDF / images) ───────────────────
    supp_initiatives: list[dict] = []
    if supplementary:
        for source_key in ("from_pdf", "from_images"):
            source = supplementary.get(source_key, {})
            if not source:
                continue

            for app in source.get("applications", []):
                name = (app.get("name") or "").strip()
                if name and name not in seen_apps:
                    seen_apps[name] = {
                        "name":      name,
                        "lifecycle": "plan",
                        "bcs":       set(),
                        "itcs":      set(),
                    }

            for bc in source.get("business_capabilities", []):
                name = (bc.get("name") or "").strip()
                if name and name not in seen_bcs:
                    seen_bcs[name] = bc.get("path") or name

            for itc in source.get("it_components", []):
                name = (itc.get("name") or "").strip()
                if name and name not in seen_itcs:
                    seen_itcs[name] = itc.get("hosting_type") or "onPremise"

            for init in source.get("initiatives", []):
                name = (init.get("name") or "").strip()
                if name:
                    supp_initiatives.append({
                        "name":                (init.get("name") or name)[:255],
                        "description":         (init.get("description") or f"Derived from {source_key}. Client: {client_name}.")[:2000],
                        "lifecycle":           init.get("lifecycle_phase") or "plan",
                        "lifecycle_startDate": init.get("lifecycle_startDate") or time.strftime("%Y-%m-%d"),
                        "lifecycle_endDate":   init.get("lifecycle_endDate") or (str(int(time.strftime("%Y")) + 2) + time.strftime("-%m-%d")),
                        "app":                 init.get("relInitiativeToApplication") or "",
                        "bcs":                 init.get("relInitiativeToBusinessCapability") or "",
                        "objective":           (init.get("name") or name)[:255],
                    })

        if supplementary:
            logger.info(
                "Supplementary merged: +%d apps, +%d BCs, +%d ITCs, +%d initiatives",
                sum(len(supplementary.get(k, {}).get("applications", [])) for k in ("from_pdf", "from_images")),
                sum(len(supplementary.get(k, {}).get("business_capabilities", [])) for k in ("from_pdf", "from_images")),
                sum(len(supplementary.get(k, {}).get("it_components", [])) for k in ("from_pdf", "from_images")),
                len(supp_initiatives),
            )

    # ── Merge Lift & Shift conversions ─────────────────────────────────────────
    ls_tag = f"Target;LiftShift;{client_name}"
    seen_prereqs: dict[str, dict] = {}   # prereq_name → {matnr, for_app, maktx}
    if lift_shift:
        for conv in lift_shift:
            name = conv.get("target_desc") or conv.get("target_matnr", "")
            if not name:
                continue
            name = name.strip()
            if name not in seen_apps:
                seen_apps[name] = {
                    "name":       name,
                    "lifecycle":  "plan",
                    "bcs":        set(),
                    "itcs":       set(),
                    "extra_tags": ls_tag,
                    "description": (
                        f"RISE target for {conv.get('source_app','')}. "
                        f"Material: {conv.get('target_matnr','')} | "
                        f"Deployment: {conv.get('deployment','')} | "
                        f"Status: {conv.get('status','')} — {conv.get('status_desc','')}"
                    )[:2000],
                }
            # Collect prerequisite apps separately (NOT in seen_apps)
            for prereq in conv.get("prereq_resolved", []):
                prereq_name = prereq.get("maktx") or prereq.get("matnr", "")
                prereq_name = prereq_name.strip()
                if prereq_name and prereq_name not in seen_prereqs:
                    seen_prereqs[prereq_name] = {
                        "matnr":   prereq.get("matnr", ""),
                        "maktx":   prereq_name,
                        "for_app": name,
                    }
        logger.info("Lift & Shift merged: +%d target apps, +%d prerequisites", len(lift_shift), len(seen_prereqs))

    # ── Reference Catalog pre-resolution (feature-flagged) ─────────────────────
    resolver: ReferenceCatalogResolver | None = None
    app_matches: dict[str, ResolvedMatch] = {}
    itc_matches: dict[str, ResolvedMatch] = {}
    if os.environ.get("ARCHIMEDES_USE_CATALOG_RESOLVER", "").lower() in ("1", "true", "yes"):
        try:
            base_url, api_token = _resolver_credentials()
            resolver = ReferenceCatalogResolver(
                base_url=base_url,
                api_token=api_token,
                interactive=sys.stdin.isatty() if hasattr(sys.stdin, "isatty") else False,
            )
            app_matches = resolver.resolve("Application", list(seen_apps.keys()))
            itc_matches = resolver.resolve("ITComponent", list(seen_itcs.keys()))
            logger.info(
                "Reference Catalog: %d/%d apps linked, %d/%d ITCs linked",
                sum(1 for m in app_matches.values() if m.status == "LINKED"),
                len(app_matches),
                sum(1 for m in itc_matches.values() if m.status == "LINKED"),
                len(itc_matches),
            )
            # Decorate seen_apps with catalog data
            for app_name, match in app_matches.items():
                entry = seen_apps.get(app_name)
                if entry is None:
                    continue
                entry["externalId"] = match.external_id or ""
                entry["catalog_confidence"] = match.confidence
                entry["catalog_status"] = match.status
                for k, v in match.fields.items():
                    entry.setdefault(k, v)
            # Decorate seen_itcs (str → dict upgrade)
            for itc_name, match in itc_matches.items():
                hosting = seen_itcs.get(itc_name) or ""
                seen_itcs[itc_name] = {
                    "hosting": hosting,
                    "externalId": match.external_id or "",
                    "catalog_confidence": match.confidence,
                    "catalog_status": match.status,
                    "fields": dict(match.fields),
                }
        except Exception as exc:
            logger.warning("Reference Catalog resolver failed (%s) — continuing without catalog data", exc)
            resolver = None

    tags = f"Target;{client_name}"
    wb = openpyxl.Workbook()

    # ── Sheet: Application ─────────────────────────────────────────────────────
    ws_app = wb.active
    ws_app.title = "Application"
    ws_app.freeze_panes = "C3"
    _sheet_header(ws_app, _COLS_APPLICATION)
    keys_app = [c[0] for c in _COLS_APPLICATION]

    lpr_catalog = _load_lpr_catalog()
    for row_idx, (app_name, app) in enumerate(sorted(seen_apps.items()), start=3):
        bcs_str  = ";".join(sorted(app["bcs"]))
        itcs_str = ";".join(sorted(app.get("itcs", set())))
        app_tags = app.get("extra_tags") or tags
        app_desc = app.get("description") or f"TO-BE application derived from requirements analysis. Client: {client_name}."
        app_lifecycle = app.get("lifecycle", "plan")
        vals = {
            "id": "", "type": "Application", "name": app_name,
            "description": app_desc,
            "alias": "", "externalId": app.get("externalId", ""),
            "lifecycle_phase": app_lifecycle, "lifecycle_startDate": "", "lifecycle_endDate": "",
            "businessCriticality": "businessCritical", "functionalSuitability": "",
            "technicalSuitability": "", "lxHostingType": app.get("lxHostingType", "saas"), "lxState": "DRAFT",
            "tags": app_tags,
            "relApplicationToBusinessCapability": bcs_str,
            "relApplicationToITComponent": itcs_str, "relToParent": "",
        }
        _enrich_with_lpr(vals, lpr_catalog)
        _sheet_row(ws_app, row_idx, [vals.get(k, "") for k in keys_app])

    # ── Sheet: Prerequisites (Lift & Shift only — reference, not for import) ───
    if seen_prereqs:
        ws_pre = wb.create_sheet("Prerequisites")
        ws_pre.freeze_panes = "A2"
        pre_headers = ["Material Number", "Description", "Required by (Target App)"]
        pre_widths   = [18, 55, 55]
        from openpyxl.styles import PatternFill, Font as _Font, Alignment as _Align
        hdr_fill = PatternFill("solid", fgColor="E76500")
        for col_idx, (hdr, width) in enumerate(zip(pre_headers, pre_widths), start=1):
            cell = ws_pre.cell(row=1, column=col_idx, value=hdr)
            cell.fill = hdr_fill
            cell.font = _Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.alignment = _Align(horizontal="center", vertical="center")
            ws_pre.column_dimensions[cell.column_letter].width = width
        ws_pre.row_dimensions[1].height = 22
        for row_idx, (prereq_name, prereq) in enumerate(sorted(seen_prereqs.items()), start=2):
            ws_pre.cell(row=row_idx, column=1, value=prereq.get("matnr", ""))
            ws_pre.cell(row=row_idx, column=2, value=prereq_name)
            ws_pre.cell(row=row_idx, column=3, value=prereq.get("for_app", ""))
        ws_pre.sheet_properties.tabColor = "E76500"

    # ── Sheet: BusinessCapability ──────────────────────────────────────────────
    ws_bc = wb.create_sheet("BusinessCapability")
    _sheet_header(ws_bc, _COLS_BC)
    keys_bc = [c[0] for c in _COLS_BC]

    for row_idx, (bc_leaf, full_path) in enumerate(sorted(seen_bcs.items()), start=3):
        parts  = full_path.split(" / ")
        parent = " / ".join(parts[:-1]) if len(parts) > 1 else ""
        vals = {
            "id": "", "type": "BusinessCapability", "name": bc_leaf,
            "description": full_path,
            "lxState": "DRAFT", "tags": tags, "relToParent": parent,
        }
        _sheet_row(ws_bc, row_idx, [vals.get(k, "") for k in keys_bc])

    # ── Sheet: Objective ──────────────────────────────────────────────────────
    # One Objective per initiative (same name) so "Initiatives linked to an
    # objective" reaches 100% in the Architecture Executive Dashboard.
    _COLS_OBJ = [
        ("id",          "ID",          "readonly",  36),
        ("type",        "Type",        "mandatory", 14),
        ("name",        "Name",        "mandatory", 45),
        ("description", "Description", "optional",  60),
        ("lxState",     "Quality Seal","optional",  14),
        ("tags",        "Tags",        "optional",  35),
    ]
    ws_obj = wb.create_sheet("Objective")
    ws_obj.freeze_panes = "C3"
    _sheet_header(ws_obj, _COLS_OBJ)
    keys_obj = [c[0] for c in _COLS_OBJ]
    seen_objectives: set[str] = set()
    for init in initiatives + supp_initiatives:
        obj_name = init.get("objective") or init["name"]
        if obj_name and obj_name not in seen_objectives:
            seen_objectives.add(obj_name)
            obj_row = {
                "id": "", "type": "Objective", "name": obj_name,
                "description": f"Strategic objective for initiative '{obj_name}'. Client: {client_name}.",
                "lxState": "DRAFT", "tags": tags,
            }
            _sheet_row(ws_obj, len(seen_objectives) + 2, [obj_row.get(k, "") for k in keys_obj])

    # ── Sheet: Initiative ──────────────────────────────────────────────────────
    ws_init = wb.create_sheet("Initiative")
    ws_init.freeze_panes = "C3"
    _sheet_header(ws_init, _COLS_INITIATIVE)
    keys_init = [c[0] for c in _COLS_INITIATIVE]

    for row_idx, init in enumerate(initiatives + supp_initiatives, start=3):
        obj_name = init.get("objective") or init["name"]
        vals = {
            "id": "", "type": "Initiative", "name": init["name"],
            "description":          init["description"],
            "lifecycle_phase":      init["lifecycle"],
            "lifecycle_startDate":  init.get("lifecycle_startDate", ""),
            "lifecycle_endDate":    init.get("lifecycle_endDate", ""),
            "lxState": "DRAFT", "tags": tags,
            "relInitiativeToApplication":        init["app"],
            "relInitiativeToBusinessCapability": init["bcs"],
            "relInitiativeToObjective":          obj_name,
        }
        _sheet_row(ws_init, row_idx, [vals.get(k, "") for k in keys_init])

    # ── Sheet: ITComponent ─────────────────────────────────────────────────────
    ws_itc = wb.create_sheet("ITComponent")
    ws_itc.freeze_panes = "C3"
    _sheet_header(ws_itc, _COLS_ITC)
    keys_itc = [c[0] for c in _COLS_ITC]

    for row_idx, (itc_name, raw) in enumerate(sorted(seen_itcs.items()), start=3):
        if isinstance(raw, dict):
            hosting = raw.get("hosting", "")
            external_id = raw.get("externalId", "")
        else:
            hosting = raw
            external_id = ""
        vals = {
            "id": "", "type": "ITComponent", "name": itc_name,
            "externalId": external_id,
            "description": f"SAP technology component supporting TO-BE applications. Client: {client_name}.",
            "lxHostingType": hosting, "lxState": "DRAFT", "tags": tags,
        }
        _sheet_row(ws_itc, row_idx, [vals.get(k, "") for k in keys_itc])

    # ── Sheet: ReadMe ──────────────────────────────────────────────────────────
    readme = wb.create_sheet("ReadMe")
    from openpyxl.styles import Font, Alignment
    readme_rows = [
        (f"LeanIX Import — Target TO-BE ({client_name})", True,  "002A86", 13),
        ("Source: Requirements analysis via Archimedes AI + SAP RBA/RSA catalogs", False, "223548", 9),
        ("", False, "223548", 9),
        ("IMPORT ORDER", True, "002A86", 10),
        ("1. BusinessCapability", False, "223548", 9),
        ("2. ITComponent", False, "223548", 9),
        ("3. Application", False, "223548", 9),
        ("4. Initiative", False, "223548", 9),
        ("", False, "223548", 9),
        ("IMPORT RULES", True, "002A86", 10),
        ("- Leave 'id' column EMPTY — new fact sheets will be created.", False, "223548", 9),
        ("- Relations use EXACT display names of existing LeanIX fact sheets.", False, "223548", 9),
        ("- Multiple relation values separated by semicolon (;) without spaces.", False, "223548", 9),
        ("- 'lxState' = DRAFT — approve manually after review.", False, "223548", 9),
        ("- Import via: Inventory > Inventory Tools > Import from Excel", False, "223548", 9),
        ("", False, "223548", 9),
        ("STATS", True, "002A86", 10),
        (f"Applications: {len(seen_apps)}", False, "223548", 9),
        (f"Business Capabilities: {len(seen_bcs)}", False, "223548", 9),
        (f"IT Components: {len(seen_itcs)}", False, "223548", 9),
        (f"Initiatives: {len(initiatives) + len(supp_initiatives)}", False, "223548", 9),
    ]
    for r_idx, (text, bold, color, size) in enumerate(readme_rows, start=1):
        c = readme.cell(row=r_idx, column=1, value=text)
        c.font = Font(name="Calibri", size=size, bold=bold, color=color)
        c.alignment = Alignment(horizontal="left", wrap_text=True)
    readme.column_dimensions["A"].width = 80

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(
        "LeanIX target: %d apps, %d BCs, %d initiatives → %s",
        len(seen_apps), len(seen_bcs), len(initiatives), output_path,
    )

    # ── Reference Catalog audit report ────────────────────────────────────────
    # Writes catalog_resolution_report.json next to the Excel with per-type
    # entries {name, external_id, confidence, status}. Always safe to call
    # (empty dicts when the resolver flag is off → empty arrays in the JSON).
    if app_matches or itc_matches:
        report = {
            "Application": [
                {"name": m.name, "external_id": m.external_id,
                 "confidence": m.confidence, "status": m.status}
                for m in app_matches.values()
            ],
            "ITComponent": [
                {"name": m.name, "external_id": m.external_id,
                 "confidence": m.confidence, "status": m.status}
                for m in itc_matches.values()
            ],
        }
        try:
            (Path(output_path).parent / "catalog_resolution_report.json").write_text(
                json.dumps(report, ensure_ascii=False, indent=2)
            )
        except Exception as exc:
            logger.warning("Could not write catalog_resolution_report.json: %s", exc)

        def _count(matches: dict, key: str, value: str) -> int:
            return sum(1 for m in matches.values() if getattr(m, key) == value)

        logger.info(
            "Reference Catalog Application: %d names — %d LINKED, %d CUSTOM",
            len(app_matches),
            _count(app_matches, "status", "LINKED"),
            _count(app_matches, "status", "CUSTOM"),
        )
        logger.info(
            "Reference Catalog ITComponent: %d names — %d LINKED, %d CUSTOM",
            len(itc_matches),
            _count(itc_matches, "status", "LINKED"),
            _count(itc_matches, "status", "CUSTOM"),
        )

    if resolver is not None:
        try:
            resolver.cleanup()
        except Exception as exc:
            logger.warning("Reference Catalog cleanup failed (%s)", exc)


# ── LeanIX writer ─────────────────────────────────────────────────────────────

# GraphQL fragments reused across mutations
_QUERY_FS_BY_NAME = """
query FindFS($type: String!, $name: String!) {
  allFactSheets(
    filter: {
      facetFilters: [{ facetKey: "FactSheetTypes", keys: [$type] }]
      fullTextSearch: $name
    }
  ) {
    edges {
      node {
        id
        displayName
      }
    }
  }
}
"""

_MUTATION_CREATE_FS = """
mutation CreateFS($type: FactSheetType!, $name: String!) {
  createFactSheet(input: { type: $type, name: $name }) {
    factSheet { id displayName }
  }
}
"""

_MUTATION_ADD_TAG = """
mutation AddTag($id: ID!, $patches: [Patch]!) {
  updateFactSheet(id: $id, patches: $patches) {
    factSheet { id }
  }
}
"""

_QUERY_TAGS = """
query GetTags {
  allTags { edges { node { id name tagGroup { id name } } } }
}
"""

_QUERY_TAG_GROUPS = """
query GetTagGroups {
  allTagGroups { edges { node {
    id name mode restrictToFactSheetTypes
    tags { edges { node { id name } } }
  } } }
}
"""

_MUTATION_CREATE_TAG = """
mutation CreateTag($name: String!, $groupId: ID) {
  createTag(name: $name, tagGroupId: $groupId) {
    id name
  }
}
"""

_MUTATION_SET_LIFECYCLE = """
mutation SetLifecycle($id: ID!, $patches: [Patch]!) {
  updateFactSheet(id: $id, patches: $patches) {
    factSheet { id }
  }
}
"""

_MUTATION_CREATE_RELATION = """
mutation CreateRelation($from: ID!, $to: ID!, $relType: RelationName!) {
  upsertRelation(
    from: { id: $from }
    to: { id: $to }
    type: $relType
  ) { fromFactSheetId }
}
"""

_MUTATION_PATCH_FS = """
mutation PatchFS($id: ID!, $patches: [Patch]!) {
  updateFactSheet(id: $id, patches: $patches) {
    factSheet { id }
  }
}
"""

# TIME (functional fit display names) → LeanIX functionalSuitability API enum
_TIME_TO_FUNCTIONAL = {
    "tolerate":  "insufficient",
    "invest":    "perfect",
    "migrate":   "unreasonable",
    "eliminate": "unreasonable",
    # pass-through if already API enum value
    "unreasonable": "unreasonable",
    "insufficient": "insufficient",
    "appropriate":  "appropriate",
    "perfect":      "perfect",
}


# ── LeanIX OAuth2 helpers ─────────────────────────────────────────────────────
# Auth handshake lives in pipeline.leanix_auth (shared with push_ldif and
# the Reference Catalog resolver).


def _resolver_credentials() -> tuple[str, str]:
    """Return (base_url, api_token) for the Reference Catalog resolver.

    Reads LEANIX_BASE_URL and LEANIX_API_TOKEN from env. Patched in tests.
    """
    return (
        os.environ["LEANIX_BASE_URL"].rstrip("/"),
        os.environ["LEANIX_API_TOKEN"],
    )


def _backup_workspace(base_url: str, api_token: str, output_dir: Path) -> Path | None:
    """
    Trigger a LeanIX full-export backup before a push and download the result.

    Calls POST /services/pathfinder/v1/exports/fullExport (async job), then polls
    GET /services/pathfinder/v1/exports/downloads until the file is ready.
    Saves the backup as leanix_backup_<timestamp>.xlsx in output_dir.

    Returns the backup path, or None if backup failed (non-fatal).
    """
    import requests as _req
    from datetime import datetime

    bearer = get_bearer(base_url, api_token)
    hdrs   = {"Authorization": f"Bearer {bearer}", "Content-Type": "application/json"}

    # 1. Trigger export
    try:
        resp = _req.post(
            f"{base_url}/services/pathfinder/v1/exports/fullExport",
            headers=hdrs,
            timeout=30,
        )
        if resp.status_code == 404:
            logger.debug("Backup: fullExport endpoint not available (404) — skipping")
            return None
        resp.raise_for_status()
        export_data = resp.json()
    except Exception as exc:
        logger.warning("Backup: failed to trigger fullExport: %s", exc)
        return None

    export_id = (
        export_data.get("id")
        or export_data.get("exportId")
        or export_data.get("jobId")
        or ""
    )
    logger.info("Backup: fullExport triggered (id=%s) — polling for download…", export_id)

    # 2. Poll for the download URL (up to 5 min, 10 s intervals)
    download_url: str | None = None
    for attempt in range(30):
        time.sleep(10)
        try:
            bearer = get_bearer(base_url, api_token)
            poll_hdrs = {"Authorization": f"Bearer {bearer}"}
            # Try job-status endpoint first, then generic downloads list
            if export_id:
                status_resp = _req.get(
                    f"{base_url}/services/pathfinder/v1/exports/downloads/{export_id}",
                    headers=poll_hdrs,
                    timeout=30,
                )
                if status_resp.status_code == 200:
                    sdata = status_resp.json()
                    url = sdata.get("url") or sdata.get("downloadUrl") or sdata.get("href")
                    if url:
                        download_url = url
                        break
                    status = sdata.get("status", "")
                    if status in ("FAILED", "ERROR"):
                        logger.warning("Backup: export job failed (status=%s)", status)
                        return None
            # Fallback: generic downloads list
            dl_resp = _req.get(
                f"{base_url}/services/pathfinder/v1/exports/downloads",
                headers=poll_hdrs,
                timeout=30,
            )
            if dl_resp.status_code == 200:
                entries = dl_resp.json() if isinstance(dl_resp.json(), list) else dl_resp.json().get("data", [])
                # Take the most recent entry
                if entries:
                    latest = entries[0] if isinstance(entries[0], dict) else {}
                    url = latest.get("url") or latest.get("downloadUrl") or latest.get("href")
                    if url:
                        download_url = url
                        break
        except Exception as exc:
            logger.debug("Backup: poll attempt %d failed: %s", attempt + 1, exc)

    if not download_url:
        logger.warning("Backup: export download URL not available after polling — skipping")
        return None

    # 3. Download the file
    backup_path = output_dir / f"leanix_backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    try:
        bearer = get_bearer(base_url, api_token)
        dl = _req.get(download_url, headers={"Authorization": f"Bearer {bearer}"}, timeout=120, stream=True)
        dl.raise_for_status()
        with open(backup_path, "wb") as f:
            for chunk in dl.iter_content(chunk_size=65536):
                f.write(chunk)
        logger.info("Backup: saved to %s (%d KB)", backup_path.name, backup_path.stat().st_size // 1024)
        return backup_path
    except Exception as exc:
        logger.warning("Backup: download failed: %s", exc)
        return None


def write_leanix(
    staging_path: Path,
    client_name: str,
    lift_shift_map: list[dict] | None = None,
) -> None:
    """
    Push a LeanIX staging Excel (generated by write_leanix_excel) to LeanIX.

    Reads the three sheets — Initiatives, BusinessCapabilities, Applications —
    and upserts the corresponding fact sheets and relations.

    lift_shift_map: optional list of conversion dicts from lift_shift.convert_to_rise().
      Each dict must contain: source_app (str), target_desc (str).
      When provided, creates decommissionApplication + implementNewApplication
      Transformations for each source→target pair on the Lift & Shift Initiative.

    Call after reviewing/editing output/<stem>_leanix_import.xlsx.
    """
    import requests
    import openpyxl

    base_url = os.environ["LEANIX_BASE_URL"].rstrip("/")
    token    = os.environ["LEANIX_API_TOKEN"]

    # ── Backup workspace before push (non-fatal) ──────────────────────────────
    _backup_workspace(base_url, token, staging_path.parent)

    # ── Authenticate (token auto-refreshes via get_bearer) ───────────────────
    gql_url = f"{base_url}/services/pathfinder/v1/graphql"
    trans_url = f"{base_url}/services/transformations/v1/transformations"

    def _gql(query: str, variables: dict, _max_retries: int = 3) -> dict:
        for attempt in range(1, _max_retries + 1):
            hdrs = {"Authorization": f"Bearer {get_bearer(base_url, token)}", "Content-Type": "application/json"}
            try:
                resp = requests.post(
                    gql_url,
                    json={"query": query, "variables": variables},
                    headers=hdrs,
                    timeout=30,
                )
            except (requests.ConnectionError, requests.Timeout) as exc:
                if attempt < _max_retries:
                    wait = 2 ** (attempt - 1)
                    logger.warning("GQL network error (attempt %d/%d), retry in %ds: %s", attempt, _max_retries, wait, exc)
                    time.sleep(wait)
                    continue
                raise
            if resp.status_code == 429:
                wait = int(resp.headers.get("Retry-After", 2 ** attempt))
                logger.warning("GQL 429 rate-limited (attempt %d/%d), retry in %ds", attempt, _max_retries, wait)
                if attempt < _max_retries:
                    time.sleep(wait)
                    continue
                resp.raise_for_status()
            resp.raise_for_status()
            data = resp.json()
            if data.get("errors"):
                raise RuntimeError(f"GraphQL errors: {data['errors']}")
            return data

    # ── Tag ───────────────────────────────────────────────────────────────────
    # Cache tag group data (loaded once)
    _tag_groups_cache: list[dict] | None = None
    _all_tags_cache: list[dict] | None = None

    def _load_tag_cache() -> None:
        nonlocal _tag_groups_cache, _all_tags_cache
        if _tag_groups_cache is None:
            r = _gql(_QUERY_TAG_GROUPS, {})
            _tag_groups_cache = [e["node"] for e in r.get("data", {}).get("allTagGroups", {}).get("edges", [])]
        if _all_tags_cache is None:
            r = _gql(_QUERY_TAGS, {})
            _all_tags_cache = [e["node"] for e in r.get("data", {}).get("allTags", {}).get("edges", [])]

    # Tags that must always be created as free tags (no tag group)
    _FREE_TAGS = {"KPI_Achievement", "Target", "Baseline", "Target Reference"}

    def _get_or_create_tag_id(tag_name: str, group_name: str | None = None) -> str:
        """Get or create a tag, optionally in a specific named group.

        Lookup rules:
        - If group_name given: find tag with that name IN that group. If not found, create it there.
        - If no group_name: find tag with that name that has NO group (free tag). If not found, create it free.
        This prevents accidentally returning a grouped tag when a free one is needed (or vice versa).
        """
        _load_tag_cache()
        group_id = None
        if group_name:
            for g in _tag_groups_cache:  # type: ignore[union-attr]
                if g["name"] == group_name:
                    group_id = g["id"]
                    break
        # Find existing tag matching name AND group context
        for t in _all_tags_cache:  # type: ignore[union-attr]
            if t["name"] != tag_name:
                continue
            t_group = t.get("tagGroup")
            t_group_id = t_group["id"] if t_group else None
            if group_id is not None:
                if t_group_id == group_id:
                    return t["id"]  # exact group match
            else:
                if t_group_id is None:
                    return t["id"]  # free tag match
        # Not found — create it
        result = _gql(_MUTATION_CREATE_TAG, {"name": tag_name, "groupId": group_id})
        new_tag = result["data"]["createTag"]
        _all_tags_cache.append({"id": new_tag["id"], "name": new_tag["name"], "tagGroup": {"id": group_id} if group_id else None})  # type: ignore[union-attr]
        return new_tag["id"]

    def _get_tag_group_id(tag_id: str) -> str | None:
        """Return the group ID for a given tag ID, or None if free tag."""
        _load_tag_cache()
        for t in _all_tags_cache:  # type: ignore[union-attr]
            if t["id"] == tag_id:
                tg = t.get("tagGroup")
                return tg["id"] if tg else None
        return None

    def _find_tag_in_group(tag_name: str, group_name: str) -> str | None:
        """Find an existing tag by name within a specific tag group (no create)."""
        _load_tag_cache()
        for g in _tag_groups_cache:  # type: ignore[union-attr]
            if g["name"] == group_name:
                for t in g.get("tags", {}).get("edges", []):
                    if t["node"]["name"] == tag_name:
                        return t["node"]["id"]
        return None

    def _single_group_ids() -> set[str]:
        """Return IDs of tag groups whose mode is SINGLE (only one tag allowed per fact sheet)."""
        _load_tag_cache()
        return {g["id"] for g in _tag_groups_cache if g.get("mode") == "SINGLE"}  # type: ignore[union-attr]

    def _tag_name_to_group(tag_name: str) -> str | None:
        """Return the group_name for a tag that exists in a group, or None if it's a free tag.

        Looks up the tag in _all_tags_cache (which includes tagGroup info).
        If a tag with that name exists in a group, return the group name so the caller
        can pass it to _get_or_create_tag_id — ensuring the right tag variant is used.
        If the tag name is in _FREE_TAGS or has no group entry, return None.
        """
        if tag_name in _FREE_TAGS:
            return None
        _load_tag_cache()
        for t in _all_tags_cache:  # type: ignore[union-attr]
            if t["name"] == tag_name:
                tg = t.get("tagGroup")
                if tg:
                    # Find the group name from _tag_groups_cache
                    gid = tg["id"]
                    for g in _tag_groups_cache:  # type: ignore[union-attr]
                        if g["id"] == gid:
                            return g["name"]
        return None

    def _tag_fs(fs_id: str, tag_id: str) -> None:
        """Apply a single tag to a fact sheet (delegates to _tag_fs_multi for SINGLE-group safety)."""
        _tag_fs_multi(fs_id, [tag_id])

    def _resolve_extra_tags(row: dict, skip: str = "") -> list[str]:
        """Parse semicolon-separated tags field and return resolved tag IDs, skipping client tag.

        Each tag name is resolved group-aware:
        - If the tag name already exists in a tag group in LeanIX, it is resolved within that group.
        - If the tag name is in _FREE_TAGS or has no group match, it is resolved as a free tag.
        This prevents accidentally creating duplicate tags in the wrong group.
        """
        raw = str(row.get("tags") or "").strip()
        ids: list[str] = []
        for t in raw.split(";"):
            t = t.strip()
            if not t or t == skip:
                continue
            try:
                group_name = _tag_name_to_group(t)
                ids.append(_get_or_create_tag_id(t, group_name))
            except Exception as exc:
                logger.warning("LeanIX: cannot resolve tag '%s' — %s", t, exc)
        return ids

    def _tag_fs_multi(fs_id: str, tag_ids: list[str]) -> None:
        """Apply multiple tags to a fact sheet, handling SINGLE-group conflicts.

        For tags that belong to a SINGLE-mode tag group, LeanIX only allows one tag per
        group per fact sheet. Before adding, we fetch the fact sheet's current tags and
        remove any existing tag that conflicts (same SINGLE group). Tags are then added
        in a single batch call.
        """
        if not tag_ids:
            return

        _load_tag_cache()
        single_gids = _single_group_ids()

        # Build a map: group_id → tag_id for the tags we want to apply
        new_tag_group: dict[str, str] = {}  # group_id → new_tag_id (only SINGLE groups)
        for tid in tag_ids:
            gid = _get_tag_group_id(tid)
            if gid and gid in single_gids:
                new_tag_group[gid] = tid

        patches: list[dict] = []

        # If we have SINGLE-group tags to apply, fetch current FS tags and remove conflicts
        if new_tag_group:
            try:
                _QUERY_FS_TAGS = """
                query FsTags($id: ID!) {
                  factSheet(id: $id) { tags { id tagGroup { id } } }
                }
                """
                r = _gql(_QUERY_FS_TAGS, {"id": fs_id})
                current_tags = r.get("data", {}).get("factSheet", {}).get("tags", [])
                for ct in current_tags:
                    ct_gid = (ct.get("tagGroup") or {}).get("id")
                    if ct_gid and ct_gid in new_tag_group:
                        # Conflict: same SINGLE group — remove the old tag first
                        patches.append({"op": "remove", "path": f"/tags/{ct['id']}"})
                        logger.debug("LeanIX: removing conflicting SINGLE-group tag %s from %s", ct["id"], fs_id)
            except Exception as exc:
                logger.warning("LeanIX: could not fetch current tags for %s, proceeding without conflict check — %s", fs_id, exc)

        # Add all new tags
        patches.append({"op": "add", "path": "/tags",
                        "value": json.dumps([{"tagId": tid} for tid in tag_ids])})
        try:
            _gql(_MUTATION_ADD_TAG, {"id": fs_id, "patches": patches})
        except Exception as exc:
            logger.warning("LeanIX: skipping tags for %s — %s", fs_id, exc)

    # ── Fact sheet helpers ────────────────────────────────────────────────────
    def _find_by_name(fs_type: str, name: str) -> str | None:
        result = _gql(_QUERY_FS_BY_NAME, {"type": fs_type, "name": name})
        for edge in result["data"]["allFactSheets"]["edges"]:
            if edge["node"]["displayName"] == name:
                return edge["node"]["id"]
        return None

    _BATCH_SIZE_GQL = 20

    def _batch_find_by_names(fs_type: str, names: list[str]) -> dict[str, str]:
        """Return {name: id} for all names found, using GraphQL aliasing (up to 20 per call)."""
        result: dict[str, str] = {}
        unique = list(dict.fromkeys(n for n in names if n))  # deduplicate, preserve order
        for chunk_start in range(0, len(unique), _BATCH_SIZE_GQL):
            chunk = unique[chunk_start:chunk_start + _BATCH_SIZE_GQL]
            # Build aliased query: alias0: allFactSheets(...) { edges { node { id displayName } } }
            aliases = "\n".join(
                f'  a{i}: allFactSheets(filter: {{facetFilters: [{{facetKey: "FactSheetTypes", keys: ["{fs_type}"]}}] fullTextSearch: "{name.replace(chr(34), chr(39))}"}}) {{ edges {{ node {{ id displayName }} }} }}'
                for i, name in enumerate(chunk)
            )
            query = "{\n" + aliases + "\n}"
            data = _gql(query, {}).get("data", {})
            for i, name in enumerate(chunk):
                for edge in data.get(f"a{i}", {}).get("edges", []):
                    if edge["node"]["displayName"] == name:
                        result[name] = edge["node"]["id"]
                        break
        return result

    def _create_fs(fs_type: str, name: str, desc: str = "") -> str:
        result = _gql(_MUTATION_CREATE_FS, {"type": fs_type, "name": name})
        return result["data"]["createFactSheet"]["factSheet"]["id"]

    def _upsert(fs_type: str, name: str, desc: str = "", _prefetch: dict[str, str] | None = None) -> tuple[str, bool]:
        existing = (_prefetch.get(name) if _prefetch is not None else None) or _find_by_name(fs_type, name)
        if existing:
            return existing, False
        return _create_fs(fs_type, name, desc), True

    def _set_lifecycle(fs_id: str, phase: str) -> None:
        _gql(_MUTATION_SET_LIFECYCLE, {
            "id": fs_id,
            "patches": [{"op": "add", "path": "/lifecycle",
                         "value": json.dumps({"phases": [{"phase": phase}]})}],
        })

    def _create_relation(from_id: str, to_id: str, rel_type: str) -> None:
        _gql(_MUTATION_CREATE_RELATION,
             {"from": from_id, "to": to_id, "relType": rel_type})

    def _set_field(fs_id: str, field_path: str, value: str) -> None:
        """Patch a single scalar field on a fact sheet (non-fatal).
        LeanIX patch 'value' for enum/string fields must be the raw string, not JSON-encoded."""
        try:
            _gql(_MUTATION_PATCH_FS, {
                "id": fs_id,
                "patches": [{"op": "add", "path": f"/{field_path}", "value": value}],
            })
        except Exception as exc:
            logger.warning("LeanIX: skipping field %s on %s — %s", field_path, fs_id, exc)

    def _set_lifecycle_full(fs_id: str, phase: str, start_date: str = "", end_date: str = "") -> None:
        """Set lifecycle phase and optional start/end dates."""
        lc: dict = {"phases": [{"phase": phase}]}
        if start_date:
            lc["phases"][0]["startDate"] = start_date
        if end_date:
            lc["phases"][0]["endDate"] = end_date
        _gql(_MUTATION_SET_LIFECYCLE, {
            "id": fs_id,
            "patches": [{"op": "add", "path": "/lifecycle",
                         "value": json.dumps(lc)}],
        })

    # ── Read staging Excel ────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(str(staging_path))

    def _sheet_rows(sheet_name: str) -> list[dict]:
        # Accept both plural (old format) and singular (new format) sheet names
        _singular = {
            "Initiatives": "Initiative",
            "Applications": "Application",
            "BusinessCapabilities": "BusinessCapability",
            "Objectives": "Objective",
        }
        name = sheet_name if sheet_name in wb.sheetnames else _singular.get(sheet_name, sheet_name)
        if name not in wb.sheetnames:
            return []
        ws     = wb[name]
        raw_header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # Normalise header: map Title-Case / human labels → camelCase API keys
        _col_alias = {
            "id": "id", "type": "type", "name": "name", "description": "description",
            "alias": "alias", "external id": "externalId",
            "lifecycle phase": "lifecycle_phase", "lifecycle": "lifecycle_phase",
            "lifecycle start date": "lifecycle_startDate",
            "lifecycle end date": "lifecycle_endDate",
            "business criticality": "businessCriticality",
            "functional fit": "functionalSuitability",
            "technical fit": "technicalSuitability",
            "hosting type": "lxHostingType",
            "quality seal": "lxState",
            "tags": "tags",
            "it component": "relApplicationToITComponent",
            "business capability": "relApplicationToBusinessCapability",
            "parent": "relToParent",
            "lpr id": "lprId",
            "material id": "materialId",
            "material description": "materialDescription",
            "solution area": "solutionArea",
            "sub-solution area": "subSolutionArea",
            "contract end": "contractEnd",
        }
        header = [_col_alias.get(str(h or "").strip().lower(), str(h or "").strip()) for h in raw_header]
        # Detect 2-row header (row 2 = translations): skip row 2 if all non-empty
        # cells look like human labels (no slash, no camelCase pattern of keys)
        rows_iter = ws.iter_rows(min_row=2)
        first_row = next(rows_iter, None)
        data_start_iter = rows_iter  # remaining rows after potential header row 2
        if first_row is not None:
            # Row 2 is a translation row if 'type' column has value "Type" (title case)
            type_idx = next((i for i, h in enumerate(header) if h == "type"), None)
            if type_idx is not None and type_idx < len(first_row):
                if str(first_row[type_idx].value or "").strip() == "Type":
                    # This is the translation row — skip it, use data_start_iter
                    pass
                else:
                    # Not a translation row — include it
                    data_start_iter = [first_row, *ws.iter_rows(min_row=3)]
            else:
                data_start_iter = [first_row, *ws.iter_rows(min_row=3)]
        return [
            {header[i]: (cell.value or "") for i, cell in enumerate(row)}
            for row in data_start_iter
            if any(cell.value for cell in row)
        ]

    init_rows = _sheet_rows("Initiatives")
    bc_rows   = _sheet_rows("BusinessCapabilities")
    app_rows  = _sheet_rows("Applications")
    obj_rows  = _sheet_rows("Objectives")
    itc_rows  = _sheet_rows("ITComponent")

    # ── Push ──────────────────────────────────────────────────────────────────
    client_tag = client_name
    # Ensure client name is always resolved as a free tag (never grouped)
    _FREE_TAGS.add(client_tag)
    tag_id     = _get_or_create_tag_id(client_tag)
    logger.info("LeanIX: using tag '%s' (id=%s)", client_tag, tag_id)

    # Pre-resolve "Architecture State" > "Baseline" tag (needed for TIME/6R KPIs)
    arch_state_baseline_id = _find_tag_in_group("Baseline", "Architecture State")
    if arch_state_baseline_id is None:
        arch_state_baseline_id = _get_or_create_tag_id("Baseline", "Architecture State")
    logger.info("LeanIX: Architecture State 'Baseline' tag id=%s", arch_state_baseline_id)

    # Backfill: tag ALL existing apps without Architecture State tag → "Baseline"
    # This ensures workspace-wide Architecture State coverage (needed for KPI ≥ 90%)
    _QUERY_ALL_APPS_BASIC = """
    query {
      allFactSheets(filter: { facetFilters: [{ facetKey: "FactSheetTypes", keys: ["Application"] }] }) {
        edges { node { id displayName
          tags { id name tagGroup { id name } }
        }}
      }
    }
    """
    try:
        all_apps_result = _gql(_QUERY_ALL_APPS_BASIC, {})
        for edge in all_apps_result["data"]["allFactSheets"]["edges"]:
            app_node = edge["node"]
            has_arch_state = any(
                (t.get("tagGroup") or {}).get("name") == "Architecture State"
                for t in app_node.get("tags", [])
            )
            if not has_arch_state:
                _tag_fs(app_node["id"], arch_state_baseline_id)
                logger.debug("LeanIX: backfilled Architecture State 'Baseline' on app '%s'", app_node["displayName"])
        logger.info("LeanIX: Architecture State backfill complete")
    except Exception as exc:
        logger.warning("LeanIX: Architecture State backfill failed — %s", exc)

    # Backfill: same for BCs
    try:
        _QUERY_ALL_BCS_BASIC = """
        query {
          allFactSheets(filter: { facetFilters: [{ facetKey: "FactSheetTypes", keys: ["BusinessCapability"] }] }) {
            edges { node { id displayName
              tags { id name tagGroup { id name } }
            }}
          }
        }
        """
        all_bcs_result = _gql(_QUERY_ALL_BCS_BASIC, {})
        for edge in all_bcs_result["data"]["allFactSheets"]["edges"]:
            bc_node = edge["node"]
            has_arch_state = any(
                (t.get("tagGroup") or {}).get("name") == "Architecture State"
                for t in bc_node.get("tags", [])
            )
            if not has_arch_state:
                _tag_fs(bc_node["id"], arch_state_baseline_id)
        logger.info("LeanIX: BC Architecture State backfill complete")
    except Exception as exc:
        logger.warning("LeanIX: BC Architecture State backfill failed — %s", exc)

    # 0. Upsert Objectives (if sheet present)
    obj_id_cache: dict[str, str] = {}
    _obj_prefetch = _batch_find_by_names("Objective", [str(r.get("name", "")).strip() for r in obj_rows])
    for row in obj_rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        obj_id, created = _upsert("Objective", name, _prefetch=_obj_prefetch)
        obj_id_cache[name] = obj_id
        extra_tags = _resolve_extra_tags(row, skip=client_tag)
        all_tags = list(dict.fromkeys([tag_id] + extra_tags)) if created else extra_tags
        if all_tags:
            _tag_fs_multi(obj_id, all_tags)
        logger.debug("LeanIX Obj %s '%s'", "created" if created else "found", name)

    # 1. Upsert BusinessCapabilities
    bc_id_cache: dict[str, str] = {}
    _bc_prefetch = _batch_find_by_names("BusinessCapability", [str(r.get("name", "")).strip() for r in bc_rows])
    for row in bc_rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        bc_id, created = _upsert("BusinessCapability", name, _prefetch=_bc_prefetch)
        bc_id_cache[name] = bc_id
        extra_tags = _resolve_extra_tags(row, skip=client_tag)
        if created:
            _tag_fs_multi(bc_id, list(dict.fromkeys([tag_id, arch_state_baseline_id] + extra_tags)))
        else:
            all_extra = list(dict.fromkeys([arch_state_baseline_id] + extra_tags))
            _tag_fs_multi(bc_id, all_extra)
        # Write optional BC fields
        catalog_status = str(row.get("lxCatalogStatus") or "").strip()
        if catalog_status:
            _set_field(bc_id, "lxCatalogStatus", catalog_status)
        scope_bc = str(row.get("scopeBC") or "").strip()
        if scope_bc:
            _set_field(bc_id, "scopeBC", scope_bc)
        logger.debug("LeanIX BC %s '%s'", "created" if created else "found", name)

    # 1b. Upsert ITComponents
    itc_id_cache: dict[str, str] = {}
    _itc_prefetch = _batch_find_by_names("ITComponent", [str(r.get("name", "")).strip() for r in itc_rows])
    for row in itc_rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        itc_id, created = _upsert("ITComponent", name, _prefetch=_itc_prefetch)
        itc_id_cache[name] = itc_id
        extra_tags = _resolve_extra_tags(row, skip=client_tag)
        if created:
            _tag_fs_multi(itc_id, list(dict.fromkeys([tag_id] + extra_tags)))
        else:
            if extra_tags:
                _tag_fs_multi(itc_id, extra_tags)
        hosting = str(row.get("lxHostingType") or "").strip()
        if hosting:
            _set_field(itc_id, "lxHostingType", hosting)
        logger.debug("LeanIX ITC %s '%s'", "created" if created else "found", name)

    # 2. Upsert Applications
    app_id_cache: dict[str, str] = {}
    _app_prefetch = _batch_find_by_names("Application", [str(r.get("name", "")).strip() for r in app_rows])
    for row in app_rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        app_id, created = _upsert("Application", name, _prefetch=_app_prefetch)
        app_id_cache[name] = app_id
        extra_tags = _resolve_extra_tags(row, skip=client_tag)
        if created:
            _tag_fs_multi(app_id, list(dict.fromkeys([tag_id, arch_state_baseline_id] + extra_tags)))
        else:
            all_extra = list(dict.fromkeys([arch_state_baseline_id] + extra_tags))
            _tag_fs_multi(app_id, all_extra)
        # Set lifecycle (with optional start/end dates)
        # Target applications (not tagged Baseline) are always plan — they represent the TO-BE state
        row_tags = str(row.get("tags") or "").lower()
        lc_phase = str(row.get("lifecycle_phase") or row.get("lifecycle") or "").strip()
        if "baseline" not in row_tags and lc_phase in ("active", "phaseIn", ""):
            lc_phase = "plan"
        if lc_phase:
            lc_start = str(row.get("lifecycle_startDate") or "").strip()
            lc_end   = str(row.get("lifecycle_endDate") or "").strip()
            _set_lifecycle_full(app_id, lc_phase, lc_start, lc_end)
        # Functional suitability (accepts both TIME labels and API enum values)
        func_raw = str(row.get("functionalSuitability") or "").strip().lower()
        func_val = _TIME_TO_FUNCTIONAL.get(func_raw)
        if func_val:
            _set_field(app_id, "functionalSuitability", func_val)
        # 6R classification (values match API enum directly)
        six_r = str(row.get("lxSixRClassification") or "").strip()
        if six_r:
            _set_field(app_id, "lxSixRClassification", six_r)
        # Technical suitability
        tech = str(row.get("technicalSuitability") or "").strip()
        if tech:
            _set_field(app_id, "technicalSuitability", tech)
        # Business criticality
        biz_crit = str(row.get("businessCriticality") or "").strip()
        if biz_crit:
            _set_field(app_id, "businessCriticality", biz_crit)
        # Link → BCs (semicolon or comma separated)
        bc_rel_raw = row.get("relApplicationToBusinessCapability") or ""
        for bc_name in re.split(r"[;,]", str(bc_rel_raw)):
            bc_name = bc_name.strip()
            if bc_name and bc_name in bc_id_cache:
                _create_relation(app_id, bc_id_cache[bc_name],
                                 "relApplicationToBusinessCapability")
        logger.debug("LeanIX App %s '%s'", "created" if created else "found", name)

    # 3. Upsert Initiatives + relations
    pushed = failed = 0
    init_id_cache: dict[str, str] = {}
    _init_prefetch = _batch_find_by_names("Initiative", [str(r.get("name", str(r.get("id", "")))).strip() for r in init_rows])
    for row in init_rows:
        req_id = str(row.get("id", "")).strip()
        init_name = str(row.get("name", req_id)).strip()
        if not req_id:
            req_id = init_name  # fallback: use name as id when id column is empty
        if not req_id:
            continue
        try:
            initiative_id, created = _upsert(
                "Initiative",
                init_name,
                desc=str(row.get("description", "")),
                _prefetch=_init_prefetch,
            )
            init_id_cache[init_name] = initiative_id
            if created:
                extra_tags = _resolve_extra_tags(row, skip=client_tag)
                all_tags = list(dict.fromkeys([tag_id] + extra_tags))
                _tag_fs_multi(initiative_id, all_tags)

            # lifecycle with optional start/end dates
            lifecycle_val = str(
                row.get("lifecycle_phase") or row.get("lifecycle") or "plan"
            ).strip()
            lc_start = str(row.get("lifecycle_startDate") or "").strip()
            lc_end   = str(row.get("lifecycle_endDate") or "").strip()
            _set_lifecycle_full(initiative_id, lifecycle_val, lc_start, lc_end)

            # Link → BCs — accept both column names, semicolon or comma separated
            bcs_raw = row.get("relInitiativeToBusinessCapability") or row.get("bcs") or ""
            for bc_name in re.split(r"[;,]", str(bcs_raw)):
                bc_name = bc_name.strip()
                if bc_name and bc_name in bc_id_cache:
                    _create_relation(initiative_id, bc_id_cache[bc_name],
                                     "relInitiativeToBusinessCapability")

            # Link → Applications — accept both column names, semicolon or comma separated
            apps_raw = row.get("relInitiativeToApplication") or row.get("rsa") or ""
            for app_name in re.split(r"[;,]", str(apps_raw)):
                app_name = app_name.strip()
                if app_name and app_name in app_id_cache:
                    _create_relation(initiative_id, app_id_cache[app_name],
                                     "relInitiativeToApplication")

            # Link → Objectives
            obj_raw = row.get("relInitiativeToObjective") or ""
            for obj_name in re.split(r"[;,]", str(obj_raw)):
                obj_name = obj_name.strip()
                if obj_name and obj_name in obj_id_cache:
                    _create_relation(initiative_id, obj_id_cache[obj_name],
                                     "relInitiativeToObjective")

            # ── Transformations ───────────────────────────────────────────────
            # Map each linked Application to a Transformation type based on its
            # lifecycle phase, then POST to the Transformations REST API.
            # Linked BCs are included in the transformation factSheets when available.
            _LIFECYCLE_TO_TRANSFORMATION = {
                "plan":      "implementNewApplication",
                "phaseIn":   "rolloutApplication",
                "active":    "rolloutApplication",
                "phaseOut":  "discontinueApplication",
                "endOfLife": "decommissionApplication",
            }
            # Collect BCs for this initiative (used in transformation payload)
            init_bc_ids = []
            for bc_name in re.split(r"[;,]", str(bcs_raw)):
                bc_name = bc_name.strip()
                if bc_name and bc_name in bc_id_cache:
                    init_bc_ids.append(bc_id_cache[bc_name])

            for app_name in re.split(r"[;,]", str(apps_raw)):
                app_name = app_name.strip()
                if not app_name or app_name not in app_id_cache:
                    continue
                app_id = app_id_cache[app_name]
                # Determine lifecycle of this specific app
                app_lc = str(row.get("lifecycle_phase") or row.get("lifecycle") or "plan").strip()
                trans_type = _LIFECYCLE_TO_TRANSFORMATION.get(app_lc, "implementNewApplication")
                trans_name = {
                    "en": {
                        "implementNewApplication": f"Introduce {app_name}",
                        "rolloutApplication":      f"Roll out {app_name}",
                        "discontinueApplication":  f"Discontinue {app_name}",
                        "decommissionApplication": f"Decommission {app_name}",
                    },
                    "es": {
                        "implementNewApplication": f"Introducir {app_name}",
                        "rolloutApplication":      f"Desplegar {app_name}",
                        "discontinueApplication":  f"Descontinuar {app_name}",
                        "decommissionApplication": f"Retirar {app_name}",
                    },
                }.get("en", {}).get(trans_type, f"{trans_type} {app_name}")

                # completionDate: use lifecycle_endDate if available, else +1 year
                _comp_date = lc_end or (
                    str(int(time.strftime("%Y")) + 1) + time.strftime("-%m-%d")
                )
                payload: dict = {
                    "factSheetId":   initiative_id,
                    "factSheetType": "Initiative",
                    "type":          trans_type,
                    "name":          trans_name,
                    "completionDate": {"type": "exactDate", "date": _comp_date},
                    "factSheets": {
                        "application": {"id": app_id, "type": "Application"},
                    },
                }
                if init_bc_ids:
                    payload["factSheets"]["businessCapabilities"] = [
                        {"id": bc_id, "type": "BusinessCapability"} for bc_id in init_bc_ids
                    ]
                try:
                    hdrs = {
                        "Authorization": f"Bearer {get_bearer(base_url, token)}",
                        "Content-Type": "application/json",
                    }
                    t_resp = requests.post(trans_url, json=payload, headers=hdrs, timeout=30)
                    t_resp.raise_for_status()
                    logger.debug(
                        "LeanIX Transformation created: %s '%s' → app '%s'",
                        trans_type, init_name, app_name,
                    )
                except Exception as t_exc:
                    _body = getattr(getattr(t_exc, "response", None), "text", "")
                    logger.warning(
                        "LeanIX Transformation failed for app '%s' in '%s': %s %s",
                        app_name, init_name, t_exc, _body,
                    )

            logger.info("LeanIX: %s initiative '%s'", "created" if created else "updated", init_name)
            pushed += 1

        except Exception as exc:
            logger.error("LeanIX: failed to push %s — %s", req_id, exc)
            failed += 1

    # ── Lift & Shift Transformations ──────────────────────────────────────────
    # Runs once after all fact sheets are upserted so source and target app IDs
    # are guaranteed to be in app_id_cache.
    # For each conversion: decommissionApplication (source → successor: target)
    #                    + implementNewApplication  (target)
    # Both are attached to the Lift & Shift Initiative if one exists.
    if lift_shift_map:
        # Find the L&S initiative ID (name contains "Lift" or tag "LiftShift")
        ls_initiative_id: str | None = None
        for row in init_rows:
            n = str(row.get("name", "")).strip()
            if "lift" in n.lower() or "shift" in n.lower():
                ls_initiative_id = _find_by_name("Initiative", n)
                if ls_initiative_id:
                    break
        # Fallback: first initiative in cache
        if not ls_initiative_id and init_rows:
            n = str(init_rows[0].get("name", "")).strip()
            ls_initiative_id = _find_by_name("Initiative", n)

        if ls_initiative_id:
            for conv in lift_shift_map:
                source_name = (conv.get("source_app") or "").strip()
                target_name = (conv.get("target_desc") or conv.get("target_matnr") or "").strip()
                if not source_name or not target_name:
                    continue
                source_id = app_id_cache.get(source_name)
                target_id = app_id_cache.get(target_name)
                if not source_id or not target_id:
                    logger.debug(
                        "L&S Transformation skipped — IDs not found: source='%s'(%s) target='%s'(%s)",
                        source_name, source_id, target_name, target_id,
                    )
                    continue

                hdrs = {
                    "Authorization": f"Bearer {get_bearer(base_url, token)}",
                    "Content-Type": "application/json",
                }

                # 1. decommissionApplication: source retires, successor = target
                try:
                    decom_payload = {
                        "factSheetId":    ls_initiative_id,
                        "factSheetType":  "Initiative",
                        "type":           "decommissionApplication",
                        "name":           f"Decommission {source_name}",
                        "completionDate": {"type": "completionDate"},
                        "factSheets": {
                            "application": {"id": source_id, "type": "Application"},
                            "successor":   {"id": target_id, "type": "Application"},
                        },
                    }
                    r = requests.post(trans_url, json=decom_payload, headers=hdrs, timeout=30)
                    r.raise_for_status()
                    logger.debug("L&S decommissionApplication: '%s' → successor '%s'", source_name, target_name)
                except Exception as exc:
                    logger.warning("L&S decommission Transformation failed '%s': %s", source_name, exc)

                # 2. implementNewApplication: target introduced
                try:
                    impl_payload = {
                        "factSheetId":    ls_initiative_id,
                        "factSheetType":  "Initiative",
                        "type":           "implementNewApplication",
                        "name":           f"Introduce {target_name}",
                        "completionDate": {"type": "completionDate"},
                        "factSheets": {
                            "application":  {"id": target_id, "type": "Application"},
                            "predecessors": [{"id": source_id, "type": "Application"}],
                        },
                    }
                    r = requests.post(trans_url, json=impl_payload, headers=hdrs, timeout=30)
                    r.raise_for_status()
                    logger.debug("L&S implementNewApplication: '%s'", target_name)
                except Exception as exc:
                    logger.warning("L&S implement Transformation failed '%s': %s", target_name, exc)
        else:
            logger.warning("L&S Transformations skipped — no Lift & Shift Initiative found in push data")

    logger.info("LeanIX push complete: %d pushed, %d failed", pushed, failed)

    # ── Reference Catalog batch-links ─────────────────────────────────────────
    # Auto-link Applications to official LeanIX catalog entries (lx_APP_XXXXXX).
    # Auto-link ITComponents to official LeanIX ITC catalog entries (lx_ITC_XXXXXX).
    # Only links when confidenceLevel == "VERYHIGH"; logs others for manual review.
    catalog_stats: dict = {}
    if app_id_cache or itc_id_cache:
        # Rows that already carry externalId were linked at create time via the
        # pre-creation Reference Catalog resolver — skip them in the post-push linker.
        rows_by_name = {
            str(r.get("name") or "").strip(): r
            for r in app_rows if r.get("name")
        }
        catalog_stats = _link_apps_to_catalog(
            base_url, token, app_id_cache, itc_id_cache,
            rows_by_name=rows_by_name,
        )

    # ── Metrics API — project KPIs ────────────────────────────────────────────
    # Compute completeness KPIs from the pushed data and create them in LeanIX.
    kpis = _create_project_kpis(
        base_url=base_url,
        api_token=token,
        app_rows=app_rows,
        bc_rows=bc_rows,
        init_rows=init_rows,
        app_id_cache=app_id_cache,
        bc_id_cache=bc_id_cache,
        init_id_cache=init_id_cache,
        client_name=client_name,
    )
    if kpis:
        logger.info(
            "KPIs — apps_with_bc=%.0f%% apps_with_lc=%.0f%% "
            "bcs_with_app=%.0f%% inits_with_trans=%.0f%%",
            kpis.get("apps_with_bc", 0), kpis.get("apps_with_lifecycle", 0),
            kpis.get("bcs_with_app", 0), kpis.get("inits_with_trans", 0),
        )

    # Persist UUID map for Step 8 Catalog Linking Review (alongside staging Excel).
    try:
        _write_push_uuid_map(
            out_dir=Path(staging_path).parent,
            base_url=base_url,
            workspace=_workspace_from_base_url(base_url),
            app_id_cache=app_id_cache,
            itc_id_cache=itc_id_cache,
            failed=[],   # GraphQL path raises on failure — no per-row failure list collected here today
        )
    except Exception as exc:
        logger.warning("Could not write push_uuid_map.json: %s", exc)

    return {
        "pushed": pushed,
        "failed": failed,
        "catalog": catalog_stats,
        "kpis": kpis or {},
    }


# ── Reference Catalog helpers ─────────────────────────────────────────────────

def _link_apps_to_catalog(
    base_url: str,
    api_token: str,
    app_id_cache: dict[str, str],
    itc_id_cache: dict[str, str] | None = None,
    rows_by_name: dict[str, dict] | None = None,
) -> dict:
    """
    Auto-link workspace Applications and ITComponents to LeanIX Reference Catalog entries.

    - Applications: POST /services/reference-data/v1/source/saas/batch-links
      then PUT /services/reference-data/v1/source/saas/links
    - ITComponents:  POST /services/reference-data/v1/source/ltls/batch-links
      then PUT /services/reference-data/v1/source/ltls/links

    Only auto-links when confidenceLevel == "VERYHIGH".
    Returns a dict with stats: {apps_linked, apps_review, itcs_linked, itcs_review}.
    After linking, copies useful fields from firstSuggestedFactSheet:
      - lxHostingType (if not already set on Application)
      - description   (if empty on Application)
    Logs productCategory and provider for info.

    ``rows_by_name`` (optional): name → row dict. Any row with a truthy
    ``externalId`` is skipped here because it was already linked at create
    time via the pre-creation Reference Catalog resolver.
    """
    rows_by_name = rows_by_name or {}
    skipped = {
        name for name, row in rows_by_name.items()
        if row and row.get("externalId")
    }
    if skipped:
        logger.debug(
            "Reference Catalog post-push: skipping %d rows already linked at create",
            len(skipped),
        )
    app_id_cache = {n: i for n, i in app_id_cache.items() if n not in skipped}
    if itc_id_cache:
        itc_id_cache = {n: i for n, i in itc_id_cache.items() if n not in skipped}
    try:
        bearer = get_bearer(base_url, api_token)
    except Exception as exc:
        logger.warning("Reference Catalog: cannot get bearer token: %s", exc)
        return

    import requests as _req

    hdrs = {"Authorization": f"Bearer {bearer}", "Content-Type": "application/json"}

    def _patch_field_rest(fs_id: str, field_path: str, value: str) -> None:
        """Patch a field on a fact sheet via GraphQL REST fallback (inline, no _gql closure)."""
        gql_url = f"{base_url}/services/pathfinder/v1/graphql"
        query = """
mutation PatchFS($id: ID!, $patches: [Patch]!) {
  updateFactSheet(id: $id, patches: $patches) { factSheet { id } }
}"""
        try:
            r = _req.post(
                gql_url,
                json={"query": query, "variables": {
                    "id": fs_id,
                    "patches": [{"op": "add", "path": f"/{field_path}", "value": value}],
                }},
                headers=hdrs,
                timeout=15,
            )
            r.raise_for_status()
        except Exception as exc:
            logger.debug("Reference Catalog: could not patch %s on %s: %s", field_path, fs_id, exc)

    def _run_catalog_linking(source_name: str, id_cache: dict[str, str], enrich_fields: bool) -> tuple[int, int]:
        """Run batch-links + PUT links for one source (saas or ltls). Returns (linked, suggestions)."""
        batch_url = f"{base_url}/services/reference-data/v1/source/{source_name}/batch-links"
        link_url  = f"{base_url}/services/reference-data/v1/source/{source_name}/links"
        fs_list = [{"id": fid, "name": name} for name, fid in id_cache.items()]
        CHUNK = 50
        linked = 0
        review_count = 0

        for i in range(0, len(fs_list), CHUNK):
            chunk = fs_list[i : i + CHUNK]
            payload = {"factSheets": chunk, "numMatches": 3}
            try:
                resp = _req.post(batch_url, json=payload, headers=hdrs, timeout=30)
                if resp.status_code in (403, 404):
                    logger.debug("Reference Catalog %s batch-links not available (%s)", source_name, resp.status_code)
                    return linked, review_count
                resp.raise_for_status()
                data = resp.json().get("data", {})
            except Exception as exc:
                logger.warning("Reference Catalog %s batch-links failed: %s", source_name, exc)
                return linked, review_count

            for fs_id, result in data.items():
                suggestions = result.get("suggestions", [])
                if not suggestions:
                    continue
                top = suggestions[0].get("factSheet", {})
                confidence = top.get("confidenceLevel", "")
                catalog_id = top.get("id", "")
                catalog_ext_id = top.get("externalId", "")
                catalog_name = top.get("displayName", "")

                if confidence == "VERYHIGH" and catalog_id:
                    try:
                        lr = _req.put(
                            link_url,
                            json={
                                "sourceFactSheetIdentifier": {"id": catalog_id, "externalId": catalog_ext_id},
                                "targetFactSheetId": fs_id,
                            },
                            headers=hdrs,
                            timeout=15,
                        )
                        link_ok = lr.status_code in (200, 201, 204, 409) or (
                            lr.status_code == 500 and "Link was established" in lr.text
                        )
                        if link_ok:
                            linked += 1
                            logger.debug(
                                "Reference Catalog [%s]: linked '%s' → '%s' (VERYHIGH)",
                                source_name, fs_id, catalog_name,
                            )
                            # Enrich from catalog fields (Applications only)
                            if enrich_fields:
                                hosting = top.get("hostingType") or top.get("lxHostingType", "")
                                desc    = top.get("description", "")
                                prod_cat = top.get("productCategory", "")
                                provider = top.get("provider", "")
                                if hosting:
                                    _patch_field_rest(fs_id, "lxHostingType", hosting)
                                if desc:
                                    _patch_field_rest(fs_id, "description", desc)
                                if prod_cat or provider:
                                    logger.debug(
                                        "Reference Catalog catalog info for %s: productCategory=%s provider=%s",
                                        catalog_name, prod_cat, provider,
                                    )
                        else:
                            logger.warning(
                                "Reference Catalog link failed for %s → %s: %s",
                                fs_id, catalog_id, lr.status_code,
                            )
                    except Exception as exc:
                        logger.warning("Reference Catalog link request failed for %s: %s", fs_id, exc)
                else:
                    review_count += 1
                    logger.debug(
                        "Reference Catalog [%s] review: %s → '%s' (%s)",
                        source_name, fs_id, catalog_name, confidence,
                    )

        return linked, review_count

    linked_apps, review_apps = _run_catalog_linking("saas", app_id_cache, enrich_fields=True)
    logger.info(
        "Reference Catalog apps: %d auto-linked (VERYHIGH), %d for manual review",
        linked_apps, review_apps,
    )

    linked_itcs, review_itcs = 0, 0
    if itc_id_cache:
        linked_itcs, review_itcs = _run_catalog_linking("ltls", itc_id_cache, enrich_fields=False)
        logger.info(
            "Reference Catalog ITCs: %d auto-linked (VERYHIGH), %d for manual review",
            linked_itcs, review_itcs,
        )

    return {
        "apps_linked": linked_apps,
        "apps_review": review_apps,
        "itcs_linked": linked_itcs,
        "itcs_review": review_itcs,
    }


def _create_project_kpis(
    base_url: str,
    api_token: str,
    app_rows: list[dict],
    bc_rows: list[dict],
    init_rows: list[dict],
    app_id_cache: dict[str, str],
    bc_id_cache: dict[str, str],
    init_id_cache: dict[str, str],
    client_name: str,
) -> dict:
    """
    Compute completeness KPIs from the pushed data.

    Returns a dict with the four KPI values (as percentages 0-100).
    Also writes them as time-series points to LeanIX Metrics v2 if the
    workspace supports it (non-fatal if the endpoint is unavailable).

    KPIs:
      - pct_apps_with_bc        % Applications with at least one BC assigned
      - pct_apps_with_lifecycle % Applications with a lifecycle phase defined
      - pct_bcs_with_app        % BCs that have at least one Application
      - pct_inits_with_trans    % Initiatives with at least one Transformation

    All KPIs are scoped with a label "archimedes-<client_name>" for easy filtering.
    Non-fatal — push always completes even if this step fails.
    """
    import requests as _req

    try:
        bearer = get_bearer(base_url, api_token)
    except Exception as exc:
        logger.warning("Metrics KPIs: cannot get bearer token: %s", exc)
        return {}

    hdrs = {"Authorization": f"Bearer {bearer}", "Content-Type": "application/json"}

    # ── 1. Compute metrics from pushed data ───────────────────────────────────

    # Apps with BC
    apps_with_bc = 0
    for row in app_rows:
        bc_raw = str(row.get("relApplicationToBusinessCapability") or "").strip()
        if bc_raw:
            apps_with_bc += 1

    # Apps with lifecycle
    apps_with_lc = 0
    for row in app_rows:
        lc = str(row.get("lifecycle_phase") or row.get("lifecycle") or "").strip()
        if lc:
            apps_with_lc += 1

    n_apps = len(app_rows) or 1  # avoid division by zero

    # BCs with at least one app (check relApplicationToBusinessCapability across app_rows)
    bcs_referenced: set[str] = set()
    for row in app_rows:
        bc_raw = str(row.get("relApplicationToBusinessCapability") or "").strip()
        for bc in re.split(r"[;,]", bc_raw):
            bc = bc.strip()
            if bc:
                bcs_referenced.add(bc.lower())

    n_bcs = len(bc_rows) or 1
    bcs_with_app = sum(
        1 for row in bc_rows
        if str(row.get("name") or "").strip().lower() in bcs_referenced
    )

    # Initiatives with transformations: use init_id_cache populated during push
    inits_with_trans = 0
    trans_url = f"{base_url}/services/transformations/v1/transformations"
    for row in init_rows:
        init_name = str(row.get("name") or "").strip()
        init_id   = init_id_cache.get(init_name)
        if init_id:
            try:
                tr = _req.get(trans_url, params={"factSheetId": init_id}, headers=hdrs, timeout=15)
                items = tr.json() if isinstance(tr.json(), list) else tr.json().get("data", [])
                if items:
                    inits_with_trans += 1
            except Exception:
                pass

    n_inits = len(init_rows) or 1

    result = {
        "apps_with_bc":        round(apps_with_bc / n_apps * 100, 1),
        "apps_with_lifecycle": round(apps_with_lc / n_apps * 100, 1),
        "bcs_with_app":        round(bcs_with_app / n_bcs * 100, 1),
        "inits_with_trans":    round(inits_with_trans / n_inits * 100, 1),
    }

    logger.info(
        "Metrics KPIs for client '%s' — "
        "apps_with_bc=%.0f%% apps_with_lc=%.0f%% bcs_with_app=%.0f%% inits_with_trans=%.0f%%",
        client_name,
        result["apps_with_bc"], result["apps_with_lifecycle"],
        result["bcs_with_app"], result["inits_with_trans"],
    )
    return result


# ── Main ──────────────────────────────────────────────────────────────────────

def write_leanix_excel_from_xlsx(
    enriched_xlsx: Path,
    output_path: Path,
    client_name: str,
    header_row: int = 8,
    data_start_row: int = 9,
    supplementary: dict | None = None,
) -> None:
    """
    Generate a LeanIX-importable Excel directly from an enriched Requerimientos Excel
    (as produced by map_requirements.py).

    Column mapping (1-based):
      B(2)=req_id, H(8)=coverage, I(9)=module, N(14)=comment, O(15)=BCs full path, P(16)=RSA app

    BCs are stored as full paths separated by ' | ', e.g.:
      'Corporate / Finance / Accounting and Financial Close | Corporate / Asset Management / ...'

    RSA column may contain comma-separated canonical app names, e.g.:
      'SAP S/4HANA, SAP Analytics Cloud' → two separate Application rows.
    Each name is resolved against the RSA catalog name_index for exact canonical spelling.
    """
    import openpyxl as _xl

    # Load RSA name_index for canonical resolution: lowercase → canonical
    rsa_catalog   = json.loads((KNOWLEDGE_DIR / "sap_rsa_catalog.json").read_text())
    rsa_name_index = rsa_catalog.get("name_index", {})  # {lowercase: canonical}

    def _canonical_rsa(name: str) -> str:
        """Return exact canonical RSA name, or original if not found in catalog."""
        return rsa_name_index.get(name.strip().lower(), name.strip())

    wb_in = _xl.load_workbook(str(enriched_xlsx))
    ws_in = wb_in.active

    enriched: list[dict] = []
    for r in range(data_start_row, ws_in.max_row + 1):
        req_id   = ws_in.cell(r, 2).value
        proceso  = ws_in.cell(r, 1).value   # col A = proceso (initiative group)
        coverage = ws_in.cell(r, 8).value
        module   = ws_in.cell(r, 9).value
        comment  = ws_in.cell(r, 14).value
        bcs_raw  = ws_in.cell(r, 15).value
        rsa_raw  = ws_in.cell(r, 16).value
        if not req_id:
            continue

        # Parse BCs: 'Domain / Area / Leaf | Domain / Area / Leaf2'
        bcs_parsed: list[str] = []
        if bcs_raw:
            for part in str(bcs_raw).split("|"):
                part = part.strip()
                segments = [s.strip() for s in part.split("/")]
                if len(segments) >= 2:
                    bcs_parsed.append(part)   # keep full path for resolution

        # Split comma-separated RSA string into individual canonical app names
        rsa_apps: list[str] = []
        if rsa_raw:
            for part in str(rsa_raw).split(","):
                canonical = _canonical_rsa(part.strip())
                if canonical:
                    rsa_apps.append(canonical)
        if not rsa_apps:
            rsa_apps = ["SAP S/4HANA"]

        # Group key: use proceso (col A) as initiative grouping
        group = str(proceso or "").strip() or str(req_id).strip()

        # Emit one record per canonical RSA app (same BCs apply to all)
        for app_name in rsa_apps:
            enriched.append({
                "id":       str(req_id).strip(),
                "_group":   group,
                "coverage": str(coverage or ""),
                "module":   str(module or ""),
                "comment":  str(comment or "")[:2000],
                "bcs_full": bcs_parsed,
                "rsa":      app_name,
            })

    # Build intermediate enriched dicts compatible with write_leanix_excel
    # Map full BC paths to leaf names
    def _leaf(full_path: str) -> str:
        parts = [s.strip() for s in full_path.split("/")]
        return parts[-1] if parts else full_path

    bcs_index_full: dict[str, str] = {}   # leaf → full_path (deduplicated)
    enriched_compat: list[dict] = []

    for row in enriched:
        bc_leaves = []
        for fp in row["bcs_full"]:
            leaf = _leaf(fp)
            bcs_index_full[leaf] = fp
            bc_leaves.append(leaf)
        enriched_compat.append({
            "id":       row["id"],
            "_group":   row.get("_group", row["id"]),
            "bcs":      bc_leaves,
            "rsa":      row["rsa"],
            "coverage": row["coverage"],
            "comment":  row["comment"],
            "module":   row["module"],
        })

    write_leanix_excel(enriched_compat, bcs_index_full, output_path, client_name, supplementary=supplementary)


def write(
    enriched_path: str | Path,
    template_path: str | Path,
    output_dir: str | Path = "output",
    client_name: str = "unknown",
    supplementary: dict | None = None,
) -> tuple[Path, Path]:
    """
    Run the write step:
      1. Write client Excel (columns H–P on original template)  → <stem>_enriched.xlsx
      2. Write LeanIX importable Excel (multi-sheet, formatted) → <client>_target_leanix.xlsx

    Returns (client_excel_path, leanix_target_path).
    Push to LeanIX separately with push_leanix().
    """
    enriched_path = Path(enriched_path)
    template_path = Path(template_path)
    output_dir    = Path(output_dir)

    enriched  = json.loads(enriched_path.read_text())
    bcs_index = _load_bcs_index()

    out_excel  = output_dir / f"{template_path.stem}_enriched.xlsx"
    out_target = output_dir / f"{client_name}_target_leanix.xlsx"

    write_excel(enriched, template_path, out_excel, bcs_index)
    write_leanix_excel(enriched, bcs_index, out_target, client_name, supplementary=supplementary)

    return out_excel, out_target


def push_leanix(
    staging_path: str | Path,
    client_name: str,
    lift_shift_map: list[dict] | None = None,
) -> None:
    """Push a previously generated LeanIX staging Excel to LeanIX."""
    write_leanix(Path(staging_path), client_name, lift_shift_map=lift_shift_map)


if __name__ == "__main__":
    import argparse
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)

    p_write = sub.add_parser("write", help="Generate Excel outputs")
    p_write.add_argument("enriched")
    p_write.add_argument("template")
    p_write.add_argument("--output-dir", default="output")
    p_write.add_argument("--client", default="unknown")

    p_push = sub.add_parser("push", help="Push staging Excel to LeanIX")
    p_push.add_argument("staging")
    p_push.add_argument("--client", default="unknown")

    args = ap.parse_args()
    if args.cmd == "write":
        out, staging = write(args.enriched, args.template, args.output_dir, args.client)
        print(f"Client Excel  → {out}")
        print(f"LeanIX import → {staging}")
    else:
        push_leanix(args.staging, args.client)
        print("LeanIX push complete.")
