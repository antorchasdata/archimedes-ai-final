"""
pipeline/push_ldif.py — Alternative LeanIX push via Integration API (LDIF).

Instead of N individual GraphQL mutations, this module builds a single LDIF
payload and calls POST /services/integration-api/v1/synchronizationRuns/withConfig.

Benefits over GraphQL push (write.py):
  - One HTTP call for the full dataset
  - Full/Partial modes with server-side deduplication via externalId
  - Built-in synclog — errors visible in LeanIX Integration Hub
  - Better suited for >100 fact sheets

Usage (CLI):
    python3 -m pipeline.push_ldif --staging output/<client>/<client>_target_leanix.xlsx \\
                                  --client <name> [--mode partial]

Usage (programmatic):
    from pipeline.push_ldif import push_leanix_ldif
    push_leanix_ldif(staging_path, client_name, mode="partial")
"""

from __future__ import annotations

import json
import logging
import os
import re
from pathlib import Path

import requests
from dotenv import load_dotenv

from pipeline.leanix_auth import get_bearer

load_dotenv()
logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
CONNECTOR_TYPE    = "archimedes-ai"
CONNECTOR_ID      = "archimedes-push"
CONNECTOR_VERSION = "1.0.0"

_LIFECYCLE_MAP = {
    "plan":      "plan",
    "phaseIn":   "phaseIn",
    "active":    "active",
    "phaseOut":  "phaseOut",
    "endOfLife": "endOfLife",
}

# ── Excel reader (mirrors write.py sheet reading logic) ───────────────────────

def _read_sheet(wb, sheet_name: str, aliases: list[str] | None = None) -> list[dict]:
    """
    Read a sheet from an openpyxl workbook.
    Row 1 = header, rows 2+ = data.
    Returns list of {header: value} dicts, skipping blank rows.
    """
    import openpyxl as xl

    target = sheet_name
    if target not in wb.sheetnames:
        for alias in (aliases or []):
            if alias in wb.sheetnames:
                target = alias
                break
        else:
            return []

    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        result.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return result


# ── LDIF builders ─────────────────────────────────────────────────────────────

def _ext_id(client_name: str, fs_type: str, name: str) -> dict:
    """Build externalId block — stable across pushes for idempotency."""
    safe = re.sub(r"[^a-zA-Z0-9_\-]", "_", name)
    return {
        "externalId":   f"archimedes-{client_name}-{fs_type.lower()}-{safe}",
        "externalType": CONNECTOR_TYPE,
    }


def _resolve_uuids_for_ldif(
    base_url: str,
    bearer: str,
    client_name: str,
    app_names: list[str],
    itc_names: list[str],
) -> tuple[dict[str, str], dict[str, str], list[dict]]:
    """After an LDIF push, fetch FS UUIDs by externalId.

    Returns (app_id_cache, itc_id_cache, failed) where:
      - app_id_cache: {app_name → workspace UUID}
      - itc_id_cache: {itc_name → workspace UUID}
      - failed:        list of {name, type, error} for rows not found
    """
    base = (base_url or "").rstrip("/")
    url = f"{base}/services/pathfinder/v1/factSheets"
    hdrs = {"Authorization": f"Bearer {bearer}"}

    app_cache: dict[str, str] = {}
    itc_cache: dict[str, str] = {}
    failed: list[dict] = []

    def _lookup(name: str, fs_type_short: str, ext_id_block: dict) -> str | None:
        params = {
            "externalId.externalId": ext_id_block["externalId"],
            "externalId.externalType": ext_id_block["externalType"],
        }
        try:
            resp = requests.get(url, headers=hdrs, params=params, timeout=30)
            if resp.status_code != 200:
                return None
            data = (resp.json() or {}).get("data") or []
            if not data:
                return None
            return data[0].get("id")
        except Exception:
            return None

    for name in app_names:
        ext = _ext_id(client_name, "APP", name)
        uid = _lookup(name, "Application", ext)
        if uid:
            app_cache[name] = uid
        else:
            failed.append({"name": name, "type": "Application", "error": "not found by externalId"})

    for name in itc_names:
        ext = _ext_id(client_name, "ITC", name)
        uid = _lookup(name, "ITComponent", ext)
        if uid:
            itc_cache[name] = uid
        else:
            failed.append({"name": name, "type": "ITComponent", "error": "not found by externalId"})

    return app_cache, itc_cache, failed


def _persist_ldif_uuid_map(
    staging_path: Path,
    ldif: dict,
    client_name: str,
    base_url: str,
    bearer: str,
) -> Path:
    """Resolve UUIDs for all App/ITC entries in the LDIF and write push_uuid_map.json."""
    app_names = [c["data"]["name"] for c in ldif["content"] if c.get("type") == "Application"]
    itc_names = [c["data"]["name"] for c in ldif["content"] if c.get("type") == "ITComponent"]

    apps, itcs, failed = _resolve_uuids_for_ldif(
        base_url=base_url, bearer=bearer, client_name=client_name,
        app_names=app_names, itc_names=itc_names,
    )

    from urllib.parse import urlparse
    workspace = (urlparse(base_url).hostname or "").split(".")[0]
    entries: dict[str, dict] = {}
    for name, uid in apps.items():
        entries[f"Application::{name}"] = {"uuid": uid, "created": True}
    for name, uid in itcs.items():
        entries[f"ITComponent::{name}"] = {"uuid": uid, "created": True}
    payload = {
        "workspace": workspace,
        "base_url": base_url,
        "entries": entries,
        "failed": failed,
    }
    out_path = Path(staging_path).parent / "push_uuid_map.json"
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False))
    return out_path


def build_create_factsheet_payload(row: dict) -> dict:
    """Build a GraphQL createFactSheet input from a row dict.

    When the row carries a catalog ``externalId`` (e.g. ``lx_APP_000123``),
    include it so the fact sheet is created already linked to the
    Reference Catalog. Otherwise, omit it.
    """
    payload: dict = {
        "name": row.get("name"),
        "type": row.get("type"),
    }
    if row.get("externalId"):
        payload["externalId"] = row["externalId"]
    if row.get("description"):
        payload["description"] = row["description"]
    return payload


def _build_bc_entry(row: dict, client_name: str) -> dict:
    name   = str(row.get("name") or "").strip()
    parent = str(row.get("relToParent") or "").strip()
    entry: dict = {
        "type": "BusinessCapability",
        "id":   {"externalId": _ext_id(client_name, "BC", name)},
        "data": {"name": name},
    }
    if parent:
        entry["relations"] = {
            "relToParent": [{"externalId": _ext_id(client_name, "BC", parent)}]
        }
    return entry


def _build_itc_entry(row: dict, client_name: str) -> dict:
    name    = str(row.get("name") or "").strip()
    hosting = str(row.get("lxDeploymentType") or row.get("lxHostingType") or "").strip()
    entry: dict = {
        "type": "ITComponent",
        "id":   {"externalId": _ext_id(client_name, "ITC", name)},
        "data": {"name": name},
    }
    if hosting:
        entry["data"]["lxDeploymentType"] = hosting
    return entry


def _build_app_entry(row: dict, client_name: str) -> dict:
    name     = str(row.get("name") or "").strip()
    lc_phase = str(row.get("lifecycle_phase") or row.get("lifecycle") or "").strip()
    hosting  = str(row.get("lxHostingType") or "").strip()
    tags_raw = str(row.get("tags") or "").strip()

    data: dict = {"name": name}
    if lc_phase and lc_phase in _LIFECYCLE_MAP:
        data["lifecycle"] = {"phases": [{"phase": _LIFECYCLE_MAP[lc_phase]}]}
    if hosting:
        data["lxHostingType"] = hosting

    entry: dict = {
        "type": "Application",
        "id":   {"externalId": _ext_id(client_name, "APP", name)},
        "data": data,
    }

    # Tags via tag field (semicolon-separated)
    tags = [t.strip() for t in tags_raw.split(";") if t.strip()]
    if tags:
        entry["data"]["tags"] = tags

    # Relations
    relations: dict = {}
    bc_raw  = str(row.get("relApplicationToBusinessCapability") or "").strip()
    itc_raw = str(row.get("relApplicationToITComponent") or "").strip()

    if bc_raw:
        relations["relApplicationToBusinessCapability"] = [
            {"externalId": _ext_id(client_name, "BC", bc.strip())}
            for bc in re.split(r"[;,]", bc_raw) if bc.strip()
        ]
    if itc_raw:
        relations["relApplicationToITComponent"] = [
            {"externalId": _ext_id(client_name, "ITC", itc.strip())}
            for itc in re.split(r"[;,]", itc_raw) if itc.strip()
        ]

    if relations:
        entry["relations"] = relations

    return entry


def _build_initiative_entry(row: dict, client_name: str) -> dict:
    name     = str(row.get("name") or "").strip()
    lc_phase = str(row.get("lifecycle_phase") or row.get("lifecycle") or "").strip()
    apps_raw = str(row.get("relInitiativeToApplication") or row.get("rsa") or "").strip()
    bcs_raw  = str(row.get("relInitiativeToBusinessCapability") or "").strip()

    data: dict = {"name": name}
    if lc_phase and lc_phase in _LIFECYCLE_MAP:
        data["lifecycle"] = {"phases": [{"phase": _LIFECYCLE_MAP[lc_phase]}]}

    entry: dict = {
        "type": "Initiative",
        "id":   {"externalId": _ext_id(client_name, "INIT", name)},
        "data": data,
    }

    relations: dict = {}
    if apps_raw:
        relations["relInitiativeToApplication"] = [
            {"externalId": _ext_id(client_name, "APP", a.strip())}
            for a in re.split(r"[;,]", apps_raw) if a.strip()
        ]
    if bcs_raw:
        relations["relInitiativeToBusinessCapability"] = [
            {"externalId": _ext_id(client_name, "BC", bc.strip())}
            for bc in re.split(r"[;,]", bcs_raw) if bc.strip()
        ]
    if relations:
        entry["relations"] = relations

    return entry


# ── Main builder ──────────────────────────────────────────────────────────────

def build_ldif(
    staging_path: Path,
    client_name: str,
    mode: str = "partial",
) -> dict:
    """
    Build a complete LDIF payload from a LeanIX staging Excel.

    Args:
        staging_path: Path to <client>_target_leanix.xlsx
        client_name:  Client name used as externalId prefix
        mode:         "partial" (default) or "full"
                      "full"    → deletes fact sheets not present in this push
                      "partial" → upserts only; safe for incremental updates

    Returns the LDIF dict ready to POST.
    """
    import openpyxl as xl

    wb = xl.load_workbook(str(staging_path), read_only=True, data_only=True)

    bc_rows   = _read_sheet(wb, "BusinessCapability", ["BusinessCapabilities"])
    itc_rows  = _read_sheet(wb, "ITComponent",        ["ITComponents"])
    app_rows  = _read_sheet(wb, "Application",        ["Applications"])
    init_rows = _read_sheet(wb, "Initiative",         ["Initiatives"])

    content = []

    # Order: BC → ITC → Application → Initiative (dependency order)
    for row in bc_rows:
        if row.get("name"):
            content.append(_build_bc_entry(row, client_name))

    for row in itc_rows:
        if row.get("name"):
            content.append(_build_itc_entry(row, client_name))

    for row in app_rows:
        if row.get("name"):
            content.append(_build_app_entry(row, client_name))

    for row in init_rows:
        if row.get("name"):
            content.append(_build_initiative_entry(row, client_name))

    ldif = {
        "connectorType":        CONNECTOR_TYPE,
        "connectorId":          CONNECTOR_ID,
        "connectorVersion":     CONNECTOR_VERSION,
        "processingDirection":  "inbound",
        "processingMode":       mode,
        "content":              content,
    }

    logger.info(
        "LDIF built: %d BCs + %d ITCs + %d Apps + %d Initiatives (mode=%s)",
        len(bc_rows), len(itc_rows), len(app_rows), len(init_rows), mode,
    )
    return ldif


# ── Push ──────────────────────────────────────────────────────────────────────

def push_leanix_ldif(
    staging_path: str | Path,
    client_name: str,
    mode: str = "partial",
    base_url: str | None = None,
    api_token: str | None = None,
    save_ldif: bool = False,
) -> dict:
    """
    Push a LeanIX staging Excel to LeanIX via the Integration API (LDIF).

    This is an alternative to push_leanix() in write.py. Use it when:
      - The dataset has >100 fact sheets (single HTTP call is more robust)
      - You want built-in synclog tracking in LeanIX Integration Hub
      - You need "full" mode to clean up stale fact sheets

    Args:
        staging_path: Path to the staging Excel
        client_name:  Client name (used as externalId prefix)
        mode:         "partial" (default) or "full"
        base_url:     LeanIX base URL (defaults to LEANIX_BASE_URL env var)
        api_token:    LeanIX API token (defaults to LEANIX_API_TOKEN env var)
        save_ldif:    If True, saves the LDIF JSON next to the staging Excel

    Returns:
        {ok, run_id, status, stats, warnings}
    """
    _base_url  = base_url  or os.environ.get("LEANIX_BASE_URL", "")
    _api_token = api_token or os.environ.get("LEANIX_API_TOKEN", "")

    if not _base_url or not _api_token:
        raise ValueError("LEANIX_BASE_URL and LEANIX_API_TOKEN must be set")

    staging_path = Path(staging_path)
    ldif = build_ldif(staging_path, client_name, mode=mode)

    if save_ldif:
        ldif_path = staging_path.with_suffix(".ldif.json")
        ldif_path.write_text(json.dumps(ldif, indent=2, ensure_ascii=False))
        logger.info("LDIF saved to %s", ldif_path)

    bearer  = get_bearer(_base_url, _api_token)
    hdrs    = {"Authorization": f"Bearer {bearer}", "Content-Type": "application/json"}
    url     = f"{_base_url}/services/integration-api/v1/synchronizationRuns/withConfig"

    logger.info("Pushing LDIF (%d entries) to %s …", len(ldif["content"]), url)
    resp = requests.post(url, json=ldif, headers=hdrs, timeout=120)

    try:
        data = resp.json()
    except Exception:
        data = {"raw": resp.text}

    if resp.status_code not in (200, 201, 202):
        logger.error("LDIF push failed: %s — %s", resp.status_code, resp.text[:500])
        return {"ok": False, "status_code": resp.status_code, "detail": resp.text[:500]}

    run_id   = data.get("id") or data.get("runId") or ""
    status   = data.get("status", "")
    stats    = data.get("stats") or {}
    warnings = data.get("warnings") or []

    logger.info(
        "LDIF push complete: run_id=%s status=%s stats=%s warnings=%d",
        run_id, status, stats, len(warnings),
    )
    if warnings:
        for w in warnings[:5]:
            logger.warning("  LDIF warning: %s", w)

    # Capture UUID map for Step 8 Catalog Linking Review.
    try:
        _persist_ldif_uuid_map(
            staging_path=staging_path,
            ldif=ldif,
            client_name=client_name,
            base_url=_base_url,
            bearer=bearer,
        )
    except Exception as exc:
        logger.warning("Could not write push_uuid_map.json: %s", exc)

    return {
        "ok":       True,
        "run_id":   run_id,
        "status":   status,
        "stats":    stats,
        "warnings": warnings,
    }


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    import sys

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s %(name)s — %(message)s")

    parser = argparse.ArgumentParser(description="Push LeanIX staging Excel via Integration API (LDIF)")
    parser.add_argument("--staging", required=True, help="Path to <client>_target_leanix.xlsx")
    parser.add_argument("--client",  required=True, help="Client name")
    parser.add_argument("--mode",    default="partial", choices=["partial", "full"],
                        help="LDIF processing mode (default: partial)")
    parser.add_argument("--save-ldif", action="store_true",
                        help="Save generated LDIF JSON next to the staging Excel")
    args = parser.parse_args()

    result = push_leanix_ldif(
        staging_path=args.staging,
        client_name=args.client,
        mode=args.mode,
        save_ldif=args.save_ldif,
    )
    print(json.dumps(result, indent=2))
    sys.exit(0 if result.get("ok") else 1)
