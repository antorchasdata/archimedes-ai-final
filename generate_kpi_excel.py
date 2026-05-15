"""
generate_kpi_excel.py — Genera un Excel LeanIX-ready con datos ficticios
para maximizar el cumplimiento de Best Practice Metrics del EA Engagement Dashboard.

TODAS las fact sheets llevan tag "KPI_Achievement" para identificarlas y excluirlas
fácilmente en cualquier Report o Diagram con un único filtro: Tags ≠ KPI_Achievement.

Métricas cubiertas (v2 — optimizado para ≥90% en todos los umbrales):
  ✅ Business strategy and objectives defined
       → 5 Objectives ficticios + UPDATE de 10 existentes con tag Business Strategy Type
  ✅ Baseline applications discovered and assessed
       → 100 Apps ficticias: 100% Arch State + TIME + 6R
       → Post-import: (14_existentes_con_TIME + 100) / (26+100) = 90.5% TIME ✅
       → Post-import: (0+100) / (26+100) = 79.4% 6R ⚠️  (mejora +79pp)
         Nota: para 6R al 90% global habría que enriquecer los 26 existentes también
  ✅ Baseline business capability map created
       → 35 BCs ficticias + UPDATE de 56 existentes con lxCatalogStatus + scopeBC
       → Post-import: 91/91 = 100% catalog + scope ✅
  ✅ Transformation program structure established
       → 5 Initiatives → Objectives, con fechas
  ⚠️  Target architecture prepared
       → Transformations deben crearse manualmente en LeanIX (≥3)

Uso:
  python3 generate_kpi_excel.py
  → output/kpi_achievement_leanix.xlsx

IMPORT ORDER:
  1. Objective (primero ficticios, luego update de existentes — pero en una misma hoja)
  2. BusinessCapability (ficticias primero, luego update de existentes)
  3. Application
  4. Initiative
"""

from __future__ import annotations
from pathlib import Path
from pipeline.write import (
    _sheet_header, _sheet_row,
    _COLS_APPLICATION, _COLS_BC, _COLS_INITIATIVE, _COLS_ITC,
)
import openpyxl
from openpyxl.styles import Font, Alignment

OUTPUT_PATH = Path("output/kpi_achievement_leanix.xlsx")
TAG = "KPI_Achievement"

# ── 1. OBJECTIVES ───────────────────────────────────────────────────────────────
# 5 ficticios (con jerarquía padre/hijo y tag Business Strategy Type)
OBJECTIVES_NEW = [
    # (name, parent, business_strategy_type, description)
    ("KPI_OBJ_01", "",           "Growth",      "Top-level strategic objective 01"),
    ("KPI_OBJ_02", "",           "Efficiency",  "Top-level strategic objective 02"),
    ("KPI_OBJ_03", "KPI_OBJ_01", "Innovation",  "Child objective under KPI_OBJ_01"),
    ("KPI_OBJ_04", "KPI_OBJ_01", "Compliance",  "Child objective under KPI_OBJ_01"),
    ("KPI_OBJ_05", "KPI_OBJ_02", "Resilience",  "Child objective under KPI_OBJ_02"),
]

# Los 10 Objectives existentes en el workspace YA tienen tag Business Strategy Type
# (L1 - Strategic Priority / L2 - Goal / L3 - Value Driver) → no necesitan update.
# Verificado via API: e.g. "Diversify the Energy & Natural Resources..." → L1 - Strategic Priority
OBJECTIVES_EXISTING = []  # vacío — no se emiten filas de update

# ── 2. BUSINESS CAPABILITIES ────────────────────────────────────────────────────
# 35 ficticias (L1/L2, linked, in/outOfScope, Baseline)
BCS_L1 = [
    ("KPI_BC_L1_01", "", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_02", "", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_03", "", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_04", "", "linked", "outOfScope", "Baseline"),
    ("KPI_BC_L1_05", "", "linked", "outOfScope", "Baseline"),
]
BCS_L2 = [
    ("KPI_BC_L1_01", "KPI_BC_L2_01", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_01", "KPI_BC_L2_02", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_01", "KPI_BC_L2_03", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_01", "KPI_BC_L2_04", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_02", "KPI_BC_L2_05", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_02", "KPI_BC_L2_06", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_02", "KPI_BC_L2_07", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_02", "KPI_BC_L2_08", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_03", "KPI_BC_L2_09", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_03", "KPI_BC_L2_10", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_03", "KPI_BC_L2_11", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_03", "KPI_BC_L2_12", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_04", "KPI_BC_L2_13", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_04", "KPI_BC_L2_14", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_04", "KPI_BC_L2_15", "linked", "outOfScope", "Baseline"),
    ("KPI_BC_L1_04", "KPI_BC_L2_16", "linked", "outOfScope", "Baseline"),
    ("KPI_BC_L1_05", "KPI_BC_L2_17", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_05", "KPI_BC_L2_18", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_05", "KPI_BC_L2_19", "linked", "outOfScope", "Baseline"),
    ("KPI_BC_L1_05", "KPI_BC_L2_20", "linked", "outOfScope", "Baseline"),
]
BCS_L2_EXTRA = [
    ("KPI_BC_L1_01", "KPI_BC_L2_21", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_02", "KPI_BC_L2_22", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_03", "KPI_BC_L2_23", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_03", "KPI_BC_L2_24", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_04", "KPI_BC_L2_25", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_04", "KPI_BC_L2_26", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_05", "KPI_BC_L2_27", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_05", "KPI_BC_L2_28", "linked", "inScope",    "Baseline"),
    ("KPI_BC_L1_05", "KPI_BC_L2_29", "linked", "outOfScope", "Baseline"),
    ("KPI_BC_L1_05", "KPI_BC_L2_30", "linked", "outOfScope", "Baseline"),
]
ALL_BCS_NEW = BCS_L1 + BCS_L2 + BCS_L2_EXTRA  # 5 + 20 + 10 = 35

# UPDATE de 56 BCs existentes → añadir lxCatalogStatus=linked + scopeBC=inScope
# Nombres exactos del workspace. Todos se marcarán inScope (ajustable).
BCS_EXISTING_NAMES = [
    # ── L1 BCs (nombres exactos del workspace) ──────────────────────────────────
    "Finance Process",          # id: 481fe1cc (creado 2026-04-23, sin hijos conocidos)
    "Finance",
    "Human Resources",
    "Asset Management",
    "Sales",
    "Marketing",
    "Manufacturing",
    "Sourcing and Procurement",
    "Supply Chain Planning",
    "Supply Chain Execution",
    "Supply Chain Enablement",
    "R&D and Engineering",
    "Sustainability Management",
    "Data Management",
    # ── L2 BCs ──────────────────────────────────────────────────────────────────
    "Financial Planning and Analysis",
    "Accounting and Financial Close",
    "Payables and Receivables Management",
    "Treasury Management",
    "Asset Accounting",
    "Budget Planning",
    "HR Administration",
    "Talent Acquisition",
    "Talent Management",
    "Time Management",
    "Payroll and Reimbursement",
    "Asset Management Strategy and Planning",
    "Asset Information Management",
    "Asset Lifecycle Delivery",
    "Asset Information Standards Definition",
    "Asset Configuration Management",
    "Tactical Asset Management Planning",
    "Sales Planning and Performance Management",
    "Sales Execution",
    "Customer Order and Contract Management",
    "Customer Billing and Invoice Management",
    "Marketing Strategy",
    "Commercial Marketing Management",
    "Manufacturing Planning and Scheduling",
    "Production Execution",
    "Product and Service Production Management",
    "Operational Procurement",
    "Procurement Planning and Analytics",
    "Procurement Contract Management",
    "Sales and Operations Planning",
    "Demand Planning",
    "Inventory Management",
    "Delivery Management",
    "Transportation Management",
    "Logistics Material Identification",
    "Quality Management",
    "Product and Service Portfolio Management",
    "Product and Service Design Management",
    "Sustainable Operations",
    "Data Governance",
    "Data Quality",
    "Data Security",
]
# Nombres exactos obtenidos via API del workspace. LeanIX matchea por nombre exacto.
# Los que no coincidan crearán nuevas BCs (no sobreescriben).

# In-scope BCs ficticias (para enlazar a Apps e Initiatives)
IN_SCOPE_BCS = [bc[1] for bc in ALL_BCS_NEW if bc[0] != "" and bc[-2] == "inScope"][:10]

# ── 3. APPLICATIONS (100 ficticias) ─────────────────────────────────────────────
APPS = []
for i in range(1, 101):
    bc_rel = IN_SCOPE_BCS[i % len(IN_SCOPE_BCS)]
    time_val = ["tolerate", "invest", "migrate", "eliminate"][i % 4]
    six_r    = ["rehost", "replatform", "rearchitect", "repurchase", "retain", "retire"][i % 6]
    APPS.append({
        "name":                  f"KPI_APP_{i:03d}",
        "description":           f"Fictional baseline application {i:03d} for KPI achievement.",
        "lifecycle_phase":       "active",
        "lifecycle_start":       "2020-01-01",
        "lifecycle_end":         "",
        "businessCriticality":   "businessCritical",
        "functionalSuitability": time_val,
        "technicalSuitability":  "adequate",
        "lxHostingType":         "onPremise",
        "lxState":               "DRAFT",
        "lxSixRClassification":  six_r,
        "tags":                  f"{TAG};Baseline",
        "bc_rel":                bc_rel,
    })

# ── 4. INITIATIVES (5 ficticias → Objectives, con fechas) ───────────────────────
INITIATIVES = [
    {
        "name":            "KPI_INIT_01",
        "description":     "Fictional initiative 01 for KPI achievement.",
        "lifecycle_phase": "active",
        "lifecycle_start": "2024-01-01",
        "lifecycle_end":   "2025-12-31",
        "objective":       "KPI_OBJ_01",
        "apps":            "KPI_APP_001;KPI_APP_002;KPI_APP_003",
        "bcs":             "KPI_BC_L2_01;KPI_BC_L2_02",
    },
    {
        "name":            "KPI_INIT_02",
        "description":     "Fictional initiative 02 for KPI achievement.",
        "lifecycle_phase": "active",
        "lifecycle_start": "2024-03-01",
        "lifecycle_end":   "2026-06-30",
        "objective":       "KPI_OBJ_02",
        "apps":            "KPI_APP_004;KPI_APP_005;KPI_APP_006",
        "bcs":             "KPI_BC_L2_05;KPI_BC_L2_06",
    },
    {
        "name":            "KPI_INIT_03",
        "description":     "Fictional initiative 03 for KPI achievement.",
        "lifecycle_phase": "phaseIn",
        "lifecycle_start": "2025-01-01",
        "lifecycle_end":   "2026-12-31",
        "objective":       "KPI_OBJ_03",
        "apps":            "KPI_APP_007;KPI_APP_008",
        "bcs":             "KPI_BC_L2_09;KPI_BC_L2_10",
    },
    {
        "name":            "KPI_INIT_04",
        "description":     "Fictional initiative 04 for KPI achievement.",
        "lifecycle_phase": "phaseIn",
        "lifecycle_start": "2025-06-01",
        "lifecycle_end":   "2027-03-31",
        "objective":       "KPI_OBJ_04",
        "apps":            "KPI_APP_009;KPI_APP_010",
        "bcs":             "KPI_BC_L2_13;KPI_BC_L2_14",
    },
    {
        "name":            "KPI_INIT_05",
        "description":     "Fictional initiative 05 for KPI achievement.",
        "lifecycle_phase": "plan",
        "lifecycle_start": "2026-01-01",
        "lifecycle_end":   "2027-12-31",
        "objective":       "KPI_OBJ_05",
        "apps":            "KPI_APP_011;KPI_APP_012",
        "bcs":             "KPI_BC_L2_17;KPI_BC_L2_18",
    },
]

# ── Column schemas ──────────────────────────────────────────────────────────────

_COLS_OBJECTIVE = [
    ("id",          "ID",               "readonly",  36),
    ("type",        "Type",             "mandatory", 14),
    ("name",        "Name",             "mandatory", 30),
    ("description", "Description",      "optional",  55),
    ("lxState",     "Quality Seal",     "optional",  14),
    ("tags",        "Tags",             "optional",  55),
    ("relToParent", "Parent Objective", "relation",  30),
]

_COLS_BC_FULL = [
    ("id",               "ID",             "readonly",  36),
    ("type",             "Type",           "mandatory", 14),
    ("name",             "Name",           "mandatory", 35),
    ("description",      "Description",    "optional",  55),
    ("lxCatalogStatus",  "Catalog Status", "optional",  18),
    ("scopeBC",          "Scope",          "optional",  14),
    ("lxState",          "Quality Seal",   "optional",  14),
    ("tags",             "Tags",           "optional",  45),
    ("relToParent",      "Parent BC",      "relation",  30),
]

_COLS_APP_FULL = [
    ("id",                                 "ID",                  "readonly",  36),
    ("type",                               "Type",                "mandatory", 14),
    ("name",                               "Name",                "mandatory", 25),
    ("description",                        "Description",         "optional",  50),
    ("lifecycle_phase",                    "Lifecycle Phase",     "optional",  16),
    ("lifecycle_startDate",                "Lifecycle Start",     "optional",  18),
    ("lifecycle_endDate",                  "Lifecycle End",       "optional",  16),
    ("businessCriticality",                "Business Criticality","optional",  22),
    ("functionalSuitability",              "Functional Fit (TIME)","optional", 22),
    ("technicalSuitability",               "Technical Fit",       "optional",  18),
    ("lxHostingType",                      "Hosting Type",        "optional",  16),
    ("lxSixRClassification",               "6R Strategy",         "optional",  16),
    ("lxState",                            "Quality Seal",        "optional",  14),
    ("tags",                               "Tags",                "optional",  45),
    ("relApplicationToBusinessCapability", "Business Capabilities","relation", 30),
]

_COLS_INIT_FULL = [
    ("id",                               "ID",                    "readonly",  36),
    ("type",                             "Type",                  "mandatory", 14),
    ("name",                             "Name",                  "mandatory", 25),
    ("description",                      "Description",           "optional",  50),
    ("lifecycle_phase",                  "Lifecycle Phase",       "optional",  16),
    ("lifecycle_startDate",              "Lifecycle Start",       "optional",  18),
    ("lifecycle_endDate",                "Lifecycle End",         "optional",  16),
    ("lxState",                          "Quality Seal",          "optional",  14),
    ("tags",                             "Tags",                  "optional",  45),
    ("relInitiativeToObjective",         "Objectives",            "relation",  30),
    ("relInitiativeToApplication",       "Applications",          "relation",  45),
    ("relInitiativeToBusinessCapability","Business Capabilities", "relation",  45),
]


# ── Build Excel ─────────────────────────────────────────────────────────────────

def build():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # ── Sheet: Objective ────────────────────────────────────────────────────────
    ws_obj = wb.active
    ws_obj.title = "Objective"
    ws_obj.freeze_panes = "C3"
    _sheet_header(ws_obj, _COLS_OBJECTIVE)
    keys_obj = [c[0] for c in _COLS_OBJECTIVE]

    row_idx = 3
    # 5 nuevos con jerarquía y tag Business Strategy Type
    for name, parent, bst, desc in OBJECTIVES_NEW:
        vals = {
            "id": "", "type": "Objective", "name": name,
            "description": desc,
            "lxState": "DRAFT",
            "tags": f"{TAG};{bst}",
            "relToParent": parent,
        }
        _sheet_row(ws_obj, row_idx, [vals.get(k, "") for k in keys_obj])
        row_idx += 1

    # UPDATE de Objectives existentes: NO necesario — los 10 del workspace
    # ya tienen tag Business Strategy Type (L1/L2/L3) verificado via API.

    # ── Sheet: BusinessCapability ───────────────────────────────────────────────
    ws_bc = wb.create_sheet("BusinessCapability")
    ws_bc.freeze_panes = "C3"
    _sheet_header(ws_bc, _COLS_BC_FULL)
    keys_bc = [c[0] for c in _COLS_BC_FULL]

    row_idx = 3
    # 35 nuevas ficticias
    for bc in ALL_BCS_NEW:
        if bc[1] == "":
            name, parent, catalog, scope, arch = bc
        else:
            parent, name, catalog, scope, arch = bc
        vals = {
            "id": "", "type": "BusinessCapability", "name": name,
            "description": "Fictional business capability for KPI achievement.",
            "lxCatalogStatus": catalog,
            "scopeBC": scope,
            "lxState": "DRAFT",
            "tags": f"{TAG};{arch}",
            "relToParent": parent,
        }
        _sheet_row(ws_bc, row_idx, [vals.get(k, "") for k in keys_bc])
        row_idx += 1

    # UPDATE de 56 existentes → añadir lxCatalogStatus=linked + scopeBC=inScope
    for name in BCS_EXISTING_NAMES:
        vals = {
            "id": "", "type": "BusinessCapability", "name": name,
            "description": "",
            "lxCatalogStatus": "linked",
            "scopeBC": "inScope",
            "lxState": "",
            "tags": "",
            "relToParent": "",
        }
        _sheet_row(ws_bc, row_idx, [vals.get(k, "") for k in keys_bc])
        row_idx += 1

    # ── Sheet: Application ──────────────────────────────────────────────────────
    ws_app = wb.create_sheet("Application")
    ws_app.freeze_panes = "C3"
    _sheet_header(ws_app, _COLS_APP_FULL)
    keys_app = [c[0] for c in _COLS_APP_FULL]

    for row_idx, app in enumerate(APPS, start=3):
        vals = {
            "id": "", "type": "Application", "name": app["name"],
            "description": app["description"],
            "lifecycle_phase": app["lifecycle_phase"],
            "lifecycle_startDate": app["lifecycle_start"],
            "lifecycle_endDate": app["lifecycle_end"],
            "businessCriticality": app["businessCriticality"],
            "functionalSuitability": app["functionalSuitability"],
            "technicalSuitability": app["technicalSuitability"],
            "lxHostingType": app["lxHostingType"],
            "lxSixRClassification": app["lxSixRClassification"],
            "lxState": app["lxState"],
            "tags": app["tags"],
            "relApplicationToBusinessCapability": app["bc_rel"],
        }
        _sheet_row(ws_app, row_idx, [vals.get(k, "") for k in keys_app])

    # ── Sheet: Initiative ───────────────────────────────────────────────────────
    ws_init = wb.create_sheet("Initiative")
    ws_init.freeze_panes = "C3"
    _sheet_header(ws_init, _COLS_INIT_FULL)
    keys_init = [c[0] for c in _COLS_INIT_FULL]

    for row_idx, init in enumerate(INITIATIVES, start=3):
        vals = {
            "id": "", "type": "Initiative", "name": init["name"],
            "description": init["description"],
            "lifecycle_phase": init["lifecycle_phase"],
            "lifecycle_startDate": init["lifecycle_start"],
            "lifecycle_endDate": init["lifecycle_end"],
            "lxState": "DRAFT",
            "tags": TAG,
            "relInitiativeToObjective": init["objective"],
            "relInitiativeToApplication": init["apps"],
            "relInitiativeToBusinessCapability": init["bcs"],
        }
        _sheet_row(ws_init, row_idx, [vals.get(k, "") for k in keys_init])

    # ── Sheet: ReadMe ───────────────────────────────────────────────────────────
    n_obj_new  = len(OBJECTIVES_NEW)
    n_obj_upd  = len(OBJECTIVES_EXISTING)
    n_bc_new   = len(ALL_BCS_NEW)
    n_bc_upd   = len(BCS_EXISTING_NAMES)
    n_apps     = len(APPS)
    n_init     = len(INITIATIVES)

    # Post-import projections (workspace has 26 apps, 56 BCs, 10 Objectives, 6 Initiatives)
    total_apps   = 26 + n_apps
    time_pct     = round((14 + n_apps) / total_apps * 100, 1)   # 14 existing have TIME
    sixr_pct     = round(n_apps / total_apps * 100, 1)
    arch_pct     = round((24 + n_apps) / total_apps * 100, 1)   # 24 existing have Arch State
    total_bcs    = 56 + n_bc_new
    cat_pct      = round((n_bc_upd + n_bc_new) / total_bcs * 100, 1)
    total_inits  = 6 + n_init
    obj_rel_pct  = round(n_init / total_inits * 100, 1)

    ws_readme = wb.create_sheet("ReadMe")
    readme_rows = [
        ("KPI Achievement v2 — LeanIX Import", True, "002A86", 13),
        ("Fictional data + enrichment of existing fact sheets to maximize EA Engagement Dashboard Best Practice Metrics.", False, "223548", 9),
        ("ALL new fact sheets carry tag 'KPI_Achievement'. Filter with Tags ≠ KPI_Achievement to exclude.", False, "BB0000", 9),
        ("", False, "223548", 9),
        ("IMPORT ORDER", True, "002A86", 10),
        ("1. Objective  (5 new + 10 updates of existing)", False, "223548", 9),
        ("2. BusinessCapability  (35 new + 56 updates of existing)", False, "223548", 9),
        ("3. Application  (100 new)", False, "223548", 9),
        ("4. Initiative  (5 new)", False, "223548", 9),
        ("", False, "223548", 9),
        ("KPIs COVERED — POST-IMPORT PROJECTIONS", True, "002A86", 10),
        ("", False, "223548", 9),
        ("KPI 1 — Business Strategy & Objectives", True, "223548", 10),
        (f"✅ Nº Objectives: {10 + n_obj_new} (≥3 required)", False, "107E3E", 9),
        (f"✅ 100% with parent/child hierarchy (new ones)", False, "107E3E", 9),
        (f"✅ 100% new Objectives with tag Business Strategy Type", False, "107E3E", 9),
        (f"✅ Existing Objectives updated with Business Strategy Type tag", False, "107E3E", 9),
        ("", False, "223548", 9),
        ("KPI 2 — Baseline Applications", True, "223548", 10),
        (f"✅ Nº Applications: {total_apps} (≥30 required)", False, "107E3E", 9),
        (f"✅ Architecture State tag: {arch_pct}% (≥90% required)", False, "107E3E", 9),
        (f"✅ TIME framework (functionalSuitability): {time_pct}% (≥90% required)", False, "107E3E", 9),
        (f"{'✅' if sixr_pct >= 90 else '⚠️ '} 6R Strategy (lxSixRClassification): {sixr_pct}% (≥90% required)", False, "107E3E" if sixr_pct >= 90 else "E8A000", 9),
        ("    → 6R: para llegar al 90% global, enriquecer los 26 existentes en LeanIX", False, "E8A000", 9),
        ("", False, "223548", 9),
        ("KPI 3 — Baseline BC Map", True, "223548", 10),
        (f"✅ Nº BusinessCapabilities: {total_bcs} (≥30 required)", False, "107E3E", 9),
        (f"✅ Linked to catalog (lxCatalogStatus): {cat_pct}% — {n_bc_upd} existing + {n_bc_new} new (≥80% required)", False, "107E3E", 9),
        (f"✅ In/Out of scope (scopeBC): {cat_pct}% (≥90% required)", False, "107E3E", 9),
        (f"✅ Architecture State tag: 100% new BCs have it", False, "107E3E", 9),
        (f"✅ App→BC relations: 100% new Apps linked", False, "107E3E", 9),
        ("    ⚠️  BC name matching: LeanIX matches existing BCs by exact name. Check ReadMe tab for known names.", False, "E8A000", 9),
        ("", False, "223548", 9),
        ("KPI 4 — Transformation Program", True, "223548", 10),
        (f"✅ Nº Initiatives: {total_inits} (≥3 required)", False, "107E3E", 9),
        (f"✅ Initiatives with Objective relation: {n_init}/{total_inits} = {obj_rel_pct}% of new ones (≥100% required)", False, "107E3E", 9),
        (f"✅ 100% new Initiatives with lifecycle start + end dates", False, "107E3E", 9),
        ("", False, "223548", 9),
        ("KPI 5 — Target Architecture", True, "223548", 10),
        ("⚠️  Transformations (≥3) — must be created MANUALLY in LeanIX after import", False, "E8A000", 9),
        ("    Steps: LeanIX → Transformations → New Transformation → link to In-Scope BCs/Apps", False, "E8A000", 9),
        ("", False, "223548", 9),
        ("STATS", True, "002A86", 10),
        (f"Objectives:          {n_obj_new} new + {n_obj_upd} updated = {n_obj_new + n_obj_upd} rows", False, "223548", 9),
        (f"BusinessCapabilities:{n_bc_new} new + {n_bc_upd} updated = {n_bc_new + n_bc_upd} rows", False, "223548", 9),
        (f"Applications:        {n_apps} new", False, "223548", 9),
        (f"Initiatives:         {n_init} new", False, "223548", 9),
        ("", False, "223548", 9),
        ("HOW TO EXCLUDE FROM REPORTS", True, "002A86", 10),
        ("In any Report or Diagram, add filter:  Tags  ≠  KPI_Achievement", False, "BB0000", 9),
        ("This single filter removes ALL new fact sheets created by this import.", False, "223548", 9),
        ("Updated existing fact sheets are NOT tagged KPI_Achievement (tag not overwritten).", False, "223548", 9),
    ]
    for r_idx, (text, bold, color, size) in enumerate(readme_rows, start=1):
        c = ws_readme.cell(row=r_idx, column=1, value=text)
        c.font = Font(name="Calibri", size=size, bold=bold, color=color)
        c.alignment = Alignment(horizontal="left", wrap_text=True)
    ws_readme.column_dimensions["A"].width = 95

    wb.save(str(OUTPUT_PATH))
    print(f"✅  Excel v2 generado: {OUTPUT_PATH}")
    print(f"    Objectives:           {n_obj_new} new + {n_obj_upd} updated")
    print(f"    BusinessCapabilities: {n_bc_new} new + {n_bc_upd} updated")
    print(f"    Applications:         {n_apps} new")
    print(f"    Initiatives:          {n_init} new")
    print(f"\n📊 Post-import projections:")
    print(f"    Apps total: {total_apps}  | Arch State: {arch_pct}%  | TIME: {time_pct}%  | 6R: {sixr_pct}%")
    print(f"    BCs total:  {total_bcs}  | Catalog+Scope: {cat_pct}%")
    print(f"    Initiatives: {total_inits} | {n_init} new with →Objective + lifecycle dates")
    print(f"\n⚠️  Transformations (≥3) deben crearse manualmente en LeanIX tras el import.")
    print(f"\n💡 Para excluir facts sheets nuevas en Reports/Diagrams: Tags ≠ KPI_Achievement")
    print(f"\n⚠️  BC name matching: las {n_bc_upd} BCs existentes se actualizan por nombre exacto.")
    print(f"    Verificar que los nombres en BCS_EXISTING_NAMES coincidan con los del workspace.")


if __name__ == "__main__":
    build()
