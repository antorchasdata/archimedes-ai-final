"""
archimedes_wizard.py — FastAPI backend for the Archimedes AI step-by-step wizard.

Endpoints:
  GET  /                                  → serve archimedes_wizard.html
  GET  /api/config                        → feature flags (leanix configured?)
  POST /api/session                       → create session (Step 0: client name)
  GET  /api/session/{id}/catalog          → catalog status (Step 1)
  POST /api/catalogs/refresh              → refresh RBA/RSA catalogs from MXP
  POST /api/session/{id}/baseline         → baseline generation (Step 2)
  POST /api/session/{id}/lift-shift/resolve → Lift & Shift: resolve app names → Material Numbers (Step 2b)
  POST /api/session/{id}/lift-shift/convert → Lift & Shift: convert SKUs to RISE targets (Step 2b)
  POST /api/session/{id}/requirements     → requirements Excel (Step 3)
  POST /api/session/{id}/contrast         → SAP Help Portal contrast (Step 3b, optional)
  POST /api/session/{id}/pdf              → PDF extraction (Step 4)
  POST /api/session/{id}/images           → image extraction (Step 5)
  POST /api/session/{id}/generate         → generate LeanIX outputs (Step 6)
  POST /api/session/{id}/push             → import to LeanIX (Step 7)
  POST /api/session/{id}/push-kpi        → 🍪 easter egg: generate + import KPI Achievement Excel
  GET  /api/session/{id}/download/{key}   → download generated file

Usage:
    python3 archimedes_wizard.py
    → http://localhost:8767
"""

from __future__ import annotations

import asyncio
import contextlib
import json
import logging
import os
import shutil
import subprocess
import sys
import uuid
import webbrowser
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse

load_dotenv()

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
WIZARD_HTML = BASE_DIR / "archimedes_wizard.html"
OUTPUT_DIR  = BASE_DIR / "output" / "sessions"
WORKSPACES_PATH = BASE_DIR / "workspaces.json"

sys.path.insert(0, str(BASE_DIR))

# ── Pipeline imports ──────────────────────────────────────────────────────────
from pipeline.catalog       import _catalog_info, RBA_PATH, RSA_PATH
from pipeline.footprint     import generate_baseline
from pipeline.extract       import extract
from pipeline.enrich        import enrich
from pipeline.validate      import validate
from pipeline.write         import write, write_leanix_excel_from_xlsx, push_leanix
from pipeline.push_ldif    import push_leanix_ldif
from pipeline.pdf_extract   import extract_pdf_factsheets
from pipeline.image_extract import extract_image_factsheets
from pipeline.help_contrast      import run_contrast, print_contrast_summary
from pipeline.lift_shift    import (
    get_material_session, check_session,
    resolve_app_to_skus, get_deployment_modes,
    convert_to_rise, resolve_prerequisites,
)
from pipeline.industry_reference import (
    get_industry_reference, compute_whitespace,
    add_reference_to_target_excel, derive_relations_for_products, INDUSTRIES
)

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s %(name)s — %(message)s")
logger = logging.getLogger("archimedes.wizard")

# ── Anthropic client (shared) ─────────────────────────────────────────────────
_anthropic_client = None

def _get_anthropic():
    global _anthropic_client
    if _anthropic_client is None:
        import anthropic
        _anthropic_client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    return _anthropic_client

MODEL = os.getenv("ENRICH_MODEL", "claude-sonnet-4-6")

# ── Session store ─────────────────────────────────────────────────────────────
# session_id → dict with pipeline state
_sessions: dict[str, dict] = {}

def _session(session_id: str) -> dict:
    if session_id not in _sessions:
        raise HTTPException(status_code=404, detail=f"Session {session_id!r} not found")
    return _sessions[session_id]


def _leanix_creds(sess: dict) -> tuple[str, str]:
    """Return (base_url, api_token) from session, falling back to env vars."""
    base_url  = sess.get("leanix_base_url")  or os.environ.get("LEANIX_BASE_URL", "")
    api_token = sess.get("leanix_api_token") or os.environ.get("LEANIX_API_TOKEN", "")
    return base_url, api_token


@contextlib.contextmanager
def _env_override(env: dict):
    """Temporarily set env vars, restoring originals on exit."""
    old = {k: os.environ.get(k) for k in env}
    os.environ.update(env)
    try:
        yield
    finally:
        for k, v in old.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v


# ── FastAPI app ───────────────────────────────────────────────────────────────
app = FastAPI(title="Archimedes Wizard", docs_url=None, redoc_url=None)


# ── Static ────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def serve_wizard():
    return WIZARD_HTML.read_text(encoding="utf-8")


@app.get("/pro", response_class=HTMLResponse)
async def serve_wizard_pro():
    p = BASE_DIR / "archimedes_wizard_pro.html"
    return p.read_text(encoding="utf-8")


@app.get("/cookie_monster.png")
async def serve_cookie_monster():
    p = BASE_DIR / "cookie_monster.png"
    return FileResponse(path=str(p), media_type="image/png")


# ── Config ────────────────────────────────────────────────────────────────────

@app.get("/api/config")
async def get_config():
    leanix_ok = bool(os.environ.get("LEANIX_API_TOKEN") and os.environ.get("LEANIX_BASE_URL"))
    anthropic_ok = bool(os.environ.get("ANTHROPIC_API_KEY"))
    return {"ok": True, "leanix_configured": leanix_ok, "anthropic_configured": anthropic_ok, "model": MODEL}


# ── Workspaces ────────────────────────────────────────────────────────────────

def _load_workspaces() -> list[dict]:
    if not WORKSPACES_PATH.exists():
        return []
    return json.loads(WORKSPACES_PATH.read_text(encoding="utf-8")).get("workspaces", [])


def _save_workspaces(workspaces: list[dict]) -> None:
    WORKSPACES_PATH.write_text(
        json.dumps({"workspaces": workspaces}, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


@app.get("/api/workspaces")
async def list_workspaces():
    ws = _load_workspaces()
    return {"ok": True, "workspaces": [{"name": w["name"], "base_url": w["base_url"]} for w in ws]}


@app.post("/api/workspaces")
async def save_workspace(body: dict):
    name      = (body.get("name") or "").strip()
    api_token = (body.get("api_token") or "").strip()
    # Strip any path/query — only scheme+host is valid for the API
    from urllib.parse import urlparse as _urlparse
    _raw     = (body.get("base_url") or "").strip()
    _p       = _urlparse(_raw)
    base_url = f"{_p.scheme}://{_p.netloc}" if _p.netloc else _raw
    if not name or not base_url or not api_token:
        raise HTTPException(status_code=400, detail="name, base_url and api_token are required")
    ws = _load_workspaces()
    # Upsert by name
    ws = [w for w in ws if w["name"] != name]
    ws.append({"name": name, "base_url": base_url, "api_token": api_token})
    _save_workspaces(ws)
    return {"ok": True, "name": name}


@app.post("/api/workspaces/validate")
async def validate_workspace(body: dict):
    name      = (body.get("name") or "").strip()
    base_url  = (body.get("base_url") or "").strip()
    api_token = (body.get("api_token") or "").strip()

    # If only name provided, look up token from workspaces.json
    if name and not api_token:
        ws = _load_workspaces()
        match = next((w for w in ws if w["name"] == name), None)
        if not match:
            return {"ok": False, "detail": f"Workspace '{name}' not found"}
        base_url  = match["base_url"]
        api_token = match["api_token"]

    if not base_url or not api_token:
        return {"ok": False, "detail": "base_url and api_token are required"}

    import requests as _req
    try:
        from pipeline.leanix_auth import get_bearer
        bearer = await asyncio.to_thread(get_bearer, base_url, api_token)
        hdrs = {"Authorization": f"Bearer {bearer}", "Content-Type": "application/json"}
        gql_url = f"{base_url}/services/pathfinder/v1/graphql"
        payload = {"query": '{ allFactSheets(filter:{facetFilters:[{facetKey:"FactSheetTypes",keys:["Application"]}]}) { totalCount } }'}
        r = await asyncio.to_thread(
            lambda: _req.post(gql_url, json=payload, headers=hdrs, timeout=15)
        )
        r.raise_for_status()
        n_apps = r.json().get("data", {}).get("allFactSheets", {}).get("totalCount", 0)
    except Exception as exc:
        return {"ok": False, "detail": str(exc)}

    return {"ok": True, "workspace_name": name or base_url, "n_applications": n_apps}



@app.post("/api/session")
async def create_session(body: dict):
    client_name = (body.get("client_name") or "").strip()
    if not client_name:
        raise HTTPException(status_code=400, detail="client_name is required")

    # Resolve workspace credentials — by name (preferred) or direct values
    leanix_base_url  = body.get("leanix_base_url") or None
    leanix_api_token = body.get("leanix_api_token") or None
    ws_name = (body.get("leanix_workspace") or "").strip()
    if ws_name and not leanix_api_token:
        ws_list = _load_workspaces()
        match = next((w for w in ws_list if w["name"] == ws_name), None)
        if match:
            leanix_base_url  = match["base_url"]
            leanix_api_token = match["api_token"]

    session_id = str(uuid.uuid4())
    output_dir = OUTPUT_DIR / session_id
    output_dir.mkdir(parents=True, exist_ok=True)

    _sessions[session_id] = {
        "client_name":       client_name,
        "output_dir":        output_dir,
        "leanix_base_url":   leanix_base_url,
        "leanix_api_token":  leanix_api_token,
        "baseline_result":   None,
        "req_excel_path":    None,
        "req_enriched_xlsx": None,
        "target_json_path":  None,
        "pdf_factsheets":    None,
        "image_factsheets":  None,
        "out_baseline":      None,
        "out_target":        None,
        "out_supplementary": None,
        "lift_shift_result": None,
    }

    logger.info("Session created: %s  client=%s  workspace=%s", session_id, client_name, ws_name or "env")
    return {"ok": True, "session_id": session_id, "client_name": client_name}


# ── Step 1 — Catalog status ───────────────────────────────────────────────────

@app.get("/api/session/{session_id}/catalog")
async def catalog_status(session_id: str):
    _session(session_id)  # validate session exists

    result = {}
    for label, path in [("rba", RBA_PATH), ("rsa", RSA_PATH)]:
        if not path.exists():
            result[label] = {"error": f"Not found: {path}"}
        else:
            result[label] = _catalog_info(path)

    return {"ok": True, **result}


@app.get("/api/catalog/rba")
async def catalog_rba_full():
    if not RBA_PATH.exists():
        raise HTTPException(status_code=404, detail="RBA catalog not found")
    return json.loads(RBA_PATH.read_text(encoding="utf-8"))


@app.get("/api/catalog/rsa")
async def catalog_rsa_full():
    if not RSA_PATH.exists():
        raise HTTPException(status_code=404, detail="RSA catalog not found")
    return json.loads(RSA_PATH.read_text(encoding="utf-8"))


@app.post("/api/catalogs/refresh")
async def catalogs_refresh():
    """
    Runs pipeline/catalog_mxp.py to refresh RBA/RSA catalogs from MXP REST API.
    Returns updated stats on success, or a helpful error message on failure (e.g. 401).
    """
    script = BASE_DIR / "pipeline" / "catalog_mxp.py"
    if not script.exists():
        return JSONResponse({"ok": False, "message": "catalog_mxp.py not found"}, status_code=500)

    try:
        result = subprocess.run(
            [sys.executable, str(script)],
            capture_output=True, text=True, timeout=120,
        )
    except subprocess.TimeoutExpired:
        return JSONResponse({"ok": False, "message": "Timeout: catalog update took >120s"}, status_code=504)

    if result.returncode != 0:
        output = (result.stdout + result.stderr).strip()
        if "401" in output or "Token" in output or "token" in output:
            msg = ("Token MXP inválido o no configurado. "
                   "Ejecuta /update-catalogs en Claude Code para actualizar via MCP session.")
        else:
            msg = f"Error actualizando catálogos: {output[:400]}"
        return JSONResponse({"ok": False, "message": msg})

    # Success — read updated stats
    updated_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    rba_info = _catalog_info(RBA_PATH) if RBA_PATH.exists() else {}
    rsa_info = _catalog_info(RSA_PATH) if RSA_PATH.exists() else {}
    return {
        "ok": True,
        "updated_at": updated_at,
        "rba": rba_info,
        "rsa": rsa_info,
    }


# ── Step 2 — Baseline / Footprint ─────────────────────────────────────────────

@app.post("/api/session/{session_id}/baseline")
async def run_baseline(
    session_id: str,
    onprem_file: Optional[UploadFile] = File(None),
    cloud_file:  Optional[UploadFile] = File(None),
):
    sess = _session(session_id)

    if not onprem_file and not cloud_file:
        return {"ok": True, "skipped": True, "n_total": 0}

    out_dir = sess["output_dir"]
    onprem_path = cloud_path = None

    if onprem_file and onprem_file.filename:
        onprem_path = out_dir / "onprem_systems.xlsx"
        onprem_path.write_bytes(await onprem_file.read())

    if cloud_file and cloud_file.filename:
        cloud_path = out_dir / "cloud_systems.xlsx"
        cloud_path.write_bytes(await cloud_file.read())

    if not onprem_path and not cloud_path:
        return {"ok": True, "skipped": True, "n_total": 0}

    client_name  = sess["client_name"]
    baseline_out = out_dir / f"{client_name}_baseline.xlsx"

    try:
        result = await asyncio.to_thread(
            generate_baseline,
            output_path=baseline_out,
            client_name=client_name,
            onprem_path=onprem_path,
            cloud_path=cloud_path,
        )
    except Exception as exc:
        logger.exception("Baseline error")
        raise HTTPException(status_code=500, detail=str(exc))

    sess["baseline_result"] = result
    sess["out_baseline"]    = baseline_out

    return {
        "ok":           True,
        "skipped":      False,
        "n_onprem":     result["n_onprem"],
        "n_cloud":      result["n_cloud"],
        "n_total":      result["n_total"],
        "download_url": f"/api/session/{session_id}/download/baseline",
    }


# ── Step 2 — Register pre-built baseline (admin/debug) ────────────────────────

@app.post("/api/session/{session_id}/baseline/register")
async def register_baseline(session_id: str, body: dict):
    """Register an existing baseline Excel file into the session (bypasses generation)."""
    sess = _session(session_id)
    file_path = Path(body.get("file_path", ""))
    if not file_path.exists():
        raise HTTPException(status_code=400, detail=f"File not found: {file_path}")

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        n_total = ws.max_row - 2  # header rows
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not open Excel: {exc}")

    sess["out_baseline"]    = file_path
    sess["baseline_result"] = {"n_onprem": 0, "n_cloud": n_total, "n_total": n_total}
    return {
        "ok":           True,
        "registered":   str(file_path),
        "n_total":      n_total,
        "download_url": f"/api/session/{session_id}/download/baseline",
    }


# ── Step 2b — Lift & Shift: Resolve ───────────────────────────────────────────

@app.post("/api/session/{session_id}/lift-shift/resolve")
async def lift_shift_resolve(session_id: str):
    """
    Phase 1: Read app names from the baseline Excel, resolve each to a Material Number
    via the SAP Material Mapping API (Claude picks best candidate when multiple exist).
    Also returns the union of available deployment modes across all resolved SKUs.
    """
    sess = _session(session_id)
    baseline_path = sess.get("out_baseline")
    if not baseline_path or not Path(baseline_path).exists():
        raise HTTPException(status_code=400, detail="Baseline not generated yet — complete Step 2 first.")

    # Extract app names from baseline Excel (col C = name, starting row 3)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(baseline_path)
        ws = wb["Applications"]
        app_names = [
            row[2] for row in ws.iter_rows(min_row=3, values_only=True)
            if row[2] and str(row[2]).strip()
        ]
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not read baseline Excel: {exc}")

    if not app_names:
        raise HTTPException(status_code=400, detail="No applications found in the baseline Excel.")

    # Build material session (Chrome cookies)
    mat_session = get_material_session()
    if not check_session(mat_session):
        raise HTTPException(
            status_code=503,
            detail="SAP Material Mapping Tool session expired — open Chrome with the app and retry."
        )

    # Resolve names → SKUs (Claude picks best candidate)
    try:
        resolved = await asyncio.to_thread(resolve_app_to_skus, app_names, mat_session)
    except Exception as exc:
        logger.exception("Lift & Shift resolve error")
        raise HTTPException(status_code=500, detail=str(exc))

    # Get deployment modes across all resolved SKUs
    skus = [r["selected"] for r in resolved if r["selected"]]
    try:
        modes = await asyncio.to_thread(get_deployment_modes, skus, mat_session)
    except Exception as exc:
        logger.exception("Lift & Shift deployment modes error")
        modes = []

    # Store mat_session is not serialisable; store resolved list in session
    sess["lift_shift_resolved"] = resolved
    sess["lift_shift_modes"]    = modes

    return {
        "ok":       True,
        "resolved": resolved,
        "modes":    modes,
        "n_apps":   len(app_names),
        "n_found":  sum(1 for r in resolved if r["selected"]),
        "n_not_found": sum(1 for r in resolved if not r["selected"]),
    }


# ── Step 2b — Lift & Shift: Convert ───────────────────────────────────────────

@app.post("/api/session/{session_id}/lift-shift/convert")
async def lift_shift_convert(session_id: str, body: dict):
    """
    Phase 2: Given user-confirmed mappings and chosen deployment mode,
    run the conversion and return target materials + prerequisites.

    Body: {
      "mappings": [{"app_name": str, "selected": str (matnr)}],
      "deployment_mode": str
    }
    """
    sess = _session(session_id)
    mappings        = body.get("mappings", [])
    deployment_mode = body.get("deployment_mode", "").strip()

    if not mappings:
        raise HTTPException(status_code=400, detail="mappings is required.")
    if not deployment_mode:
        raise HTTPException(status_code=400, detail="deployment_mode is required.")

    mat_session = get_material_session()
    if not check_session(mat_session):
        raise HTTPException(
            status_code=503,
            detail="SAP Material Mapping Tool session expired — open Chrome with the app and retry."
        )

    try:
        conversions = await asyncio.to_thread(convert_to_rise, mappings, deployment_mode, mat_session)
    except Exception as exc:
        logger.exception("Lift & Shift convert error")
        raise HTTPException(status_code=500, detail=str(exc))

    # Resolve all unique prerequisite SKUs to names
    all_prereq_skus = list({sku for c in conversions for sku in c["prereq_list"]})
    try:
        prereq_names = await asyncio.to_thread(resolve_prerequisites, all_prereq_skus, mat_session)
        prereq_map   = {p["matnr"]: p["maktx"] for p in prereq_names}
    except Exception:
        prereq_map = {}

    # Enrich conversions with resolved prereq names
    for c in conversions:
        c["prereq_resolved"] = [
            {"matnr": sku, "maktx": prereq_map.get(sku, sku)}
            for sku in c["prereq_list"]
        ]

    sess["lift_shift_result"] = {
        "deployment_mode": deployment_mode,
        "conversions":     conversions,
    }

    return {
        "ok":          True,
        "conversions": conversions,
        "n_converted": len(conversions),
        "deployment_mode": deployment_mode,
    }


# ── Step 3 — Requirements Excel ───────────────────────────────────────────────

@app.post("/api/session/{session_id}/requirements")
async def run_requirements(
    session_id: str,
    req_file: Optional[UploadFile] = File(None),
    no_validate: bool = Form(False),
):
    sess = _session(session_id)

    if not req_file or not req_file.filename:
        return {"ok": True, "skipped": True}

    out_dir     = sess["output_dir"]
    client_name = sess["client_name"]
    req_path    = out_dir / req_file.filename
    req_path.write_bytes(await req_file.read())
    sess["req_excel_path"] = req_path

    # Detect if already enriched (cols O/P populated from row 9)
    import openpyxl as _xl
    wb_check = _xl.load_workbook(str(req_path))
    ws_check = wb_check.active
    already_enriched = any(
        ws_check.cell(r, 15).value or ws_check.cell(r, 16).value
        for r in range(9, min(15, ws_check.max_row + 1))
    )

    if already_enriched:
        sess["req_enriched_xlsx"] = req_path
        n_rows = sum(
            1 for r in range(9, ws_check.max_row + 1)
            if ws_check.cell(r, 2).value
        )
        return {
            "ok":              True,
            "skipped":         False,
            "already_enriched": True,
            "n_requirements":  n_rows,
            "validation_ok":   True,
        }

    # Run extract → enrich → (validate) pipeline
    try:
        raw_path = await asyncio.to_thread(extract, req_path, out_dir)
        enriched_path = await asyncio.to_thread(enrich, raw_path, out_dir)

        validation_ok       = True
        validation_warnings = 0
        if not no_validate:
            validation_ok = await asyncio.to_thread(validate, enriched_path, out_dir)

        sess["target_json_path"] = enriched_path

        enriched_data = json.loads(enriched_path.read_text())
        n_requirements = len([r for r in enriched_data if not r.get("_error")])

        return {
            "ok":               True,
            "skipped":          False,
            "already_enriched": False,
            "n_requirements":   n_requirements,
            "validation_ok":    validation_ok,
        }

    except Exception as exc:
        logger.exception("Requirements error")
        raise HTTPException(status_code=500, detail=str(exc))


# ── Step 3b — SAP Help Portal Contrast (optional) ────────────────────────────

@app.post("/api/session/{session_id}/contrast")
async def run_help_contrast(session_id: str):
    sess = _session(session_id)

    target_json      = sess.get("target_json_path")
    req_enriched_xlsx = sess.get("req_enriched_xlsx")
    out_dir          = sess["output_dir"]

    # If pre-enriched xlsx (no JSON), build a minimal JSON from cols O/P for contrast
    if not target_json and req_enriched_xlsx:
        import openpyxl as _xl
        wb = _xl.load_workbook(str(req_enriched_xlsx))
        ws = wb.active
        reqs = []
        for r in range(9, ws.max_row + 1):
            req_id = ws.cell(r, 2).value
            rsa    = ws.cell(r, 16).value  # col P
            desc   = ws.cell(r, 3).value or ""
            if req_id and rsa:
                reqs.append({"_source_id": str(req_id), "rsa": str(rsa), "description": str(desc)})
        if not reqs:
            return {"ok": True, "skipped": True, "reason": "No RSA products found in Excel"}
        tmp_json = out_dir / "reqs_for_contrast.json"
        tmp_json.write_text(json.dumps(reqs, ensure_ascii=False, indent=2))
        target_json = tmp_json

    if not target_json:
        return {"ok": True, "skipped": True, "reason": "No enriched requirements available"}

    try:
        report_path = await asyncio.to_thread(run_contrast, target_json, out_dir)
        sess["out_contrast"] = report_path

        report      = json.loads(report_path.read_text())
        validated   = sum(1 for r in report if r["validated"])
        unvalidated = sum(1 for r in report if not r["validated"] and r["rsa_product"])
        skipped     = sum(1 for r in report if not r["rsa_product"])

        return {
            "ok":            True,
            "skipped":       False,
            "n_total":       len(report),
            "n_validated":   validated,
            "n_unvalidated": unvalidated,
            "n_skipped":     skipped,
            "download_url":  f"/api/session/{session_id}/download/contrast",
        }
    except Exception as exc:
        logger.exception("SAP Help contrast error")
        raise HTTPException(status_code=500, detail=str(exc))


# ── Step 3c — Industry Reference Architecture / Whitespace ───────────────────

@app.get("/api/industries")
async def list_industries():
    return {"ok": True, "industries": [{"key": k, "label": v} for k, v in INDUSTRIES.items()]}


@app.post("/api/session/{session_id}/industry-reference")
async def run_industry_reference(session_id: str, body: dict):
    sess = _session(session_id)

    industry_key = (body.get("industry_key") or "").strip()
    if not industry_key:
        raise HTTPException(status_code=400, detail="industry_key is required")

    # Fetch reference products from SAP API Hub
    try:
        reference = await asyncio.to_thread(get_industry_reference, industry_key)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logger.exception("Industry reference fetch error")
        raise HTTPException(status_code=500, detail=str(exc))

    # Compute whitespace against baseline AND target
    baseline_path = sess.get("out_baseline")
    target_path   = sess.get("out_target")
    whitespace = await asyncio.to_thread(
        compute_whitespace, reference["products"], baseline_path, target_path
    )

    sess["industry_reference"] = {
        "industry_key":   industry_key,
        "industry_label": reference["industry_label"],
        "whitespace":     whitespace,
    }

    return {
        "ok":             True,
        "industry_key":   industry_key,
        "industry_label": reference["industry_label"],
        "n_reference":    reference["n_products"],
        "whitespace":     whitespace,
    }


@app.post("/api/session/{session_id}/industry-reference/apply")
async def apply_industry_reference(session_id: str, body: dict):
    sess = _session(session_id)

    selected: list[str] = body.get("selected_products", [])
    if not selected:
        return {"ok": True, "skipped": True, "n_added": 0}

    ref = sess.get("industry_reference", {})
    industry_label = ref.get("industry_label", body.get("industry_label", "Reference"))
    out_dir = sess["output_dir"]
    client_name = sess["client_name"]

    # Create or reuse target Excel
    out_target = sess.get("out_target")
    if not out_target or not out_target.exists():
        # Create a minimal target Excel
        import openpyxl as xl
        out_target = out_dir / f"{client_name}_target_leanix.xlsx"
        wb = xl.Workbook()
        ws = wb.active
        ws.title = "Application"
        ws.append(["type", "displayName", "description", "tags"])
        wb.save(str(out_target))
        sess["out_target"] = out_target

    try:
        # Derive BC/ITC relations via Claude API
        relations = await asyncio.to_thread(
            derive_relations_for_products,
            selected, industry_label, _get_anthropic(), MODEL,
        )
        await asyncio.to_thread(
            add_reference_to_target_excel,
            out_target, selected, industry_label, client_name, relations,
            *_leanix_creds(sess),
        )
    except Exception as exc:
        logger.exception("Apply industry reference error")
        raise HTTPException(status_code=500, detail=str(exc))

    return {
        "ok":           True,
        "skipped":      False,
        "n_added":      len(selected),
        "download_url": f"/api/session/{session_id}/download/target",
    }


# ── Step 4 — PDF ──────────────────────────────────────────────────────────────

@app.post("/api/session/{session_id}/pdf")
async def run_pdf(
    session_id: str,
    pdf_file: Optional[UploadFile] = File(None),
):
    sess = _session(session_id)

    if not pdf_file or not pdf_file.filename:
        return {"ok": True, "skipped": True}

    out_dir     = sess["output_dir"]
    client_name = sess["client_name"]
    pdf_path    = out_dir / pdf_file.filename
    pdf_path.write_bytes(await pdf_file.read())

    try:
        result = await asyncio.to_thread(
            extract_pdf_factsheets,
            pdf_path, client_name, _get_anthropic(), MODEL,
        )
    except Exception as exc:
        logger.exception("PDF extract error")
        raise HTTPException(status_code=500, detail=str(exc))

    sess["pdf_factsheets"] = result

    return {
        "ok":                       True,
        "skipped":                  False,
        "summary":                  result.get("summary", ""),
        "n_applications":           len(result.get("applications", [])),
        "n_business_capabilities":  len(result.get("business_capabilities", [])),
        "n_organizations":          len(result.get("organizations", [])),
        "n_initiatives":            len(result.get("initiatives", [])),
        "n_it_components":          len(result.get("it_components", [])),
    }


# ── Step 5 — Images / Diagrams ────────────────────────────────────────────────

@app.post("/api/session/{session_id}/images")
async def run_images(
    session_id: str,
    images: list[UploadFile] = File(default=[]),
):
    sess = _session(session_id)

    valid = [f for f in images if f and f.filename]
    if not valid:
        return {"ok": True, "skipped": True}

    out_dir     = sess["output_dir"]
    client_name = sess["client_name"]
    saved_paths = []

    for f in valid:
        p = out_dir / f.filename
        p.write_bytes(await f.read())
        saved_paths.append(p)

    try:
        result = await asyncio.to_thread(
            extract_image_factsheets,
            saved_paths, client_name, _get_anthropic(), MODEL,
        )
    except Exception as exc:
        logger.exception("Image extract error")
        raise HTTPException(status_code=500, detail=str(exc))

    sess["image_factsheets"] = result

    return {
        "ok":                       True,
        "skipped":                  False,
        "n_files_processed":        len(saved_paths),
        "diagram_type":             result.get("diagram_type", ""),
        "summary":                  result.get("summary", ""),
        "n_applications":           len(result.get("applications", [])),
        "n_business_capabilities":  len(result.get("business_capabilities", [])),
        "n_it_components":          len(result.get("it_components", [])),
        "n_interfaces":             len(result.get("interfaces", [])),
    }


# ── Step 6 — Generate outputs ─────────────────────────────────────────────────

@app.post("/api/session/{session_id}/generate")
async def run_generate(session_id: str):
    sess        = _session(session_id)
    client_name = sess["client_name"]
    out_dir     = sess["output_dir"]
    outputs     = []

    # Baseline (already generated in Step 2, just expose download link)
    if sess.get("out_baseline") and sess["out_baseline"].exists():
        br = sess["baseline_result"] or {}
        outputs.append({
            "key":          "baseline",
            "label":        "Baseline AS-IS",
            "description":  "Applications actuales (on-premise + cloud) — importar en LeanIX",
            "filename":     sess["out_baseline"].name,
            "download_url": f"/api/session/{session_id}/download/baseline",
            "stats":        f"{br.get('n_total', 0)} apps ({br.get('n_onprem', 0)} on-prem + {br.get('n_cloud', 0)} cloud)",
        })

    # Target — requirements → LeanIX Excel
    out_target = None
    try:
        if sess.get("req_enriched_xlsx"):
            out_target = out_dir / f"{client_name}_target_leanix.xlsx"
            await asyncio.to_thread(
                write_leanix_excel_from_xlsx,
                sess["req_enriched_xlsx"], out_target, client_name,
            )
        elif sess.get("target_json_path") and sess.get("req_excel_path"):
            enriched_xlsx, out_target = await asyncio.to_thread(
                write,
                sess["target_json_path"], sess["req_excel_path"], out_dir, client_name,
            )

        if out_target and out_target.exists():
            sess["out_target"] = out_target
            import openpyxl as _xl
            wb = _xl.load_workbook(str(out_target))
            stats_parts = []
            for sh in ["Application", "BusinessCapability", "Initiative", "ITComponent"]:
                if sh in wb.sheetnames:
                    n = wb[sh].max_row - 2
                    if n > 0:
                        stats_parts.append(f"{sh[:3]}: {n}")
            outputs.append({
                "key":          "target",
                "label":        "Target TO-BE",
                "description":  "Fact sheets derivados de requerimientos — importar en LeanIX",
                "filename":     out_target.name,
                "download_url": f"/api/session/{session_id}/download/target",
                "stats":        "  |  ".join(stats_parts),
            })
    except Exception as exc:
        logger.exception("Generate target error")
        raise HTTPException(status_code=500, detail=str(exc))

    # Supplementary JSON (PDF + images)
    supplementary = {}
    if sess.get("pdf_factsheets"):
        supplementary["from_pdf"] = sess["pdf_factsheets"]
    if sess.get("image_factsheets"):
        supplementary["from_images"] = sess["image_factsheets"]

    if supplementary:
        out_supp = out_dir / f"{client_name}_supplementary_factsheets.json"
        out_supp.write_text(json.dumps(supplementary, ensure_ascii=False, indent=2))
        sess["out_supplementary"] = out_supp
        n_apps = len(supplementary.get("from_pdf", {}).get("applications", [])) + \
                 len(supplementary.get("from_images", {}).get("applications", []))
        outputs.append({
            "key":          "supplementary",
            "label":        "Fact sheets adicionales (PDF + imágenes)",
            "description":  "Revisar manualmente antes de importar en LeanIX",
            "filename":     out_supp.name,
            "download_url": f"/api/session/{session_id}/download/supplementary",
            "stats":        f"{n_apps} apps identificadas",
        })

    # Lift & Shift Target Excel (written as the main Target file)
    ls_result = sess.get("lift_shift_result")
    if ls_result and ls_result.get("conversions"):
        try:
            from pipeline.write import write_leanix_excel
            ls_convs    = ls_result["conversions"]
            ls_out_path = out_dir / f"{client_name}_target_leanix.xlsx"
            await asyncio.to_thread(
                write_leanix_excel,
                [],        # no enriched requirements — apps come from lift_shift param
                {},        # no bcs_index
                ls_out_path,
                client_name,
                lift_shift=ls_convs,
            )
            sess["out_target"]    = ls_out_path
            sess["out_liftshift"] = ls_out_path
            n_ls = len({c.get("target_matnr") for c in ls_convs})
            n_prereqs = len({p["matnr"] for c in ls_convs for p in c.get("prereq_resolved", [])})
            outputs.append({
                "key":          "target",
                "label":        "Target TO-BE (Lift & Shift)",
                "description":  f"RISE conversion targets ({ls_result.get('deployment_mode','')}) — importar en LeanIX como TO-BE",
                "filename":     ls_out_path.name,
                "download_url": f"/api/session/{session_id}/download/target",
                "stats":        f"{n_ls} targets · {n_prereqs} prerequisites",
            })
        except Exception as exc:
            logger.exception("Lift & Shift generate error")

    if not outputs:
        raise HTTPException(status_code=400, detail="No hay datos para generar. Completa al menos el Paso 2 o el Paso 3.")

    return {"ok": True, "outputs": outputs}


# ── Step 7 — Import to LeanIX ─────────────────────────────────────────────────

@app.post("/api/session/{session_id}/push")
async def run_push(session_id: str, body: dict):
    sess = _session(session_id)

    base_url, api_token = _leanix_creds(sess)
    if not api_token or not base_url:
        raise HTTPException(status_code=400, detail="No hay workspace LeanIX configurado. Selecciona uno en el Step 0.")

    client_name    = sess["client_name"]
    # Accept both naming conventions (wizard.html uses push_*, wizard_pro uses import_*)
    push_baseline  = body.get("push_baseline") or body.get("import_baseline") or False
    push_target    = body.get("push_target")   or body.get("import_target")   or False
    pushed = []
    errors = []

    if push_baseline and sess.get("out_baseline"):
        try:
            with _env_override({"LEANIX_BASE_URL": base_url, "LEANIX_API_TOKEN": api_token}):
                result = await asyncio.to_thread(push_leanix, sess["out_baseline"], client_name)
            pushed.append("baseline")
            sess["push_baseline_stats"] = result or {}
        except Exception as exc:
            errors.append(f"Baseline: {exc}")

    if push_target and sess.get("out_target"):
        try:
            ls_map = (sess.get("lift_shift_result") or {}).get("conversions") or None
            with _env_override({"LEANIX_BASE_URL": base_url, "LEANIX_API_TOKEN": api_token}):
                result = await asyncio.to_thread(push_leanix, sess["out_target"], client_name, ls_map)
            pushed.append("target")
            sess["push_target_stats"] = result or {}
        except Exception as exc:
            errors.append(f"Target: {exc}")

    catalog = {}
    for key in ("push_baseline_stats", "push_target_stats"):
        c = (sess.get(key) or {}).get("catalog", {})
        for k, v in c.items():
            catalog[k] = catalog.get(k, 0) + v

    return {
        "ok":     not errors,
        "pushed": pushed,
        "errors": errors,
        "catalog": catalog,
    }


@app.post("/api/session/{session_id}/push-ldif")
async def run_push_ldif(session_id: str, body: dict):
    """
    Alternative push via Integration API (LDIF) — single HTTP call, built-in synclog.

    Body: { push_target: bool, mode: "partial"|"full" }
    Recommended for large datasets (>100 apps) or when synclog visibility is needed.
    """
    sess = _session(session_id)

    base_url, api_token = _leanix_creds(sess)
    if not api_token or not base_url:
        raise HTTPException(status_code=400, detail="No hay workspace LeanIX configurado. Selecciona uno en el Step 0.")

    client_name  = sess["client_name"]
    push_target  = body.get("push_target", True)
    mode         = body.get("mode", "partial")
    errors: list[str] = []
    results: list[dict] = []

    if push_target and sess.get("out_target"):
        try:
            with _env_override({"LEANIX_BASE_URL": base_url, "LEANIX_API_TOKEN": api_token}):
                result = await asyncio.to_thread(
                    push_leanix_ldif, sess["out_target"], client_name, mode,
                )
            results.append({"sheet": "target", **result})
            if not result.get("ok"):
                errors.append(f"Target LDIF: {result.get('detail', 'unknown error')}")
        except Exception as exc:
            errors.append(f"Target LDIF: {exc}")

    return {
        "ok":      not errors,
        "results": results,
        "errors":  errors,
    }


@app.get("/api/session/{session_id}/transformations")
async def get_transformations(session_id: str):
    """
    Return Transformations created in LeanIX for this session's client.

    Queries all Initiatives via GraphQL, then fan-outs to the Transformations
    REST API for each one. EN/ES labels included for the frontend.
    """
    _ = session_id  # accepted for URL consistency, not required
    sess_obj  = _sessions.get(session_id, {})
    base_url, api_token = _leanix_creds(sess_obj)
    if not base_url or not api_token:
        return {"ok": True, "transformations": [], "warning": "LeanIX not configured"}

    import requests as _req
    from pipeline.leanix_auth import get_bearer

    try:
        bearer = get_bearer(base_url, api_token)
    except Exception as exc:
        return {"ok": False, "transformations": [], "error": str(exc)}

    hdrs     = {"Authorization": f"Bearer {bearer}", "Content-Type": "application/json"}
    gql_url  = f"{base_url}/services/pathfinder/v1/graphql"
    trans_url = f"{base_url}/services/transformations/v1/transformations"

    # 1. Get all Initiatives
    q = """
    query {
      allFactSheets(filter: {facetFilters: [{facetKey:"FactSheetTypes", keys:["Initiative"]}]}) {
        edges { node { id displayName } }
      }
    }
    """
    try:
        gr = await asyncio.to_thread(
            lambda: _req.post(gql_url, json={"query": q}, headers=hdrs, timeout=30)
        )
        gr.raise_for_status()
        edges = gr.json().get("data", {}).get("allFactSheets", {}).get("edges", [])
    except Exception as exc:
        return {"ok": False, "transformations": [], "error": f"GraphQL error: {exc}"}

    initiatives = [e["node"] for e in edges if e.get("node")]

    # 2. Fan-out: GET /transformations?factSheetId=<initiative_id>
    transformations: list[dict] = []
    for init in initiatives:
        try:
            tr = await asyncio.to_thread(
                lambda iid=init["id"]: _req.get(
                    trans_url, params={"factSheetId": iid}, headers=hdrs, timeout=30
                )
            )
            if tr.status_code != 200:
                continue
            items = tr.json() if isinstance(tr.json(), list) else tr.json().get("data", [])
            for item in items:
                # Resolve application name from factSheets
                app_name = ""
                fs = item.get("factSheets") or {}
                app_fs = fs.get("application") or {}
                if isinstance(app_fs, dict):
                    app_name = app_fs.get("displayName") or app_fs.get("name") or ""

                transformations.append({
                    "initiative_name": init["displayName"],
                    "app_name":        app_name,
                    "type":            item.get("type", ""),
                    "name":            item.get("name", ""),
                    "status":          item.get("execution", item.get("status", "")),
                    "completion_date": (item.get("completionDate") or {}).get("date", ""),
                })
        except Exception as exc:
            logger.warning("Transformations fetch for initiative %s failed: %s", init["id"], exc)

    return {"ok": True, "transformations": transformations}


@app.get("/api/session/{session_id}/projections")
async def get_projections(session_id: str, date: Optional[str] = None):
    """
    Return a TO-BE scenario projection using the LeanIX Impacts API.

    Queries all Initiatives, collects all Transformation IDs attached to them,
    then calls POST /services/impacts/v1/projections with those transformations
    and a target date to show the projected landscape.

    Query params:
        date (optional): ISO date YYYY-MM-DD for projection. Defaults to end of next year.

    Returns:
        { ok, date, impacts: [ { factSheetId, factSheetName, factSheetType, field, from, to } ] }
    """
    # session_id accepted for URL consistency but not required for this endpoint
    _ = session_id
    sess_obj  = _sessions.get(session_id, {})
    base_url, api_token = _leanix_creds(sess_obj)
    if not base_url or not api_token:
        return {"ok": True, "impacts": [], "warning": "LeanIX not configured"}

    import requests as _req
    from pipeline.leanix_auth import get_bearer

    try:
        bearer = get_bearer(base_url, api_token)
    except Exception as exc:
        return {"ok": False, "impacts": [], "error": str(exc)}

    hdrs      = {"Authorization": f"Bearer {bearer}", "Content-Type": "application/json"}
    gql_url   = f"{base_url}/services/pathfinder/v1/graphql"
    trans_url = f"{base_url}/services/transformations/v1/transformations"
    proj_url  = f"{base_url}/services/impacts/v1/projections"

    # Default date: last day of next calendar year
    if not date:
        import time as _time
        date = f"{int(_time.strftime('%Y')) + 1}-12-31"

    # 1. Get all Initiatives
    q = """
    query {
      allFactSheets(filter: {facetFilters: [{facetKey:"FactSheetTypes", keys:["Initiative"]}]}) {
        edges { node { id displayName } }
      }
    }
    """
    try:
        gr = await asyncio.to_thread(
            lambda: _req.post(gql_url, json={"query": q}, headers=hdrs, timeout=30)
        )
        gr.raise_for_status()
        edges = gr.json().get("data", {}).get("allFactSheets", {}).get("edges", [])
    except Exception as exc:
        return {"ok": False, "impacts": [], "error": f"GraphQL error: {exc}"}

    initiatives = [e["node"] for e in edges if e.get("node")]
    if not initiatives:
        return {"ok": True, "date": date, "impacts": [], "note": "No initiatives found"}

    # 2. Fan-out: collect all transformation IDs
    transformation_ids: list[str] = []
    fact_sheet_ids: list[str] = []
    for init in initiatives:
        try:
            tr = await asyncio.to_thread(
                lambda iid=init["id"]: _req.get(
                    trans_url, params={"factSheetId": iid}, headers=hdrs, timeout=30
                )
            )
            if tr.status_code != 200:
                continue
            items = tr.json() if isinstance(tr.json(), list) else tr.json().get("data", [])
            for item in items:
                tid = item.get("id")
                if tid:
                    transformation_ids.append(tid)
                # Collect affected application IDs
                fs = item.get("factSheets") or {}
                app_fs = fs.get("application") or {}
                if isinstance(app_fs, dict) and app_fs.get("id"):
                    fact_sheet_ids.append(app_fs["id"])
        except Exception as exc:
            logger.warning("Projections: transformations fetch for %s failed: %s", init["id"], exc)

    if not transformation_ids:
        return {"ok": True, "date": date, "impacts": [], "note": "No transformations found"}

    # Deduplicate
    transformation_ids = list(dict.fromkeys(transformation_ids))
    fact_sheet_ids     = list(dict.fromkeys(fact_sheet_ids))

    # 3. Call Impacts API projections
    proj_body = {
        "factSheetIds":     fact_sheet_ids[:100],  # cap to avoid huge requests
        "date":             date,
        "transformationIds": transformation_ids[:50],
    }
    try:
        pr = await asyncio.to_thread(
            lambda: _req.post(proj_url, json=proj_body, headers=hdrs, timeout=60)
        )
        if pr.status_code in (403, 404):
            return {"ok": True, "date": date, "impacts": [], "note": f"Impacts API not available ({pr.status_code})"}
        pr.raise_for_status()
        proj_data = pr.json()
    except Exception as exc:
        return {"ok": False, "impacts": [], "error": f"Impacts API error: {exc}"}

    # Normalize response — API returns list of impact objects
    raw_impacts = proj_data if isinstance(proj_data, list) else proj_data.get("data", proj_data.get("impacts", []))
    impacts: list[dict] = []
    for impact in raw_impacts:
        impacts.append({
            "factSheetId":   impact.get("factSheetId", ""),
            "factSheetName": impact.get("factSheetName", impact.get("displayName", "")),
            "factSheetType": impact.get("factSheetType", "Application"),
            "field":         impact.get("fieldName", impact.get("field", "")),
            "from":          impact.get("currentValue", impact.get("from", "")),
            "to":            impact.get("projectedValue", impact.get("to", "")),
        })

    return {
        "ok":                 True,
        "date":               date,
        "transformation_count": len(transformation_ids),
        "impacts":            impacts,
    }
async def get_synclog(session_id: str):
    """
    Return recent synchronization runs from LeanIX Integration Hub (synclog).

    Calls GET /services/synclog/v1/synchronizations.
    Useful for debugging push failures — shows status, timestamps and warnings
    for the most recent integrations in this workspace.
    """
    _ = session_id  # accepted for URL consistency, no session needed
    sess_obj  = _sessions.get(session_id, {})
    base_url, api_token = _leanix_creds(sess_obj)
    if not base_url or not api_token:
        return {"ok": True, "synchronizations": [], "warning": "LeanIX not configured"}

    import requests as _req
    from pipeline.leanix_auth import get_bearer

    try:
        bearer = get_bearer(base_url, api_token)
    except Exception as exc:
        return {"ok": False, "synchronizations": [], "error": str(exc)}

    hdrs     = {"Authorization": f"Bearer {bearer}"}
    sync_url = f"{base_url}/services/synclog/v1/synchronizations"

    try:
        resp = await asyncio.to_thread(
            lambda: _req.get(sync_url, params={"pageSize": 20}, headers=hdrs, timeout=30)
        )
        if resp.status_code in (403, 404):
            return {"ok": True, "synchronizations": [], "warning": "Synclog not available (insufficient permissions or endpoint not found)"}
        resp.raise_for_status()
        raw = resp.json()
    except Exception as exc:
        return {"ok": False, "synchronizations": [], "error": str(exc)}

    # Normalise: the endpoint may return a list or {data: [...]}
    items = raw if isinstance(raw, list) else raw.get("data", raw.get("synchronizations", []))

    syncs = []
    for item in items[:20]:
        syncs.append({
            "id":           item.get("id", ""),
            "connector":    item.get("connectorType") or item.get("connector") or "",
            "status":       item.get("status", ""),
            "started_at":   item.get("startedAt") or item.get("createdAt") or "",
            "finished_at":  item.get("finishedAt") or item.get("updatedAt") or "",
            "fact_sheets_created":  item.get("factSheetsCreated") or item.get("created") or 0,
            "fact_sheets_updated":  item.get("factSheetsUpdated") or item.get("updated") or 0,
            "warnings":     len(item.get("warnings") or []),
            "errors":       len(item.get("errors")   or []),
        })

    return {"ok": True, "synchronizations": syncs}


# ── Step 9 Easter Egg — KPI Achievement Import ────────────────────────────────

@app.post("/api/session/{session_id}/push-kpi")
async def run_push_kpi(session_id: str):
    """🍪 Easter egg: generate KPI Achievement Excel and import it into LeanIX."""
    sess = _session(session_id)

    base_url, api_token = _leanix_creds(sess)
    if not api_token or not base_url:
        raise HTTPException(status_code=400, detail="No hay workspace LeanIX configurado. Selecciona uno en el Step 0.")

    client_name = sess["client_name"]
    out_dir     = sess["output_dir"]
    kpi_path    = out_dir / "kpi_achievement_leanix.xlsx"

    # ── Generate the KPI Excel into the session output dir ──────────────────
    def _build_kpi():
        import sys
        sys.path.insert(0, str(BASE_DIR))
        import importlib, types

        # Import generate_kpi_excel and override OUTPUT_PATH to session dir
        import generate_kpi_excel as _kpi_mod
        original_path = _kpi_mod.OUTPUT_PATH
        _kpi_mod.OUTPUT_PATH = kpi_path
        try:
            _kpi_mod.build()
        finally:
            _kpi_mod.OUTPUT_PATH = original_path

    await asyncio.to_thread(_build_kpi)

    if not kpi_path.exists():
        raise HTTPException(status_code=500, detail="Error generando el fichero KPI Achievement Excel.")

    # ── Push to LeanIX in order: Objective → BusinessCapability → Application → Initiative ──
    try:
        with _env_override({"LEANIX_BASE_URL": base_url, "LEANIX_API_TOKEN": api_token}):
            await asyncio.to_thread(push_leanix, kpi_path, client_name)
    except Exception as exc:
        return {"ok": False, "errors": [str(exc)], "path": str(kpi_path)}

    sess["out_kpi"] = kpi_path
    return {
        "ok":       True,
        "pushed":   ["kpi_achievement"],
        "path":     str(kpi_path),
        "stats": {
            "objectives":           5,
            "business_capabilities": 91,
            "applications":          100,
            "initiatives":           5,
        },
    }


# ── Downloads ─────────────────────────────────────────────────────────────────

_DOWNLOAD_KEYS = {
    "baseline":      "out_baseline",
    "target":        "out_target",
    "supplementary": "out_supplementary",
    "contrast":      "out_contrast",
    "liftshift":     "out_liftshift",
}

@app.get("/api/session/{session_id}/download/{key}")
async def download_file(session_id: str, key: str):
    sess = _session(session_id)

    if key not in _DOWNLOAD_KEYS:
        raise HTTPException(status_code=404, detail=f"Unknown download key: {key!r}")

    path: Optional[Path] = sess.get(_DOWNLOAD_KEYS[key])
    if not path or not path.exists():
        raise HTTPException(status_code=404, detail=f"File not yet generated for key: {key!r}")

    media = "application/json" if path.suffix == ".json" else \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(path=str(path), filename=path.name, media_type=media)


@app.get("/api/session/{session_id}/catalog-report", response_class=HTMLResponse)
async def catalog_report_html(session_id: str):
    sess = _session(session_id)  # raises 404 if missing
    session_dir = OUTPUT_DIR / session_id
    uuid_map_path = session_dir / "push_uuid_map.json"
    if not uuid_map_path.exists():
        raise HTTPException(status_code=404, detail="push_uuid_map.json not found — run Step 7 first")

    html_path = session_dir / "catalog_report.html"
    if not html_path.exists():
        # Render on demand
        from pipeline.catalog_report import generate_report
        try:
            generate_report(session_dir=session_dir, client_name=sess.get("client_name") or "")
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc))
    return HTMLResponse(content=html_path.read_text())


@app.get("/api/session/{session_id}/catalog-report.xlsx")
async def catalog_report_xlsx(session_id: str):
    sess = _session(session_id)
    session_dir = OUTPUT_DIR / session_id
    uuid_map_path = session_dir / "push_uuid_map.json"
    if not uuid_map_path.exists():
        raise HTTPException(status_code=404, detail="push_uuid_map.json not found — run Step 7 first")

    xlsx_path = session_dir / "catalog_report.xlsx"
    if not xlsx_path.exists():
        from pipeline.catalog_report import generate_report
        try:
            generate_report(session_dir=session_dir, client_name=sess.get("client_name") or "")
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc))
    return FileResponse(
        path=str(xlsx_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="catalog_report.xlsx",
    )


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("""
╔══════════════════════════════════════════════════════╗
║        Archimedes AI — Wizard Server                 ║
║  URL: http://localhost:8767                          ║
║  Press Ctrl+C to stop                               ║
╚══════════════════════════════════════════════════════╝
""")
    import threading
    threading.Timer(1.2, lambda: webbrowser.open("http://localhost:8767")).start()
    uvicorn.run(app, host="127.0.0.1", port=8767, log_level="warning")
